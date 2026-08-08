import os
import sys
import socket
import threading
import tkinter as tk
from tkinter import messagebox, ttk
import pandas as pd
import warnings

import shutil
import openpyxl
from openpyxl import load_workbook
#import module function
from module import globalconfig
from module.utils import inhopdongnhancong
from module.utils import inhopdong_ca_nhan
from module.utils import in_phu_luc_to_doi
from module.utils import in_hop_dong_giao_khoan_to_truong
from module.utils import get_DSNC_path
from module.utils import convert_number_to_vietnamese
from module.utils import nhap_luong_dialog
from module.utils import quan_ly_danh_muc_cong_viec
from module.utils import phu_luc_cong_viec_hop_dong
from module.utils import addcontract
from module.utils import hien_thi_danh_sach_hop_dong
from module.license_utils import check_license_and_launch
from module.globalconfig import khoitaothongso
from pathlib import Path
import unicodedata
import customtkinter as ctk
from PIL import Image
# Bỏ qua cảnh báo từ openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
def normalize_text(text):
    return unicodedata.normalize('NFC', str(text))
def get_project_root():
    # 1. Xác định vị trí file thực thi hoặc file script hiện tại
    if getattr(sys, 'frozen', False):
        # Đang chạy .exe: Root chính là thư mục chứa .exe
        return Path(sys.executable).parent
    else:
        # Đang chạy .py: Bắt đầu từ vị trí file hiện tại
        current_path = Path(__file__).resolve().parent.parent          
        return current_path

def get_path(relative_path):
    """Hàm tiện ích gom chung cho mọi loại đường dẫn"""
    print(f"root: {get_project_root()}")
    base_dir = get_project_root()
    
    path_at_root = os.path.join(base_dir, relative_path)
    path_at_internal = os.path.join(base_dir, "_internal", relative_path)
    if os.path.exists(path_at_root):
        return path_at_root
    # Nếu không, kiểm tra xem nó có bị đẩy vào _internal không
    if os.path.exists(path_at_internal):
        return path_at_internal    
        
    return path_at_root 

def clean_cell(val):
    """Làm sạch giá trị ô để hiển thị text thuần túy"""
    if pd.isna(val): 
        return ""
    # Chuyển về string, xóa khoảng trắng thừa
    s = str(val).strip()
    
    # Nếu là dạng số float kết thúc bằng .0, loại bỏ .0
    if s.endswith('.0'):
        s = s[:-2]
        
    return s

# Hàm mở file Excel
def open_excel_file():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    dsnc_path = get_DSNC_path()#os.path.join(parent_dir, 'Data', 'dulieuduan.xlsx')
    
    if os.path.exists(dsnc_path):
        os.startfile(dsnc_path)  # Lệnh mở file mặc định trên Windows
    else:
        messagebox.showerror("Lỗi", "Không tìm thấy file để mở!")
        
def load_names():
    # 1. Cổng kiểm soát license
    is_valid = check_license_and_launch(root, status_label)
    if not is_valid:
        print("Bản quyền không hợp lệ. Đang dừng chương trình...")
        return

    # 2. Kiểm tra xem Tree đã được tạo chưa
    # Nếu chưa tồn tại (None) hoặc chưa được khởi tạo, ta mới tạo UI
    if not hasattr(globalconfig, 'tree_todoi') or globalconfig.tree_todoi is None:
        table_frame = tk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        columns = ('HoTen', 'MST', 'CCCD')
        tree = ttk.Treeview(table_frame, columns=columns, show='tree headings')
        
        # Cấu hình UI (chỉ chạy 1 lần)
        #tree.bind("<Button-3>", show_context_menu)
        #tree.bind("<Button-3>", show_custom_menu)
        tree.bind("<Button-3>", lambda event: show_custom_menu(event, root))
        tree.heading("#0", text="Tổ đội / Tổ trưởng", anchor="w")
        tree.column("#0", width=250, anchor="w")
        tree.heading('HoTen', text='Họ và tên')
        tree.heading('MST', text='MST')
        tree.heading('CCCD', text='CMT/CCCD/Hộ chiếu')
        tree.column('HoTen', width=150)
        tree.column('MST', width=100)
        tree.column('CCCD', width=150)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Lưu vào globalconfig
        globalconfig.tree_todoi = tree
    else:
        # Nếu đã tồn tại, lấy ra và xóa sạch dữ liệu cũ để nạp mới
        tree = globalconfig.tree_todoi
        for item in tree.get_children():
            tree.delete(item)

    # 3. Nạp dữ liệu (Phần này chung cho cả 2 trường hợp)
    try:
        nctnpath = globalconfig.folder_nhancongthuengoai
        dsnc_path = get_path(os.path.join(nctnpath,"dulieuduan.xlsx"))
        print (f"dsnc_path {dsnc_path}  nctnpath: {nctnpath} ")
        df = pd.read_excel(
            dsnc_path, 
            sheet_name='NCList', 
            engine='openpyxl',
            dtype={'MST': str, 'CMT/CCCD/Hộ chiếu': str}
        )
        df.columns = df.columns.str.strip()
        df['Tổ đội'] = df['Tổ đội'].apply(lambda x: unicodedata.normalize('NFC', str(x)))
        grouped = df.groupby('Tổ đội')
        
        for to_doi, group in grouped:
            group['Công việc'] = group['Công việc'].apply(normalize_text)
            quan_ly_df = group[group['Công việc'].str.strip().str.lower() == 'tổ trưởng']
            
            to_truong_name = quan_ly_df.iloc[0]['Họ và tên'] if not quan_ly_df.empty else "Chưa xác định"
            
            node_cha = tree.insert("", "end", text=f"{to_doi} - Tổ trưởng: {to_truong_name}", open=True)
            
            for _, row in group.iterrows():
                tree.insert(node_cha, "end", values=(
                    clean_cell(row['Họ và tên']), 
                    clean_cell(row['MST']), 
                    clean_cell(row['CMT/CCCD/Hộ chiếu'])
                ))
        globalconfig.root = root
        #root.withdraw()
    except Exception as e:
        print(f"Lỗi khi load dữ liệu: {e}")      
                          
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể tải danh sách: {e}")

def get_selection():
    selected_item = tree.selection()
    if selected_item:
        item_data = tree.item(selected_item)['values']
        messagebox.showinfo("Đã chọn", f"Bạn đã chọn: {item_data[0]} (MST: {item_data[1]})")
    else:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn một hàng trong bảng!")
def print_selected_team():
    #print (f"do the print_selected_team")
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy text của item được chọn (tên Tổ đội)
    team_name = tree.item(selected_item, 'text')
    
    if messagebox.askyesno("Xác nhận", f"Bạn muốn in toàn bộ hợp đồng cho {team_name}?"):
        inhopdongnhancong(selected_team=team_name)
        messagebox.showinfo("Hoàn tất", "Đã in xong Cam kết cho tổ đội!")
def print_selected_person():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    
    # Kiểm tra xem người dùng có chọn đúng nhân viên không
    if not tree.parent(selected_item):
        messagebox.showwarning("Thông báo", "Vui lòng chọn một nhân viên cụ thể!")
        return
        
    values = tree.item(selected_item, 'values') # (HoTen, MST, CCCD)
    hoten, mst, cccd = values[0], values[1], values[2]
    
    if messagebox.askyesno("Xác nhận", f"In cam kết cho {hoten} (CCCD: {cccd})?"):
        # Truyền 3 tham số định danh, không cần team_name
        inhopdong_ca_nhan(hoten, mst, cccd)

def print_team_appendix():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    
    if messagebox.askyesno("Xác nhận", f"In phụ lục danh sách cho: {team_name}?"):
        in_phu_luc_to_doi(team_name)
        #messagebox.showinfo("Thành công", "Đã xuất phụ lục danh sách tổ!")

def danh_muc_cong_viec():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    team_name = unicodedata.normalize('NFC', team_name)    
    quan_ly_danh_muc_cong_viec(team_name)
def add_hopdong():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
    os.makedirs(folder_path, exist_ok=True)
    catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    if not os.path.exists(catalog_path):
        messagebox.showerror("Lỗi", "Chưa có file danh mục công việc!")
        return
    addcontract(team_name)
    #globalconfig.root = root
    root.withdraw()
def hien_thi_ds_hopdong():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    hien_thi_danh_sach_hop_dong(team_name)    
def print_team_contract():
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    
    if messagebox.askyesno("Xác nhận", f"In phụ lục danh sách cho: {team_name}?"):
        in_hop_dong_giao_khoan_to_truong(team_name)

def is_valid_date(date_val):
    # 1. Kiểm tra nếu giá trị là NaN, None hoặc rỗng
    if pd.isna(date_val) or date_val == "":
        return False        
    # 2. Chuyển sang chuỗi và xóa khoảng trắng thừa
    date_str = str(date_val).strip()    
    # 3. Nếu sau khi chuyển thành chuỗi mà nó vẫn là "nan" (do lỗi ép kiểu)
    if date_str.lower() == 'nan' or date_str.lower() == 'nat':
        return False
    try:
        pd.to_datetime(date_str, format='%d/%m/%Y', errors='raise')
        return True
    except (ValueError, TypeError):
        return False
def is_number(s):
    try:
        float(s) # Thử chuyển sang số thực (bao gồm cả số nguyên)
        return True
    except ValueError:
        return False
        
def open_contract_input_dialog(team_name, callback):
    # Đường dẫn file
    dsnc_path = get_DSNC_path()
    # 1. Đọc sheet NCList vào DataFrame
    df_nc = pd.read_excel(dsnc_path, sheet_name='NCList', engine='openpyxl')
    
    # 2. Lọc dữ liệu theo team_name
    # Tìm dòng mà cột 'Tổ đội' bằng team_name
    team_data = df_nc[df_nc['Tổ đội'] == team_name]
    manager_data = team_data[team_data['Công việc'] == 'Tổ trưởng']
    # Đọc cấu hình đơn vị từ sheet Validate
    try:
        df_val = pd.read_excel(dsnc_path, sheet_name='Validate', engine='openpyxl')
        units = df_val['Đơn vị thời gian hợp đồng'].dropna().tolist()
    except:
        units = ['Ngày', 'Tháng', 'Năm', 'Giờ'] # Fallback nếu lỗi
        
    if not team_data.empty:
        # Lấy giá trị của dòng đầu tiên tìm thấy (nếu tổ đội có nhiều dòng, nó lấy dòng đầu tiên)
        row = manager_data.iloc[0]
        # Lấy giá trị từ các cột tương ứng
        
        default_start = str(row['Ngày bắt đầu hợp đồng']) if 'Ngày bắt đầu hợp đồng' in row.index else pd.Timestamp.now().strftime('%d/%m/%Y')
        if not is_valid_date(default_start):
            default_start = pd.Timestamp.now().strftime('%d/%m/%Y')
            print(f"default_start is not valid")
            
        try:
            default_duration = str(int(row['Thời gian hợp đồng'])) if 'Thời gian hợp đồng' in row.index else ''
        except:
           default_duration = ''
           
        default_unit = str(row['Đơn vị thời gian hợp đồng']) if 'Đơn vị thời gian hợp đồng' in row.index else units[0]
        try:
            default_price = str(int(row['Giá trị khoán'])) if 'Giá trị khoán' in row.index else ''
        except:
            default_price = ''  
        try:
            so_hop_dong = str(row['Số hợp đồng']) if 'Số hợp đồng' in row.index else ''
            print(f"Số hợp đồng đọc được : {so_hop_dong}")
        except:
            print(f"Không tìm thấy số hợp đồng trong  {row['Tổ đội']}")
            so_hop_dong = ''  
            
    else:
        # Nếu không tìm thấy tổ đội, dùng giá trị mặc định hoặc rỗng
        default_start = pd.Timestamp.now().strftime('%d/%m/%Y')
        default_duration = ''
        default_unit = units[0]
        default_price = ''
        so_hop_dong = ''
    # Tạo cửa sổ phụ
    dialog = tk.Toplevel()
    dialog.title(f"Nhập thông tin hợp đồng: {team_name}")
    dialog.geometry("350x350")

    tk.Label(dialog, text="Ngày bắt đầu (DD/MM/YYYY):").pack(pady=5)
    ent_start = tk.Entry(dialog)
    ent_start.pack()
    #ent_start.insert(0, pd.Timestamp.now().strftime('%d/%m/%Y'))
    ent_start.insert(0, default_start)
    
    tk.Label(dialog, text="Thời gian hợp đồng:").pack(pady=5)
    ent_duration = tk.Entry(dialog)
    ent_duration.pack()
    ent_duration.insert(0, default_duration)

    tk.Label(dialog, text="Đơn vị thời gian:").pack(pady=5)
    combo_unit = ttk.Combobox(dialog, values=units)
    combo_unit.pack()
    if default_unit in units:
        combo_unit.set(default_unit)
    else:
        # Nếu giá trị trong file Excel không nằm trong danh sách đơn vị (ví dụ bị sai dữ liệu), 
        # ta quay về chọn mặc định đầu tiên
        combo_unit.current(0)

    tk.Label(dialog, text="Tiền khoán:").pack(pady=5)
    ent_price = tk.Entry(dialog)
    ent_price.pack()
    # Load dữ liệu cũ nếu có
    ent_price.insert(0, default_price)
    # 1. Thêm Label để hiển thị chữ
    lbl_amount_text = tk.Label(dialog, text="", fg="blue", font=("Arial", 9, "italic"))
    lbl_amount_text.pack(pady=5)
    
    tk.Label(dialog, text="Số hợp đồng:").pack(pady=5)
    ent_sohopdong = tk.Entry(dialog)
    ent_sohopdong.pack()
    # Load dữ liệu cũ nếu có
    ent_sohopdong.insert(0, so_hop_dong)
    #2. Hàm cập nhật chữ khi gõ
    def update_amount_text(event=None):
        raw_val = ent_price.get()
        # Loại bỏ các ký tự không phải số
        clean_val = ''.join(filter(str.isdigit, raw_val))
        if clean_val:
            # Gọi hàm chuyển đổi của bạn
            chu_so = convert_number_to_vietnamese(int(clean_val))
            lbl_amount_text.config(text=f"Bằng chữ: {chu_so}")
        else:
            lbl_amount_text.config(text="")
    # 3. Gán sự kiện 'KeyRelease' (ngay khi gõ xong 1 ký tự)
    ent_price.bind('<KeyRelease>', update_amount_text)
    def on_submit():
        start_date_str = ent_start.get()
        # 1. Kiểm tra tính hợp lệ của ngày tháng
        # Sử dụng hàm is_valid_date mà chúng ta đã xây dựng trước đó
        if not is_valid_date(start_date_str):
            messagebox.showerror("Lỗi định dạng", "Ngày bắt đầu không hợp lệ!\nVui lòng nhập theo định dạng DD/MM/YYYY")
            # 2. Bôi đen toàn bộ nội dung trong ent_start
            ent_start.select_range(0, 'end')
            ent_start.focus_set() # Đưa con trỏ chuột vào lại ô đó
            return # Dừng lại, không thực hiện các bước tiếp theo
        
        
        if not is_number(ent_duration.get()):
            messagebox.showerror("Lỗi định dạng", "Thời gian hợp đồng không hợp lệ!\nVui lòng nhập thời gian hợp đồng là con số")
            ent_duration.select_range(0, 'end')
            ent_duration.focus_set() # Đưa con trỏ chuột vào lại ô đó
            return # Dừng lại, không thực hiện các bước tiếp theo
        if not is_number(ent_price.get()):
            messagebox.showerror("Lỗi định dạng", "Giá trị hợp đồng không hợp lệ!\nVui lòng nhập giá trị hợp đồng là con số")
            ent_price.select_range(0, 'end')
            ent_price.focus_set() # Đưa con trỏ chuột vào lại ô đó
            return # Dừng lại, không thực hiện các bước tiếp theo    
        if combo_unit.get() not in units:
            messagebox.showerror("Lỗi định dạng", "Giá trị đơn vị thời gian hợp đồng không hợp lệ!\nVui lòng chọn một giá trị trong list")
            combo_unit.current(0)
            combo_unit.select_range(0, 'end')
            combo_unit.focus_set()
            return # Dừng lại, không thực hiện các bước tiếp theo
        data = {
            'start': ent_start.get(),
            'duration': ent_duration.get(),
            'unit': combo_unit.get(),
            'price': ent_price.get(), # Lấy giá trị tiền khoán
            'so_hop_dong': ent_sohopdong.get()
        }
        dialog.destroy()
        callback(team_name, data)

    tk.Button(dialog, text="Xác nhận & In", command=on_submit).pack(pady=20)
def process_contract_printing(team_name, data):
    dsnc_path = get_DSNC_path()
    # 1. Mở workbook
    wb = openpyxl.load_workbook(dsnc_path)
    ws = wb['NCList']
    
    # 2. Tìm hoặc tạo cột 'Giá trị khoán'
    header = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    
    if 'Giá trị khoán' not in header:
        new_col_idx = ws.max_column + 1
        ws.cell(row=1, column=new_col_idx, value='Giá trị khoán')
        header['Giá trị khoán'] = new_col_idx
    if 'Số hợp đồng' not in header:
        new_col_idx = ws.max_column + 1
        ws.cell(row=1, column=new_col_idx, value='Số hợp đồng')
        header['Số hợp đồng'] = new_col_idx
    # Xác định các cột cần thiết
    col_to_doi = header.get('Tổ đội')
    col_start = header.get('Ngày bắt đầu hợp đồng')
    col_duration = header.get('Thời gian hợp đồng')
    col_unit = header.get('Đơn vị thời gian hợp đồng')
    col_gia_tri = header.get('Giá trị khoán')
    col_cong_viec = header.get('Công việc') # Giả sử cột này tên là 'Công việc'
    col_so_hop_dong = header.get('Số hợp đồng')
    if not all([col_to_doi, col_start, col_duration, col_unit, col_gia_tri]):
        messagebox.showerror("Lỗi", "Không tìm thấy đủ các cột cần thiết trong file Excel!")
        return

    # 3. Duyệt qua các dòng để tìm và cập nhật
    updated = False
    for row in range(2, ws.max_row + 1):
        team_val = ws.cell(row=row, column=col_to_doi).value
        
        if team_val == team_name:
            # Cập nhật thông tin chung cho các dòng của tổ đội
            ws.cell(row=row, column=col_start).value = data['start']
            ws.cell(row=row, column=col_duration).value = data['duration']
            ws.cell(row=row, column=col_unit).value = data['unit']
            #ws.cell(row=row, column=col_so_hop_dong).value = data['so_hop_dong']
            # Cập nhật giá trị khoán riêng cho người có Công việc = 'Tổ trưởng'
            # Kiểm tra nếu cột Công việc tồn tại
            if col_cong_viec:
                cong_viec = ws.cell(row=row, column=col_cong_viec).value
                if cong_viec == 'Tổ trưởng':
                    ws.cell(row=row, column=col_gia_tri).value = data.get('price', 0)
                    ws.cell(row=row, column=col_so_hop_dong).value = data.get('so_hop_dong', 0)
            updated = True
    
    # 4. Lưu workbook
    if updated:
        wb.save(dsnc_path)
        print(f"Đã cập nhật dữ liệu cho tổ: {team_name}")
        in_hop_dong_giao_khoan_to_truong(team_name)
        messagebox.showinfo("Hoàn tất", f"Đã cập nhật dữ liệu và in HĐ cho {team_name}")
    else:
        messagebox.showwarning("Thông báo", f"Không tìm thấy tổ đội: {team_name}")
def get_safe_team_name():
    selected_item = tree.selection()
    if not selected_item:
        return None
    
    item_values = tree.item(selected_item[0])['values']
    #team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    # Kiểm tra xem values có tồn tại và không rỗng
    if item_values and len(item_values) > 0:       
        return item_values[0].split(" - ")[0].split(" (TT:")[0].strip()
    else:
        # Trường hợp Node cha không có giá trị ở cột 0, 
        # ta lấy text của chính nó (thường là tên tổ đội)
        return tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
# Hàm kích hoạt menu khi chuột phải

def show_custom_menu(event, root_window):
    # 1. Tạo menu
    menu = ctk.CTkToplevel()
    menu.overrideredirect(True)
    menu.attributes("-topmost", True)
    menu.geometry(f"+{event.x_root}+{event.y_root}")
    menu.focus_force() 

    # Hàm đóng menu an toàn
    def close_menu():
        if menu.winfo_exists():
            menu.destroy()

    # 2. Xử lý FocusOut với độ trễ (Debounce)
    # Nếu click ra ngoài, đợi 150ms rồi mới đóng. 
    # Khoảng thời gian 150ms này đủ để nút bấm thực thi nếu người dùng click vào menu.
    def on_focus_out(e):
        menu.after(80, close_menu)

    menu.bind("<FocusOut>", on_focus_out)
    
    # Khung chứa menu
    frame = ctk.CTkFrame(menu, corner_radius=10, fg_color="#FFFFFF", border_width=1, border_color="#D1D1D1")
    frame.pack(padx=2, pady=2)

    # Hàm tạo mục menu chuẩn
    def add_menu_item(text, icon_path, command_func):
        img = None
        if icon_path and os.path.exists(icon_path):
            img = ctk.CTkImage(light_image=Image.open(icon_path), size=(20, 20))
        
        def on_click():
            # ƯU TIÊN SỐ 1: Đóng menu ngay lập tức khi click nút (không cần đợi)
            close_menu()
            
            # ƯU TIÊN SỐ 2: Chạy hàm xử lý sau 50ms để giao diện ổn định
            root_window.after(50, lambda: execute_task(command_func))

        def execute_task(func):
            try:
                func()
            except Exception as e:
                print(f"Lỗi thực thi: {e}")

        btn = ctk.CTkButton(
            frame, text=text, image=img, compound="left", anchor="w",
            fg_color="transparent", text_color="#333333", hover_color="#E8E8E8",
            height=35, width=200, command=on_click
        )
        btn.pack(fill="x", padx=5, pady=2)

    # 3. Logic xác định item trong Treeview
    tree = globalconfig.tree_todoi
    item = tree.identify_row(event.y)
    
    if item:
        tree.selection_set(item)
        is_team_node = (tree.parent(item) == "")
        
        if is_team_node:
            add_menu_item("In cam kết cho tổ", get_path("icons/print.png"), print_selected_team)
            add_menu_item("In phụ lục danh sách tổ", get_path("icons/list.png"), print_team_appendix)
            add_menu_item("Danh mục công việc", get_path("icons/edit.png"), danh_muc_cong_viec)
            add_menu_item("Tạo hợp đồng khoán", get_path("icons/edit.png"), add_hopdong)
            add_menu_item("Danh mục hợp đồng", get_path("icons/edit.png"), hien_thi_ds_hopdong)
        else:
            add_menu_item("In Cam kết cho người này", get_path("icons/edit.png"), print_selected_person)
    else:
        menu.destroy()
    
def push_to_server_function():
    return
    
def get_from_server_function():
    return
    
def check_internet():
    """Kiểm tra kết nối internet bằng cách thử kết nối tới Google DNS."""
    try:
        socket.setdefaulttimeout(3)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(("8.8.8.8", 53))
        return True
    except Exception:
        return False

def update_status():
    """Cập nhật màu sắc của bong bóng 3D dựa trên trạng thái mạng."""
    is_online = check_internet()
    color = "green" if is_online else "red"
    text = "Connected" if is_online else "No internet connection"
    
    # Cập nhật màu bong bóng
    canvas_status.itemconfig(bubble, fill=color)
    lbl_status_text.config(text=text)
    
    # Tự động kiểm tra lại sau mỗi 10 giây
    root.after(10000, update_status)
def open_output():
    # 1. Xác định đường dẫn thư mục Output
    # Thường nằm cùng cấp với file .exe hoặc file script đang chạy
    output_dir = os.path.join(os.getcwd(), "Output")

    # 2. Kiểm tra nếu chưa tồn tại thì tạo mới để tránh lỗi
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo thư mục Output: {e}")
            return

    # 3. Mở thư mục bằng File Explorer của Windows
    try:
        os.startfile(output_dir)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể mở thư mục: {e}")

# --- THIẾT LẬP GIAO DIỆN ---                
#root = tk.Tk()
root = ctk.CTk()
#create sidebar menu
# 1. Khởi tạo Sidebar
sidebar = ctk.CTkFrame(master=root, width=200, corner_radius=0)
sidebar.pack(side="left", fill="y")

# 2. Logic đóng mở menu con
def toggle_menu_con():
    if menu_con_frame.winfo_ismapped():
        menu_con_frame.pack_forget()
    else:
        # Khi hiện ra, đặt nó nằm ngay dưới nút cha
        menu_con_frame.pack(fill="x", pady=0) 

# 3. Nút Menu Cha (đặt master=sidebar)
btn_cha = ctk.CTkButton(
    master=sidebar, 
    text="📂 Danh mục quản lý ▼", 
    fg_color="#3B8ED0",
    hover_color="#275E8D",
    anchor="w",         # Căn lề trái cho chữ
    command=toggle_menu_con
)
btn_cha.pack(fill="x", pady=10, padx=10)

# 4. Khung chứa các Menu Con (đặt master=sidebar)
# Để "transparent" để nó hòa vào màu nền của sidebar
menu_con_frame = ctk.CTkFrame(master=sidebar, fg_color="transparent")
def create_menu_button(text, command):
    return ctk.CTkButton(
        menu_con_frame,
        text=text,
        fg_color="transparent", # Trong suốt để hòa vào nền sidebar
        text_color="black",
        hover_color="#3B8ED0",
        anchor="w",             # Căn trái chữ
        height=40,
        command=command
    )    
# Thêm các mục vào menu
create_menu_button("📄Danh mục công việc", lambda: print("Chọn công việc")).pack(fill="x", padx=10, pady=5)
create_menu_button("📋 Hợp đồng khoán", lambda: print("Chọn hợp đồng")).pack(fill="x", padx=10, pady=5)
create_menu_button("⚙️ Cấu hình", lambda: print("Chọn cấu hình")).pack(fill="x", padx=10, pady=5)

# 1. Cấu hình Treeview (Bảng)
style = ttk.Style()
style.theme_use("clam")

# 1. Cấu hình bảng dữ liệu
style = ttk.Style()
style.theme_use("clam")

# Cấu hình bảng sáng - tạo không gian mở
style.configure("Treeview", 
                background="#FFFFFF",       # Nền trắng tinh khôi
                foreground="#333333",       # Chữ đen xám (dịu hơn đen tuyền)
                fieldbackground="#FFFFFF",
                rowheight=35,               # Giãn cách dòng thoáng đãng
                borderwidth=0)

# Cấu hình Header - màu xanh xám nhạt (Industrial style)
style.configure("Treeview.Heading", 
                background="#E1E4E8",       # Màu xám xanh nhạt
                foreground="#2C3E50",       # Chữ đậm nổi bật
                relief="flat",
                font=('Segoe UI', 12, 'bold'))

# Cấu hình khi chọn hàng - Màu xanh dương dịu
style.map('Treeview', 
          background=[('selected', '#5C91D4')], 
          foreground=[('selected', '#FFFFFF')])
                
                
root.title("Smart AC Management")
root.geometry("600x500")
#context_menu
#context_menu = tk.Menu(root, tearoff=0)
# --- 1. TẠO THANH MENU ---
menubar = tk.Menu(root)
root.config(menu=menubar)

# Menu File
file_menu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="Tải danh sách tổ đội", command=load_names)
file_menu.add_command(label="Mở file dữ liệu dự án để soạn thảo", command=open_excel_file)
file_menu.add_command(label="Mở thư mục in ấn", command=open_output)
#file_menu.add_command(label="Sao lưu dữ liệu dự án", command=open_excel_file)
file_menu.add_separator()
file_menu.add_command(label="Thoát", command=root.quit)

# Menu Server
server_menu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="Server", menu=server_menu)
server_menu.add_command(label="Push to server", command=push_to_server_function) # Thay bằng tên hàm của bạn
server_menu.add_command(label="Get from server", command=get_from_server_function) # Thay bằng tên hàm của bạn

# --- PHẦN GIAO DIỆN CŨ ---
main_frame = tk.Frame(root, padx=20, pady=20)
main_frame.pack(fill=tk.BOTH, expand=True)

# --- STATUS BAR ---
status_bar = tk.Frame(root, bd=1, relief=tk.SUNKEN)
status_bar.pack(side=tk.BOTTOM, fill=tk.X)

# Tạo canvas để vẽ "bong bóng" (indicator)
canvas_status = tk.Canvas(status_bar, width=20, height=20, highlightthickness=0)
canvas_status.pack(side=tk.LEFT, padx=5, pady=2)
bubble = canvas_status.create_oval(3, 3, 17, 17, fill="red") # Mặc định là đỏ

# Dòng chữ trạng thái
lbl_status_text = tk.Label(status_bar, text="Checking connection...", bd=0)
lbl_status_text.pack(side=tk.LEFT, padx=5)

# Thêm Label để hiển thị trạng thái license
status_label = tk.Label(status_bar, text="License status", anchor="w")
status_label.pack(side=tk.LEFT, fill=tk.X, padx=5)

# Khởi chạy luồng kiểm tra
threading.Thread(target=update_status, daemon=True).start()
root.state('zoomed')
#khởi tạo thông số ban đầu: thư mục ketoan_nhan công thuê ngoài
#->globalconfig.folder_nhancongthuengoai
globalconfig.folder_nhancongthuengoai = khoitaothongso()#globalconfig.folder_nhancongthuengoai = 
#print (f"globalconfig.folder_nhancongthuengoai:  {globalconfig.folder_nhancongthuengoai}")
root.mainloop()