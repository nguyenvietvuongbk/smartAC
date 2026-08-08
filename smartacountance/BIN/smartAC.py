import os
import sys
import socket
import threading
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import pandas as pd
import warnings

import shutil
import openpyxl
from openpyxl import load_workbook
#import module function
from module import globalconfig
from module.utils import inhopdongnhancong
from module.utils import inhopdong_ca_nhan
from module.utils import in_phu_luc_to_doi
from module.utils import in_hop_dong_giao_khoan_to_truong
from module.utils import get_DSNC_path
from module.utils import convert_number_to_vietnamese
from module.utils import nhap_luong_dialog
from module.utils import quan_ly_danh_muc_cong_viec
from module.utils import phu_luc_cong_viec_hop_dong
from module.utils import addcontract
from module.utils import hien_thi_danh_sach_hop_dong
from module.utils import clear_container
from module.utils import load_names
from module.utils import mo_dialog_hop_dong
from module.utils import back_up_system
from module.utils import restore_system
from module.utils import add_new_team
from module.utils import nhap_thong_tin_du_an
from module.license_utils import check_license_and_launch
from module.license_utils import get_hwid
from module.globalconfig import khoitaothongso
from pathlib import Path
import unicodedata
import customtkinter as ctk
from PIL import Image
# Bỏ qua cảnh báo từ openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
def normalize_text(text):
    return unicodedata.normalize('NFC', str(text))
def get_project_root():
    # 1. Xác định vị trí file thực thi hoặc file script hiện tại
    if getattr(sys, 'frozen', False):
        # Đang chạy .exe: Root chính là thư mục chứa .exe
        return Path(sys.executable).parent
    else:
        # Đang chạy .py: Bắt đầu từ vị trí file hiện tại
        current_path = Path(__file__).resolve().parent.parent          
        return current_path

def get_path(relative_path):
    """Hàm tiện ích gom chung cho mọi loại đường dẫn"""
    print(f"root: {get_project_root()}")
    base_dir = get_project_root()
    
    path_at_root = os.path.join(base_dir, relative_path)
    path_at_internal = os.path.join(base_dir, "_internal", relative_path)
    if os.path.exists(path_at_root):
        return path_at_root
    # Nếu không, kiểm tra xem nó có bị đẩy vào _internal không
    if os.path.exists(path_at_internal):
        return path_at_internal    
        
    return path_at_root 



# Hàm mở file Excel
def open_excel_file():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    dsnc_path = get_DSNC_path()#os.path.join(parent_dir, 'Data', 'dulieuduan.xlsx')
    
    if os.path.exists(dsnc_path):
        os.startfile(dsnc_path)  # Lệnh mở file mặc định trên Windows
    else:
        messagebox.showerror("Lỗi", "Không tìm thấy file để mở!")
        


def get_selection():
    selected_item = tree.selection()
    if selected_item:
        item_data = tree.item(selected_item)['values']
        messagebox.showinfo("Đã chọn", f"Bạn đã chọn: {item_data[0]} (MST: {item_data[1]})")
    else:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn một hàng trong bảng!")
def print_selected_team():
    #print (f"do the print_selected_team")
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy text của item được chọn (tên Tổ đội)
    team_name = tree.item(selected_item, 'text')
    
    if messagebox.askyesno("Xác nhận", f"Bạn muốn in toàn bộ hợp đồng cho {team_name}?"):
        inhopdongnhancong(selected_team=team_name)
        messagebox.showinfo("Hoàn tất", "Đã in xong Cam kết cho tổ đội!")


def print_team_appendix():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    
    if messagebox.askyesno("Xác nhận", f"In phụ lục danh sách cho: {team_name}?"):
        in_phu_luc_to_doi(team_name)
        #messagebox.showinfo("Thành công", "Đã xuất phụ lục danh sách tổ!")

def danh_muc_cong_viec():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    team_name = unicodedata.normalize('NFC', team_name)    
    quan_ly_danh_muc_cong_viec(team_name)
def add_hopdong():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
    os.makedirs(folder_path, exist_ok=True)
    catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    if not os.path.exists(catalog_path):
        messagebox.showerror("Lỗi", "Chưa có file danh mục công việc!")
        return
    #addcontract(team_name)
    mo_dialog_hop_dong(team_name)
    #globalconfig.root = root
    #root.withdraw()
def tao_hop_dong_nhan_cong():
    tree = globalconfig.tree_todoi
    if tree is None:
        messagebox.showwarning("Thông báo", "Vui lòng mở danh sách tổ đội!")
        return # Dừng hàm lại để tránh lỗi tiếp theo
    #selected_item = tree.selection()[0]
    tree = globalconfig.tree_todoi
    if not tree.winfo_exists():
        if not globalconfig.curent_team_name =="":
           mo_dialog_hop_dong(globalconfig.curent_team_name)
        else:
           messagebox.showwarning("Thông báo", "Vui lòng chọn một tổ đội hoặc nhân viên trong bảng danh sách tổ đội!") 
    else:
        
        selection = tree.selection() # Lấy tuple chứa các ID đang được chọn

        if selection:  # Kiểm tra nếu selection không rỗng (chắc chắn có item được chọn)
            selected_item = selection[0]
            parent_id = tree.parent(selected_item)
        
            # 3. Quyết định target_item
            # Nếu parent_id là "", tức là đang chọn cha -> giữ nguyên.
            # Nếu parent_id khác "", tức là đang chọn con -> lấy cha của nó.
            target_item = selected_item if parent_id == "" else parent_id
            # ... viết code xử lý của bạn ở đây với selected_item ...
            #print(f"Item được chọn là: {selected_item}")
            team_name = tree.item(target_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
            mo_dialog_hop_dong(team_name)    
        else:
            # Xử lý trường hợp người dùng chưa chọn gì
            if not globalconfig.curent_team_name =="":
               mo_dialog_hop_dong(globalconfig.curent_team_name)
            else:
               messagebox.showwarning("Thông báo", "Vui lòng chọn một tổ đội hoặc nhân viên trong bảng danh sách tổ đội!")
        return
def hien_thi_ds_hopdong():
    tree = globalconfig.tree_todoi
    #selected_item = tree.selection()[0]
    tree = globalconfig.tree_todoi
    if not tree.winfo_exists():
        if not globalconfig.curent_team_name =="":
           hien_thi_danh_sach_hop_dong(globalconfig.curent_team_name)
        else:
           messagebox.showwarning("Thông báo", "Vui lòng chọn một tổ đội hoặc nhân viên trong bảng danh sách tổ đội!") 
    else:
        
        selection = tree.selection() # Lấy tuple chứa các ID đang được chọn

        if selection:  # Kiểm tra nếu selection không rỗng (chắc chắn có item được chọn)
            selected_item = selection[0]
            parent_id = tree.parent(selected_item)
        
            # 3. Quyết định target_item
            # Nếu parent_id là "", tức là đang chọn cha -> giữ nguyên.
            # Nếu parent_id khác "", tức là đang chọn con -> lấy cha của nó.
            target_item = selected_item if parent_id == "" else parent_id
            # ... viết code xử lý của bạn ở đây với selected_item ...
            #print(f"Item được chọn là: {selected_item}")
            team_name = tree.item(target_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
            hien_thi_danh_sach_hop_dong(team_name)    
        else:
            # Xử lý trường hợp người dùng chưa chọn gì
            if not globalconfig.curent_team_name =="":
               hien_thi_danh_sach_hop_dong(globalconfig.curent_team_name)
            else:
               messagebox.showwarning("Thông báo", "Vui lòng chọn một tổ đội hoặc nhân viên trong bảng danh sách tổ đội!")
        return                
def print_team_contract():
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    
    if messagebox.askyesno("Xác nhận", f"In phụ lục danh sách cho: {team_name}?"):
        in_hop_dong_giao_khoan_to_truong(team_name)

def is_valid_date(date_val):
    # 1. Kiểm tra nếu giá trị là NaN, None hoặc rỗng
    if pd.isna(date_val) or date_val == "":
        return False        
    # 2. Chuyển sang chuỗi và xóa khoảng trắng thừa
    date_str = str(date_val).strip()    
    # 3. Nếu sau khi chuyển thành chuỗi mà nó vẫn là "nan" (do lỗi ép kiểu)
    if date_str.lower() == 'nan' or date_str.lower() == 'nat':
        return False
    try:
        pd.to_datetime(date_str, format='%d/%m/%Y', errors='raise')
        return True
    except (ValueError, TypeError):
        return False
def is_number(s):
    try:
        float(s) # Thử chuyển sang số thực (bao gồm cả số nguyên)
        return True
    except ValueError:
        return False
        
def open_contract_input_dialog(team_name, callback):
    # Đường dẫn file
    dsnc_path = get_DSNC_path()
    # 1. Đọc sheet NCList vào DataFrame
    df_nc = pd.read_excel(dsnc_path, sheet_name='NCList', engine='openpyxl')
    
    # 2. Lọc dữ liệu theo team_name
    # Tìm dòng mà cột 'Tổ đội' bằng team_name
    team_data = df_nc[df_nc['Tổ đội'] == team_name]
    manager_data = team_data[team_data['Công việc'] == 'Tổ trưởng']
    # Đọc cấu hình đơn vị từ sheet Validate
    try:
        df_val = pd.read_excel(dsnc_path, sheet_name='Validate', engine='openpyxl')
        units = df_val['Đơn vị thời gian hợp đồng'].dropna().tolist()
    except:
        units = ['Ngày', 'Tháng', 'Năm', 'Giờ'] # Fallback nếu lỗi
        
    if not team_data.empty:
        # Lấy giá trị của dòng đầu tiên tìm thấy (nếu tổ đội có nhiều dòng, nó lấy dòng đầu tiên)
        row = manager_data.iloc[0]
        # Lấy giá trị từ các cột tương ứng
        
        default_start = str(row['Ngày bắt đầu hợp đồng']) if 'Ngày bắt đầu hợp đồng' in row.index else pd.Timestamp.now().strftime('%d/%m/%Y')
        if not is_valid_date(default_start):
            default_start = pd.Timestamp.now().strftime('%d/%m/%Y')
            print(f"default_start is not valid")
            
        try:
            default_duration = str(int(row['Thời gian hợp đồng'])) if 'Thời gian hợp đồng' in row.index else ''
        except:
           default_duration = ''
           
        default_unit = str(row['Đơn vị thời gian hợp đồng']) if 'Đơn vị thời gian hợp đồng' in row.index else units[0]
        try:
            default_price = str(int(row['Giá trị khoán'])) if 'Giá trị khoán' in row.index else ''
        except:
            default_price = ''  
        try:
            so_hop_dong = str(row['Số hợp đồng']) if 'Số hợp đồng' in row.index else ''
            print(f"Số hợp đồng đọc được : {so_hop_dong}")
        except:
            print(f"Không tìm thấy số hợp đồng trong  {row['Tổ đội']}")
            so_hop_dong = ''  
            
    else:
        # Nếu không tìm thấy tổ đội, dùng giá trị mặc định hoặc rỗng
        default_start = pd.Timestamp.now().strftime('%d/%m/%Y')
        default_duration = ''
        default_unit = units[0]
        default_price = ''
        so_hop_dong = ''
    # Tạo cửa sổ phụ
    dialog = tk.Toplevel()
    dialog.title(f"Nhập thông tin hợp đồng: {team_name}")
    dialog.geometry("350x350")

    tk.Label(dialog, text="Ngày bắt đầu (DD/MM/YYYY):").pack(pady=5)
    ent_start = tk.Entry(dialog)
    ent_start.pack()
    #ent_start.insert(0, pd.Timestamp.now().strftime('%d/%m/%Y'))
    ent_start.insert(0, default_start)
    
    tk.Label(dialog, text="Thời gian hợp đồng:").pack(pady=5)
    ent_duration = tk.Entry(dialog)
    ent_duration.pack()
    ent_duration.insert(0, default_duration)

    tk.Label(dialog, text="Đơn vị thời gian:").pack(pady=5)
    combo_unit = ttk.Combobox(dialog, values=units)
    combo_unit.pack()
    if default_unit in units:
        combo_unit.set(default_unit)
    else:
        # Nếu giá trị trong file Excel không nằm trong danh sách đơn vị (ví dụ bị sai dữ liệu), 
        # ta quay về chọn mặc định đầu tiên
        combo_unit.current(0)

    tk.Label(dialog, text="Tiền khoán:").pack(pady=5)
    ent_price = tk.Entry(dialog)
    ent_price.pack()
    # Load dữ liệu cũ nếu có
    ent_price.insert(0, default_price)
    # 1. Thêm Label để hiển thị chữ
    lbl_amount_text = tk.Label(dialog, text="", fg="blue", font=("Arial", 9, "italic"))
    lbl_amount_text.pack(pady=5)
    
    tk.Label(dialog, text="Số hợp đồng:").pack(pady=5)
    ent_sohopdong = tk.Entry(dialog)
    ent_sohopdong.pack()
    # Load dữ liệu cũ nếu có
    ent_sohopdong.insert(0, so_hop_dong)
    #2. Hàm cập nhật chữ khi gõ
    def update_amount_text(event=None):
        raw_val = ent_price.get()
        # Loại bỏ các ký tự không phải số
        clean_val = ''.join(filter(str.isdigit, raw_val))
        if clean_val:
            # Gọi hàm chuyển đổi của bạn
            chu_so = convert_number_to_vietnamese(int(clean_val))
            lbl_amount_text.config(text=f"Bằng chữ: {chu_so}")
        else:
            lbl_amount_text.config(text="")
    # 3. Gán sự kiện 'KeyRelease' (ngay khi gõ xong 1 ký tự)
    ent_price.bind('<KeyRelease>', update_amount_text)
    def on_submit():
        start_date_str = ent_start.get()
        # 1. Kiểm tra tính hợp lệ của ngày tháng
        # Sử dụng hàm is_valid_date mà chúng ta đã xây dựng trước đó
        if not is_valid_date(start_date_str):
            messagebox.showerror("Lỗi định dạng", "Ngày bắt đầu không hợp lệ!\nVui lòng nhập theo định dạng DD/MM/YYYY")
            # 2. Bôi đen toàn bộ nội dung trong ent_start
            ent_start.select_range(0, 'end')
            ent_start.focus_set() # Đưa con trỏ chuột vào lại ô đó
            return # Dừng lại, không thực hiện các bước tiếp theo
        
        
        if not is_number(ent_duration.get()):
            messagebox.showerror("Lỗi định dạng", "Thời gian hợp đồng không hợp lệ!\nVui lòng nhập thời gian hợp đồng là con số")
            ent_duration.select_range(0, 'end')
            ent_duration.focus_set() # Đưa con trỏ chuột vào lại ô đó
            return # Dừng lại, không thực hiện các bước tiếp theo
        if not is_number(ent_price.get()):
            messagebox.showerror("Lỗi định dạng", "Giá trị hợp đồng không hợp lệ!\nVui lòng nhập giá trị hợp đồng là con số")
            ent_price.select_range(0, 'end')
            ent_price.focus_set() # Đưa con trỏ chuột vào lại ô đó
            return # Dừng lại, không thực hiện các bước tiếp theo    
        if combo_unit.get() not in units:
            messagebox.showerror("Lỗi định dạng", "Giá trị đơn vị thời gian hợp đồng không hợp lệ!\nVui lòng chọn một giá trị trong list")
            combo_unit.current(0)
            combo_unit.select_range(0, 'end')
            combo_unit.focus_set()
            return # Dừng lại, không thực hiện các bước tiếp theo
        data = {
            'start': ent_start.get(),
            'duration': ent_duration.get(),
            'unit': combo_unit.get(),
            'price': ent_price.get(), # Lấy giá trị tiền khoán
            'so_hop_dong': ent_sohopdong.get()
        }
        dialog.destroy()
        callback(team_name, data)

    tk.Button(dialog, text="Xác nhận & In", command=on_submit).pack(pady=20)
def process_contract_printing(team_name, data):
    dsnc_path = get_DSNC_path()
    # 1. Mở workbook
    wb = openpyxl.load_workbook(dsnc_path)
    ws = wb['NCList']
    
    # 2. Tìm hoặc tạo cột 'Giá trị khoán'
    header = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    
    if 'Giá trị khoán' not in header:
        new_col_idx = ws.max_column + 1
        ws.cell(row=1, column=new_col_idx, value='Giá trị khoán')
        header['Giá trị khoán'] = new_col_idx
    if 'Số hợp đồng' not in header:
        new_col_idx = ws.max_column + 1
        ws.cell(row=1, column=new_col_idx, value='Số hợp đồng')
        header['Số hợp đồng'] = new_col_idx
    # Xác định các cột cần thiết
    col_to_doi = header.get('Tổ đội')
    col_start = header.get('Ngày bắt đầu hợp đồng')
    col_duration = header.get('Thời gian hợp đồng')
    col_unit = header.get('Đơn vị thời gian hợp đồng')
    col_gia_tri = header.get('Giá trị khoán')
    col_cong_viec = header.get('Công việc') # Giả sử cột này tên là 'Công việc'
    col_so_hop_dong = header.get('Số hợp đồng')
    if not all([col_to_doi, col_start, col_duration, col_unit, col_gia_tri]):
        messagebox.showerror("Lỗi", "Không tìm thấy đủ các cột cần thiết trong file Excel!")
        return

    # 3. Duyệt qua các dòng để tìm và cập nhật
    updated = False
    for row in range(2, ws.max_row + 1):
        team_val = ws.cell(row=row, column=col_to_doi).value
        
        if team_val == team_name:
            # Cập nhật thông tin chung cho các dòng của tổ đội
            ws.cell(row=row, column=col_start).value = data['start']
            ws.cell(row=row, column=col_duration).value = data['duration']
            ws.cell(row=row, column=col_unit).value = data['unit']
            #ws.cell(row=row, column=col_so_hop_dong).value = data['so_hop_dong']
            # Cập nhật giá trị khoán riêng cho người có Công việc = 'Tổ trưởng'
            # Kiểm tra nếu cột Công việc tồn tại
            if col_cong_viec:
                cong_viec = ws.cell(row=row, column=col_cong_viec).value
                if cong_viec == 'Tổ trưởng':
                    ws.cell(row=row, column=col_gia_tri).value = data.get('price', 0)
                    ws.cell(row=row, column=col_so_hop_dong).value = data.get('so_hop_dong', 0)
            updated = True
    
    # 4. Lưu workbook
    if updated:
        wb.save(dsnc_path)
        print(f"Đã cập nhật dữ liệu cho tổ: {team_name}")
        in_hop_dong_giao_khoan_to_truong(team_name)
        messagebox.showinfo("Hoàn tất", f"Đã cập nhật dữ liệu và in HĐ cho {team_name}")
    else:
        messagebox.showwarning("Thông báo", f"Không tìm thấy tổ đội: {team_name}")
def get_safe_team_name():
    selected_item = tree.selection()
    if not selected_item:
        return None
    
    item_values = tree.item(selected_item[0])['values']
    #team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    # Kiểm tra xem values có tồn tại và không rỗng
    if item_values and len(item_values) > 0:       
        return item_values[0].split(" - ")[0].split(" (TT:")[0].strip()
    else:
        # Trường hợp Node cha không có giá trị ở cột 0, 
        # ta lấy text của chính nó (thường là tên tổ đội)
        return tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
# Hàm kích hoạt menu khi chuột phải


    
def push_to_server_function():
    return
    
def get_from_server_function():
    return
    
def check_internet():
    """Kiểm tra kết nối internet bằng cách thử kết nối tới Google DNS."""
    try:
        socket.setdefaulttimeout(3)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(("8.8.8.8", 53))
        return True
    except Exception:
        return False

def update_status():
    """Cập nhật màu sắc của bong bóng 3D dựa trên trạng thái mạng."""
    is_online = check_internet()
    color = "green" if is_online else "red"
    text = "Connected" if is_online else "No internet connection"
    
    # Cập nhật màu bong bóng
    canvas_status.itemconfig(bubble, fill=color)
    lbl_status_text.config(text=text)
    
    # Tự động kiểm tra lại sau mỗi 10 giây
    root.after(10000, update_status)
def open_output():
    # 1. Xác định đường dẫn thư mục Output
    # Thường nằm cùng cấp với file .exe hoặc file script đang chạy
    if globalconfig.curent_team_name == "":
        messagebox.showerror("Lỗi", "Chưa chọn tổ đội")
        return
    output_dir = os.path.join(os.getcwd(), "Output", globalconfig.ketoan_nhancongthuengoai, globalconfig.curent_team_name)

    # 2. Kiểm tra nếu chưa tồn tại thì tạo mới để tránh lỗi
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo thư mục Output: {e}")
            return

    # 3. Mở thư mục bằng File Explorer của Windows
    try:
        os.startfile(output_dir)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể mở thư mục: {e}")
def open_restore_system():
    # 1. Xác định đường dẫn thư mục Output
    # Thường nằm cùng cấp với file .exe hoặc file script đang chạy
    output_dir = os.path.join(os.getcwd(), "backupsystem")

    # 2. Kiểm tra nếu chưa tồn tại thì tạo mới để tránh lỗi
    if not os.path.exists(output_dir):        
       messagebox.showerror("Lỗi", "Chưa có bản sao lưu nào")
       return

    # 3. Mở thư mục bằng File Explorer của Windows
    try:
        os.startfile(output_dir)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể mở thư mục: {e}") 
def open_danh_muc_congviec():
    tree = globalconfig.tree_todoi
    if tree is None:
        messagebox.showwarning("Thông báo", "Vui lòng mở danh sách tổ đội!")
        return # Dừng hàm lại để tránh lỗi tiếp theo
    #selected_item = tree.selection()[0]
    tree = globalconfig.tree_todoi
    if not tree.winfo_exists():
        if not globalconfig.curent_team_name =="":
           quan_ly_danh_muc_cong_viec(globalconfig.curent_team_name)
        else:
           messagebox.showwarning("Thông báo", "Vui lòng chọn một tổ đội hoặc nhân viên trong bảng danh sách tổ đội!") 
    else:
        
        selection = tree.selection() # Lấy tuple chứa các ID đang được chọn

        if selection:  # Kiểm tra nếu selection không rỗng (chắc chắn có item được chọn)
            selected_item = selection[0]
            parent_id = tree.parent(selected_item)
        
            # 3. Quyết định target_item
            # Nếu parent_id là "", tức là đang chọn cha -> giữ nguyên.
            # Nếu parent_id khác "", tức là đang chọn con -> lấy cha của nó.
            target_item = selected_item if parent_id == "" else parent_id
            # ... viết code xử lý của bạn ở đây với selected_item ...
            #print(f"Item được chọn là: {selected_item}")
            team_name = tree.item(target_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
            quan_ly_danh_muc_cong_viec(team_name)    
        else:
            # Xử lý trường hợp người dùng chưa chọn gì
            if not globalconfig.curent_team_name =="":
               quan_ly_danh_muc_cong_viec(globalconfig.curent_team_name)
            else:
               messagebox.showwarning("Thông báo", "Vui lòng chọn một tổ đội hoặc nhân viên trong bảng danh sách tổ đội!")
        return    
def checkvar():
    print(f"globalconfig.curent_team_name: {globalconfig.curent_team_name}") 

    #back up get_path("Data")-> get_path("Back_up")
    #Tạo file nén back_up_{number}.7z từ Data và ghi vào get_path("Back_up")
        
# --- THIẾT LẬP GIAO DIỆN ---                
#root = tk.Tk()
root = ctk.CTk()
root.option_add('*Menu.font', ('Segoe UI', 12))
#create sidebar menu
# 1. Khởi tạo Sidebar


# Hàm tạo nút con (Tái sử dụng)
def create_menu_button(text, command, parent_):
    return ctk.CTkButton(
        parent_,
        text=text,
        fg_color="transparent",
        text_color="black",
        hover_color="#3B8ED0",
        anchor="w",
        height=40,
        command=command
    )
# --- 1. SIDEBAR ---
sidebar = ctk.CTkFrame(master=root, width=200, corner_radius=0)
sidebar.pack(side="left", fill="y")

# --- MENU 1: KẾ TOÁN NHÂN CÔNG ---
def toggle_menu_1():
    # Kiểm tra xem menu 1 đang hiện hay ẩn
    if menu_con_frame_1.winfo_ismapped():
        menu_con_frame_1.pack_forget() # Ẩn menu con 1
    else:
        # Khi hiện, pack nó vào giữa btn_cha_1 và btn_cha_2
        # Việc này sẽ tự động đẩy mọi thứ phía sau nó (bao gồm btn_cha_2) xuống
        menu_con_frame_1.pack(fill="x", after=btn_cha_1) 

btn_cha_1 = ctk.CTkButton(sidebar, text="📂 Kế toán nhân công ▼", anchor="w", command=toggle_menu_1)
btn_cha_1.pack(fill="x", pady=10, padx=10) # 1. Nút 1

menu_con_frame_1 = ctk.CTkFrame(sidebar, fg_color="transparent")
# Thêm nút vào khung 1
create_menu_button("Nhập thông tin dự án", lambda: nhap_thong_tin_du_an(globalconfig.root), menu_con_frame_1).pack(fill="x", padx=10)
#create_menu_button("Tải danh mục nhân công", lambda: load_names(), menu_con_frame_1).pack(fill="x", padx=10)
create_menu_button("Tạo mới tổ đội", lambda: add_new_team(globalconfig.root), menu_con_frame_1).pack(fill="x", padx=10)#-> danh_sach_to_doi.JSON
create_menu_button("Khai báo danh mục công việc", lambda: open_danh_muc_congviec(), menu_con_frame_1).pack(fill="x", padx=10)
create_menu_button("Tạo mới hợp đồng nhân công", lambda: tao_hop_dong_nhan_cong(), menu_con_frame_1).pack(fill="x", padx=10)
create_menu_button("Mở danh mục hợp đồng nhân cồng", lambda: hien_thi_ds_hopdong(), menu_con_frame_1).pack(fill="x", padx=10)
create_menu_button("Mở thư mục in", lambda: open_output(), menu_con_frame_1).pack(fill="x", padx=10)

# --- MENU 2: PHẢI THU (Sẽ nằm dưới menu 1) ---
def toggle_menu_2():
    if menu_con_frame_2.winfo_ismapped():
        menu_con_frame_2.pack_forget()
    else:
        menu_con_frame_2.pack(fill="x", after=btn_cha_2)

def display_about():
    about_win = ctk.CTkToplevel(globalconfig.root)
    about_win.title("Giới thiệu phần mềm")
    about_win.geometry("440x360")
    about_win.resizable(False, False)
    about_win.attributes("-topmost", True)
    about_win.grab_set()

    main_frame = ctk.CTkFrame(about_win, corner_radius=10, fg_color="#FFFFFF", border_width=1, border_color="#D1D1D1")
    main_frame.pack(fill="both", expand=True, padx=15, pady=15)

    lbl_title = ctk.CTkLabel(
        main_frame, 
        text="Smart Accountance Management", 
        font=("Arial", 16, "bold"), 
        text_color="#1F6AA5"
    )
    lbl_title.pack(pady=(15, 5))

    lbl_version = ctk.CTkLabel(
        main_frame, 
        text="Version: 1.0.0.1", 
        font=("Arial", 12, "italic"), 
        text_color="#555555"
    )
    lbl_version.pack(pady=(0, 10))

    # --- PHẦN HIỂN THỊ PRODUCT ID CÓ THỂ COPY ---
    frame_hwid = ctk.CTkFrame(main_frame, fg_color="transparent")
    frame_hwid.pack(fill="x", padx=20, pady=(0, 10))

    lbl_pid_title = ctk.CTkLabel(frame_hwid, text="Product ID:", font=("Arial", 12, "bold"), text_color="#333333")
    lbl_pid_title.pack(anchor="w", pady=(0, 2))

    # Frame con chứa ô Entry và Nút Copy nằm ngang
    sub_frame = ctk.CTkFrame(frame_hwid, fg_color="transparent")
    sub_frame.pack(fill="x")

    try:
        product_id = get_hwid()
    except Exception:
        product_id = "Không xác định"

    # Dùng CTkEntry nhưng cấu hình để người dùng có thể bôi đen/copy nhưng không sửa được
    ent_pid = ctk.CTkEntry(
        sub_frame, 
        font=("Arial", 11), 
        fg_color="#F0F0F0", 
        text_color="#333333",
        border_color="#D1D1D1",
        height=32
    )
    ent_pid.insert(0, product_id)
    ent_pid.configure(state="readonly") # Khóa không cho gõ sửa, nhưng cho phép click, bôi đen, Ctrl+C
    ent_pid.pack(side="left", fill="x", expand=True, padx=(0, 5))

    # Hàm thực hiện copy vào bộ nhớ tạm (Clipboard)
    def copy_to_clipboard():
        about_win.clipboard_clear()
        about_win.clipboard_append(product_id)
        about_win.update() # Cập nhật clipboard hệ thống
        # Hiệu ứng đổi tên nút tạm thời để báo thành công
        btn_copy.configure(text="Đã chép!", fg_color="#28A745")
        about_win.after(1500, lambda: btn_copy.configure(text="Sao chép", fg_color="#3B8ED0"))

    btn_copy = ctk.CTkButton(
        sub_frame, 
        text="Sao chép", 
        width=80, 
        height=32,
        font=("Arial", 11, "bold"),
        fg_color="#3B8ED0",
        hover_color="#275E8D",
        command=copy_to_clipboard
    )
    btn_copy.pack(side="right")
    # ---------------------------------------------

    # Các thông tin còn lại giữ nguyên dạng Label
    info_text = (
        "Nhà phát triển: SME Support\n"
        "Địa chỉ: Xuân Phương, Hà Nội\n"
        "Điện thoại: 0972833961\n"
        "Email: mpsoftwarepro1@gmail.com"
    )
    
    lbl_info = ctk.CTkLabel(
        main_frame, 
        text=info_text, 
        font=("Arial", 12), 
        justify="left", 
        text_color="#333333"
    )
    lbl_info.pack(padx=20, pady=(5, 15), anchor="w")

    def close_about():
        about_win.grab_release()
        about_win.destroy()

    btn_close = ctk.CTkButton(
        main_frame, 
        text="Đóng", 
        width=120, 
        height=35, 
        corner_radius=8,
        fg_color="#3B8ED0", 
        hover_color="#275E8D",
        font=("Arial", 12, "bold"),
        command=close_about
    )
    btn_close.pack(pady=(0, 15))

    about_win.protocol("WM_DELETE_WINDOW", close_about)
    
btn_cha_2 = ctk.CTkButton(sidebar, text="💰Phải thu ▼", anchor="w", command=toggle_menu_2)
btn_cha_2.pack(fill="x", pady=10, padx=10) # 2. Nút 2 (Luôn nằm dưới Nút 1)

menu_con_frame_2 = ctk.CTkFrame(sidebar, fg_color="transparent")
# Thêm nút vào khung 2
create_menu_button("📄 Lập hóa đơn", lambda: print("C"), menu_con_frame_2).pack(fill="x", padx=10)
create_menu_button("📋 Công nợ", lambda: print("D"), menu_con_frame_2).pack(fill="x", padx=10)
# 1. Cấu hình Treeview (Bảng)
style = ttk.Style()
style.theme_use("clam")

# 1. Cấu hình bảng dữ liệu
style = ttk.Style()
style.theme_use("clam")

# Cấu hình bảng sáng - tạo không gian mở
style.configure("Treeview", 
                background="#FFFFFF",       # Nền trắng tinh khôi
                foreground="#333333",       # Chữ đen xám (dịu hơn đen tuyền)
                fieldbackground="#FFFFFF",
                rowheight=35,               # Giãn cách dòng thoáng đãng
                borderwidth=0)

# Cấu hình Header - màu xanh xám nhạt (Industrial style)
style.configure("Treeview.Heading", 
                background="#E1E4E8",       # Màu xám xanh nhạt
                foreground="#2C3E50",       # Chữ đậm nổi bật
                relief="flat",
                font=('Segoe UI', 12, 'bold'))
style.configure("Treeview", 
                background="#FFFFFF", 
                foreground="#333333", 
                fieldbackground="#FFFFFF",
                rowheight=40,            # Tăng rowheight lên một chút cho tương xứng với font mới
                font=('Segoe UI', 12),   # <--- THÊM DÒNG NÀY (Chỉnh kích thước 12 hoặc hơn)
                borderwidth=0)                

# Cấu hình khi chọn hàng - Màu xanh dương dịu
style.map('Treeview', 
          background=[('selected', '#5C91D4')], 
          foreground=[('selected', '#FFFFFF')])
                
                
root.title("Smart AC Management")
#root.geometry("600x500")
#context_menu
#context_menu = tk.Menu(root, tearoff=0)
# --- 1. TẠO THANH MENU ---
menubar = tk.Menu(root)
root.config(menu=menubar)

# Menu File
file_menu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="File", menu=file_menu)
#file_menu.add_command(label="Tải danh sách tổ đội", font=("Segoe UI", 12), command=load_names)
#file_menu.add_command(label="Mở file dữ liệu dự án", font=("Segoe UI", 12), command=open_excel_file)
#file_menu.add_command(label="Mở thư mục in ấn", font=("Segoe UI", 12), command=open_output)
file_menu.add_command(label="Sao lưu dữ liệu dự án_Archive", font=("Segoe UI", 12), command= lambda: back_up_system())
file_menu.add_command(label="Phục hồi dữ liệu dự án_Retrive", font=("Segoe UI", 12), command= lambda: restore_system())
file_menu.add_command(label="Mở thư mục chứa các bản sao", font=("Segoe UI", 12), command= lambda: open_restore_system())
file_menu.add_separator()
file_menu.add_command(label="Thoát",font=("Segoe UI", 12), command=root.quit)

# Menu Server
server_menu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="Server", menu=server_menu)
server_menu.add_command(label="Push to server", command=push_to_server_function) # Thay bằng tên hàm của bạn
server_menu.add_command(label="Get from server", command=get_from_server_function) # Thay bằng tên hàm của bạn
#Help menu
help_menu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="Help", menu=help_menu)
help_menu.add_command(label="About", command=display_about) # Thay bằng tên hàm của bạn
# --- PHẦN GIAO DIỆN CŨ ---
main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack(fill=tk.BOTH, expand=True)

global_container = tk.Frame(main_frame) 
global_container.pack(fill=tk.BOTH, expand=True, pady=0)
globalconfig.global_container = global_container # Lưu vào config để dễ gọi
# --- STATUS BAR ---
status_bar = tk.Frame(root, bd=1, relief=tk.SUNKEN)
status_bar.pack(side=tk.BOTTOM, fill=tk.X)

# Tạo canvas để vẽ "bong bóng" (indicator)
canvas_status = tk.Canvas(status_bar, width=20, height=20, highlightthickness=0)
canvas_status.pack(side=tk.LEFT, padx=5, pady=2)
bubble = canvas_status.create_oval(3, 3, 17, 17, fill="red") # Mặc định là đỏ

# Dòng chữ trạng thái
lbl_status_text = tk.Label(status_bar, text="Checking connection...", bd=0)
lbl_status_text.pack(side=tk.LEFT, padx=5)

# Thêm Label để hiển thị trạng thái license
status_label = tk.Label(status_bar, text="License status", anchor="w")
status_label.pack(side=tk.LEFT, fill=tk.X, padx=5)
globalconfig.status_label = status_label
# Khởi chạy luồng kiểm tra
threading.Thread(target=update_status, daemon=True).start()
#root.state('zoomed')
globalconfig.root = root
#khởi tạo thông số ban đầu: thư mục ketoan_nhan công thuê ngoài
#->globalconfig.folder_nhancongthuengoai
globalconfig.folder_nhancongthuengoai = khoitaothongso()#globalconfig.folder_nhancongthuengoai = 
#hiển thị danh sách nhân công
load_names()
#print (f"globalconfig.folder_nhancongthuengoai:  {globalconfig.folder_nhancongthuengoai}")
# Phóng to cửa sổ
root.update()
root.state('zoomed')
root.mainloop()