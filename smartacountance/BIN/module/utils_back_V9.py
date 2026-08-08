#updater
#16/7/2026
#thêm chức năng thêm biên bản nghiệm thu
#17/7/2026
#thêm chức năng thêm đê nghị thanh toán
#20/7/2026
#bổ sung tên nghiệm thu vào dữ liệu nghiệm thu
#tạo menu in_hop_dong
#21/07/2026
#thêm chức năng in bảng chia thu nhập - thanh toán
#27/7/2026
#hợp nhất các dialog chính trên một root , hiển thị danh sách hợp đồng trong globalconfig.global_container
#update utils.addcontract(team_name) : để treeview fit content include header and cell, điều chỉnh chiều cao của treeview để fit được nội dung của bảng
#auto_resize_columns(tree) để fit được các nội dung trong bảng
#29/7/2026
#update utils.mo_dialog_hop_dong(team_name) chống lưu hợp đồng khi khối lượng <=0
#click vào dòng chọn của tree -> hiển thị khối lượng trên ent_kl_input using tree.bind("<<TreeviewSelect>>", cap_nhat_dialog)
#Chỉ cho phép sửa với nghiệm thu cuối cùng, thanh toán cuối cùng, -> hien_thi_danh_sach_hop_dong
#áp dụng auto_resize_columns đối với mo_cua_so_nghiem_thu
#áp dụng auto_resize_columns đối với mo_cua_so_thanh_toan
#áp dụng nút đẹp đối với mo_cua_so_thanh_toan
#áp dụng căn giữa đối với các bảng treeview của danh mục hợp đồng, mở cửa sổ nghiệm thu, mở cửa số thanh toán
#sửa lỗi bấm nút tạo hợp đồng khi chưa load danh sach tổ đội (chưa có gì để chọn team_name)
#tăng kích thức font chữ
#30/7/2026
#Chuyển PDF in phụ lục, nghiệm thu, thanh toán vào đúng thư mục của phân hệ kế toán ketoan_nhancongthuengoai
#hiển thị tên tổ đội trong bảng danh mục hợp đồng
#thêm mức thu nhập chịu thuế -> Dataduan.JSON
#thêm Tên công ty, Tên công trình , Giám đốc, Chỉ huy trưởng, Kế toán trưởng , Người lập biểu, Địa điểm công trình -> Dataduan.JSON
#sử dụng Dataduan.JSON thay thế cho excel file

#Tạo chương trình đẩy các file mới cập nhật lên Gdrive
#Tạo service đọc thông tin update từ Gdrive và auto update cho phiên bản software hiện tại : temple, logiccode , library OTA
#----------------------------------------------------------------------------------------------------------------------------
import shutil
import os
import sys
from openpyxl import load_workbook
import pandas as pd
from copy import copy
from tkinter import messagebox

from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
from datetime import datetime
from num2words import num2words

from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import tkinter as tk
from tkinter import ttk, font
from tkinter import messagebox, filedialog, Toplevel, Label, Entry, Button # Thêm Toplevel vào đây
import json
from . import globalconfig
from pathlib import Path
import unicodedata
import customtkinter as ctk
from module.license_utils import check_license_and_launch
from PIL import Image
import re
import py7zr


    
def auto_resize_columns(tree):
    tree.update_idletasks()
    style = ttk.Style()
    
    # 1. Lấy font riêng cho Header và Data
    # Header thường có style là 'Treeview.Heading'
    header_font_tuple = style.lookup('Treeview.Heading', 'font', default="TkDefaultFont")
    header_font = font.Font(font=header_font_tuple)
    
    data_font_tuple = style.lookup('Treeview', 'font', default="TkDefaultFont")
    data_font = font.Font(font=data_font_tuple)
    
    # Hàm đệ quy lấy tất cả item
    all_items = tree.get_children("")
    
    # 2. Duyệt qua từng cột
    for col in tree["columns"]:
        # Đo tiêu đề (dùng header_font)
        max_width = header_font.measure(tree.heading(col, "text"))
        
        # Đo nội dung dòng (dùng data_font)
        for item in tree.get_children(""): # Chỉ cần get_children("") nếu bạn đã load đầy đủ hoặc dùng đệ quy
            text = tree.item(item, "text") if col == "#0" else tree.set(item, col)
            item_width = data_font.measure(text)
            if item_width > max_width:
                max_width = item_width
        
        # 3. Tăng padding: Header cần khoảng 40px thay vì 20px 
        # (Để chừa chỗ cho icon sắp xếp và khoảng đệm hệ thống)
        tree.column(col, width=max_width + 40, stretch=False)

def normalize_text(text):
    return unicodedata.normalize('NFC', str(text))
def get_project_root():
    # 1. Xác định vị trí file thực thi hoặc file script hiện tại
    if getattr(sys, 'frozen', False):
        # Đang chạy .exe: Root chính là thư mục chứa .exe
        return Path(sys.executable).parent
    else:
        # Đang chạy .py: Bắt đầu từ vị trí file hiện tại
        current_path = Path(__file__).resolve().parent.parent.parent          
        return current_path

def get_path(relative_path):
    """Hàm tiện ích gom chung cho mọi loại đường dẫn"""
    print(f"root: {get_project_root()}")
    base_dir = get_project_root()
    
    path_at_root = os.path.join(base_dir, relative_path)
    path_at_internal = os.path.join(base_dir, "_internal", relative_path)
    if os.path.exists(path_at_root):
        return path_at_root
    # Nếu không, kiểm tra xem nó có bị đẩy vào _internal không
    if os.path.exists(path_at_internal):
        return path_at_internal    
        
    return path_at_root
def back_up_system():
    data_dir = get_path("Data")           # Thư mục cần nén
    backup_dir = get_path("backupsystem") # Thư mục lưu file backup
    
    # 1. Kiểm tra và tạo thư mục backup nếu chưa tồn tại
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        print(f"Đã tạo thư mục: {backup_dir}")
    
    # 2. Tìm số backup lớn nhất hiện có
    max_num = 0
    files = os.listdir(backup_dir)
    
    # Regex để lọc các file có dạng backup_{number}.7z
    pattern = re.compile(r"backup_(\d+)\.7z")
    
    for f in files:
        match = pattern.search(f)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
    
    next_num = max_num + 1
    output_filename = os.path.join(backup_dir, f"backup_{next_num}.7z")
    
    # 3. Tiến hành nén dữ liệu
    try:
        if os.path.exists(data_dir):
            with py7zr.SevenZipFile(output_filename, 'w') as archive:
                archive.writeall(data_dir, 'Data') # Nén thư mục Data vào file
            print(f"Đã sao lưu thành công tại: {output_filename}")
            messagebox.showinfo("Thông báo", f"Dữ liệu đã được lưu thành công!-> {output_filename}")
            return output_filename
        else:
            print(f"Lỗi: Thư mục '{data_dir}' không tồn tại!")
            return None
    except Exception as e:
        print(f"Có lỗi xảy ra khi nén file: {e}")
        return None   
def restore_system():
    # 1. Chọn file backup
    backup_dir = get_path("backupsystem")
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        
    file_path = filedialog.askopenfilename(
        initialdir=backup_dir,
        title="Chọn file sao lưu để khôi phục",
        filetypes=[("7z files", "*.7z")]
    )
    
    if not file_path:
        return # Người dùng hủy chọn

    # 2. Cảnh báo xác nhận (Rất quan trọng vì sẽ ghi đè dữ liệu)
    confirm = messagebox.askyesno(
        "Cảnh báo", 
        "Dữ liệu hiện tại trong thư mục Data sẽ bị ghi đè bởi bản sao lưu này.\nBạn có chắc chắn muốn tiếp tục?"
    )
    
    if not confirm:
        return

    # 3. Tiến hành giải nén và khôi phục
    temp_dir = get_path("temp_restore_folder")
    try:
        # Giải nén ra thư mục tạm
        with py7zr.SevenZipFile(file_path, 'r') as archive:
            archive.extractall(path=temp_dir)
        
        # Đường dẫn thư mục Data bên trong file nén
        # Dựa theo cách hàm back_up_system trước đó (nén thư mục Data vào trong)
        source_data_path = os.path.join(temp_dir, "Data")
        dest_data_path = "Data"
        
        # Đảm bảo thư mục đích tồn tại
        if not os.path.exists(dest_data_path):
            os.makedirs(dest_data_path)
            
        # 4. Copy và ghi đè
        # dirs_exist_ok=True cho phép copy đè lên các file/thư mục đã tồn tại
        shutil.copytree(source_data_path, dest_data_path, dirs_exist_ok=True)
        
        messagebox.showinfo("Thành công", "Khôi phục dữ liệu hoàn tất!")
        load_names()
    except Exception as e:
        messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi khôi phục: {str(e)}")
        
    finally:
        # 5. Dọn dẹp thư mục tạm
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)        
def open_output():
    # 1. Xác định đường dẫn thư mục Output
    # Thường nằm cùng cấp với file .exe hoặc file script đang chạy
    team_name = globalconfig.curent_team_name
    team_name = unicodedata.normalize('NFC', team_name)
    output_dir = get_path(f"Output/{globalconfig.ketoan_nhancongthuengoai}/{team_name}")

    # 2. Kiểm tra nếu chưa tồn tại thì tạo mới để tránh lỗi
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo thư mục Output: {e}")
            return    
    # 3. Mở thư mục bằng File Explorer của Windows
    try:
        os.startfile(output_dir)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể mở thư mục: {e}")        
def is_valid_date(date_string):
    try:
        datetime.strptime(date_string, '%d/%m/%Y')
        return True
    except ValueError:
        return False
def lay_don_vi(team_name, ten_cong_viec):
    # Đường dẫn file
    file_path = get_path(os.path.join("Data", globalconfig.ketoan_nhancongthuengoai, team_name, f"khai báo danh mục công việc_{team_name}.JSON"))
    
    try:
        if not os.path.exists(file_path):
            print(f"Không tìm thấy file: {file_path}")
            return None

        with open(file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            
        # Duyệt qua danh sách để tìm công việc
        for item in data:
            if item["cong_viec"] == ten_cong_viec:
                return item["don_vi"]
        
        return "Không tìm thấy"
    
    except Exception as e:
        print(f"Lỗi khi đọc file: {e}")
        return None
def lay_don_gia(team_name, ten_cong_viec):
    # Đường dẫn đến file JSON của team
    file_path = get_path(os.path.join("Data", globalconfig.ketoan_nhancongthuengoai, team_name, f"khai báo danh mục công việc_{team_name}.JSON"))
    
    try:
        if not os.path.exists(file_path):
            print(f"Không tìm thấy file: {file_path}")
            return 0
        
        with open(file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            
        # Tạo từ điển tra cứu nhanh: {tên_công_việc: đơn_giá}
        # Lưu ý: Chuyển don_gia sang kiểu số (float hoặc int) để dễ tính toán sau này
        lookup_dict = {item["cong_viec"]: float(item["don_gia"]) for item in data}
        
        # Trả về đơn giá, mặc định là 0 nếu không tìm thấy
        return lookup_dict.get(ten_cong_viec, 0)
    
    except Exception as e:
        print(f"Lỗi khi đọc file hoặc xử lý dữ liệu: {e}")
        return 0        
def get_so_tai_khoan(ten_nguoi_lao_dong):
    """
    Hàm này lấy thông tin số tài khoản, ngân hàng và chủ tài khoản 
    của người lao động từ file JSON danh sách nhân công.
    """
    try:
        dsnc_json_path = get_DSNC_JSON()
        if not os.path.exists(dsnc_json_path):
            return "Chưa cập nhật STK"

        with open(dsnc_json_path, "r", encoding="utf-8") as f:
            dsnc_data = json.load(f)

        # Chuẩn hóa tên người lao động cần tìm
        ten_nld_norm = unicodedata.normalize('NFC', str(ten_nguoi_lao_dong)).strip().lower()

        # Duyệt qua các tổ đội trong file JSON để tìm nhân viên khớp tên
        for team_name, team_info in dsnc_data.items():
            workmen_list = team_info.get("workmen", [])
            for worker in workmen_list:
                w_name = unicodedata.normalize('NFC', str(worker.get('Họ và tên', ''))).strip().lower()
                if w_name == ten_nld_norm:
                    so_tk = str(worker.get('Số TK', '')).replace('.0', '').strip()
                    nh = str(worker.get('Tên ngân hàng', '')).strip()
                    chu_tk = str(worker.get('Tên chủ TK', '')).strip()

                    # Nếu thiếu thông tin có thể xử lý hiển thị linh hoạt
                    if so_tk or nh or chu_tk:
                        return f"{so_tk}-{nh}-{chu_tk}"
                    else:
                        return "Chưa cập nhật STK"

        return "Chưa cập nhật STK"

    except Exception as e:
        print(f"Lỗi khi lấy STK từ JSON: {e}")
        return "Lỗi kết nối"
def get_ten_doi_truong(team_name):
    """
    Hàm lấy tên đội trưởng dựa trên tên tổ đội từ file JSON danh sách nhân công.
    """
    try:
        dsnc_json_path = get_DSNC_JSON()
        if not os.path.exists(dsnc_json_path):
            return "Chưa cập nhật đội trưởng"

        with open(dsnc_json_path, "r", encoding="utf-8") as f:
            dsnc_data = json.load(f)

        team_name_norm = unicodedata.normalize('NFC', str(team_name)).strip().lower()

        # Tìm đúng tổ đội trong dữ liệu JSON
        target_team_info = None
        for k, v in dsnc_data.items():
            if unicodedata.normalize('NFC', str(k)).strip().lower() == team_name_norm:
                target_team_info = v
                break

        # Duyệt danh sách nhân công trong tổ để tìm Tổ trưởng
        if target_team_info and "workmen" in target_team_info:
            for worker in target_team_info["workmen"]:
                cong_viec = unicodedata.normalize('NFC', str(worker.get('Công việc', ''))).strip().lower()
                if "tổ trưởng" in cong_viec or cong_viec == "tổ trưởng":
                    ten_dt = str(worker.get('Họ và tên', '')).strip()
                    if ten_dt:
                        return ten_dt

        return "Chưa cập nhật đội trưởng"
        
    except Exception as e:
        print(f"Lỗi khi lấy tên đội trưởng từ JSON: {e}")
        return "N/A"
    
def calculate_deadline(start_date, duration, unit):
    """
    start_date: đối tượng datetime
    duration: số lượng
    unit: 'giờ', 'ngày', 'tháng', 'năm'
    output_format: 'datetime' (trả về ngày+giờ) hoặc 'date' (chỉ trả về ngày)
    """
    if isinstance(start_date, str):
        try:
            start_date = datetime.strptime(start_date, '%d/%m/%Y')
        except ValueError:
            # Nếu định dạng khác, thử xử lý phù hợp hoặc báo lỗi
            raise ValueError("Định dạng ngày bắt đầu phải là DD/MM/YYYY")
    unit_lower = str(unit).strip().lower()
    # 1. Tính toán thời gian
    if unit_lower == 'giờ':
        deadline = start_date + timedelta(hours=int(duration))
        return deadline
    elif unit_lower == 'ngày':
        deadline = start_date + timedelta(days=int(duration))
        return deadline.date()
    elif unit_lower == 'tháng':
        deadline = start_date + relativedelta(months=int(duration))
        return deadline.date()
    elif unit_lower == 'năm':
        deadline = start_date + relativedelta(years=int(duration))
        return deadline.date()
    else:
        raise ValueError("Đơn vị không hợp lệ!")
    return deadline # Trả về đối tượng datetime (mặc định)
def convert_number_to_vietnamese(number):
    # 1. Kiểm tra nếu giá trị là NaN hoặc None (thường gặp trong file Excel)
    if pd.isna(number) or number is None:
        return "không đồng"

    try:
        # 2. Chuyển sang kiểu số nguyên (đảm bảo không có dấu phẩy/chấm)
        # Ép kiểu int để chắc chắn num2words hiểu được
        val = int(float(number))
        
        # 3. Chuyển đổi
        text = num2words(val, lang='vi')
        
        # 4. Trả về định dạng chuẩn
        return f"{text.capitalize()} đồng chẵn"
    except Exception as e:
        # In lỗi ra console để debug nếu vẫn không chạy
        print(f"Lỗi convert_number_to_vietnamese: {e}, giá trị đầu vào: {number}")
        return str(number) # Trả về số dạng chuỗi nếu lỗi
def clone_and_fill_excel(source_path, target_folder, new_filename, name, MST, CCCD, ngaycap, noicap, taikoan, diachicutru, Noilamviecchinh, congviec, totruong):
    # 1. Đảm bảo thư mục đích đã tồn tại
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)
        
    target_path = os.path.join(target_folder, name)
    
    # 2. Clone file (Sao chép file nguồn sang file đích)
    shutil.copy2(source_path, target_path)
    
    # 3. Mở file mới tạo để điền thông tin
    wb = load_workbook(target_path)
    ws = wb.active # Hoặc chọn sheet cụ thể: ws = wb['Sheet1']
    
    # Điền thông tin vào ô A1
    ws['A1'] = "Thông tin mới cần điền"
    
    # 4. Lưu file
    wb.save(target_path)
    print(f"Đã tạo file tại: {target_path}")

def safe_write(ws, cell_address, value):
    cell = ws[cell_address]
    
    # Kiểm tra xem ô đó có nằm trong bất kỳ vùng merge nào không
    for merged_range in ws.merged_cells.ranges:
        if cell_address in merged_range:
            # Nếu có, lấy ô đầu tiên (góc trên bên trái) của vùng gộp đó
            top_left_cell = merged_range.start_cell
            top_left_cell.value = value
            return
            
    # Nếu không nằm trong vùng gộp nào, cứ ghi bình thường
    cell.value = value

def inhopdongnhancong(selected_team=None):
    # 1. Khởi tạo đường dẫn
    template_dir = get_path('templete')
    target_folder_base = get_path(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai))
    
    # Lấy đường dẫn các file JSON
    ttda_path = get_TTDA_JSON()
    dsnc_json_path = get_DSNC_JSON()

    # 2. Đọc thông tin chung dự án từ thong_tin_du_an.JSON
    meta = {
        'ten_cong_ty': "",
        'ten_cong_trinh': "",
        'dia_diem': "",
        'nam': "",
        'thu_nhap_mien_thue': ""
    }
    if os.path.exists(ttda_path):
        try:
            with open(ttda_path, "r", encoding="utf-8") as f:
                proj_data = json.load(f)
                meta = {
                    'ten_cong_ty': proj_data.get("Tên công ty", ""),
                    'ten_cong_trinh': proj_data.get("Tên công trình", ""),
                    'dia_diem': proj_data.get("Địa điểm", ""),
                    'nam': str(proj_data.get("Năm thực hiện", "")),
                    'thu_nhap_mien_thue': str(proj_data.get("Thu nhập năm miễn thuế", ""))
                }
        except Exception as e:
            print(f"Lỗi đọc thông tin dự án: {e}")

    # 3. Đọc danh sách nhân công từ danh_sach_nhan_cong.JSON
    all_workers = []
    if os.path.exists(dsnc_json_path):
        try:
            with open(dsnc_json_path, "r", encoding="utf-8") as f:
                data_json = json.load(f)
                
                # Duyệt qua từng tổ đội trong dictionary JSON
                for to_doi_name, team_info in data_json.items():
                    workmen_list = team_info.get("workmen", [])
                    for worker in workmen_list:
                        # Gắn thêm trường 'Tổ đội' vào từng dictionary của worker để dễ lọc
                        worker_copy = worker.copy()
                        worker_copy['Tổ đội'] = unicodedata.normalize('NFC', to_doi_name.strip())
                        all_workers.append(worker_copy)
        except Exception as e:
            print(f"Lỗi đọc dữ liệu nhân công từ JSON: {e}")

    if not all_workers:
        messagebox.showwarning("Cảnh báo", "Không tìm thấy dữ liệu nhân công hoặc file JSON trống!")
        return

    # 4. Lọc dữ liệu theo tổ đội được chọn
    if selected_team:
        team_name = selected_team.split(" - ")[0].split(" (TT:")[0].strip()
        team_name = unicodedata.normalize('NFC', team_name)
        data_to_print = [w for w in all_workers if w.get('Tổ đội') == team_name]
    else:
        data_to_print = all_workers

    # 5. Setup Jinja2 để render HTML
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template('CK.html')

    # 6. Duyệt và tạo báo cáo PDF cho từng cá nhân
    for row in data_to_print:
        to_doi = str(row.get('Tổ đội', 'Khac')).strip()
        to_doi_path = os.path.join(target_folder_base, to_doi)
        os.makedirs(to_doi_path, exist_ok=True)
        
        ho_ten = row.get('Họ và tên', '')
        mst_raw = str(row.get('MST', '')).strip()
        mst_formatted = mst_raw.zfill(10) if mst_raw else ""
        raw_thu_nhap = str(meta.get('thu_nhap_mien_thue', '')).replace('.', '').replace(',', '').strip()
        thu_nhap_val = int(raw_thu_nhap) if raw_thu_nhap.isdigit() else 0
        # Chuẩn bị dữ liệu cho template
        context = {
            **meta, # Gom các thông tin chung từ dự án
            'ten_cong_ty': meta['ten_cong_ty'], 
            'ho_ten': ho_ten,
            'mst': mst_formatted,
            'cmnd': str(row.get('CMT/CCCD/Hộ chiếu', '')),
            'ngay_cap': str(row.get('Ngày cấp', '')),
            'noi_cap': row.get('Nơi cấp', ''),
            'dia_chi': row.get('Địa chỉ cư trú', ''),
            'noi_lam_viec': row.get('Nơi làm việc chính', ''),
            'dia_diem_ngay': f"{meta['dia_diem']}, ngày ... tháng ... năm {meta['nam']}",
            'thu_nhap_mien_thue': f"{thu_nhap_val:,}".replace(",", " "), 
            'VND': convert_number_to_vietnamese(thu_nhap_val)
        }

        # Render HTML và xuất PDF
        try:
            html_out = template.render(context)
            safe_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in ho_ten])
            pdf_path = os.path.join(to_doi_path, f"Cam_ket_Ca_Nhan_{safe_name}.pdf")
            
            HTML(string=html_out).write_pdf(pdf_path)
            print(f"Đã tạo Cam kết PDF cho: {ho_ten}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi tạo PDF cho {ho_ten}: {str(e)}")
        

def inhopdong_ca_nhan(hoten, mst, cccd):
    # 1. Khởi tạo đường dẫn
    template_dir = get_path('templete')
    target_folder_base = get_path(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai))
    
    ttda_path = get_TTDA_JSON()
    dsnc_json_path = get_DSNC_JSON()

    # 2. Đọc thông tin chung dự án từ thong_tin_du_an.JSON
    meta = {
        'ten_cong_ty': "",
        'ten_cong_trinh': "",
        'dia_diem': "",
        'nam': "",
        'thu_nhap_mien_thue': ""
    }
    if os.path.exists(ttda_path):
        try:
            with open(ttda_path, "r", encoding="utf-8") as f:
                proj_data = json.load(f)
                meta = {
                    'ten_cong_ty': proj_data.get("Tên công ty", ""),
                    'ten_cong_trinh': proj_data.get("Tên công trình", ""),
                    'dia_diem': proj_data.get("Địa điểm", ""),
                    'nam': str(proj_data.get("Năm thực hiện", "")),
                    'thu_nhap_mien_thue': str(proj_data.get("Thu nhập năm miễn thuế", ""))
                }
        except Exception as e:
            print(f"Lỗi đọc thông tin dự án: {e}")

    # Xử lý tính toán số liệu thu nhập miễn thuế an toàn
    raw_thu_nhap = meta['thu_nhap_mien_thue'].replace('.', '').replace(',', '').strip()
    thu_nhap_val = int(raw_thu_nhap) if raw_thu_nhap.isdigit() else 0

    # 3. Đọc danh sách nhân công từ danh_sach_nhan_cong.JSON
    target_worker = None
    found_to_doi = ""
    
    if os.path.exists(dsnc_json_path):
        try:
            with open(dsnc_json_path, "r", encoding="utf-8") as f:
                data_json = json.load(f)
                
                # Chuẩn hóa thông tin đầu vào cần tìm
                hoten_norm = unicodedata.normalize('NFC', hoten).strip()
                mst_str = str(mst).replace('.0', '').strip()
                cccd_str = str(cccd).replace('.0', '').strip()
                
                # Duyệt qua các tổ đội để tìm cá nhân khớp thông tin
                for to_doi_name, team_info in data_json.items():
                    workmen_list = team_info.get("workmen", [])
                    for worker in workmen_list:
                        w_name = unicodedata.normalize('NFC', str(worker.get('Họ và tên', '')).strip())
                        w_mst = str(worker.get('MST', '')).replace('.0', '').strip()
                        w_cccd = str(worker.get('CMT/CCCD/Hộ chiếu', '')).replace('.0', '').strip()
                        
                        if w_name == hoten_norm and w_mst == mst_str and w_cccd == cccd_str:
                            target_worker = worker
                            found_to_doi = unicodedata.normalize('NFC', to_doi_name.strip())
                            break
                    if target_worker:
                        break
        except Exception as e:
            print(f"Lỗi đọc dữ liệu nhân công từ JSON: {e}")

    if not target_worker:
        messagebox.showerror("Lỗi", "Không tìm thấy dữ liệu nhân viên trong hệ thống!")
        return

    # 4. Chuẩn bị dữ liệu cho Template HTML
    raw_mst = str(target_worker.get('MST', '')).replace('.0', '').strip()
    mst_formatted = raw_mst.ljust(10, ' ')

    context = {
        **meta,
        'ten_cong_ty': meta['ten_cong_ty'],
        'ho_ten': target_worker.get('Họ và tên', ''),
        'mst': mst_formatted,
        'cmnd': str(target_worker.get('CMT/CCCD/Hộ chiếu', '')).replace('.0', ''),
        'ngay_cap': str(target_worker.get('Ngày cấp', '')),
        'noi_cap': target_worker.get('Nơi cấp', ''),
        'dia_chi': target_worker.get('Địa chỉ cư trú', ''),
        'noi_lam_viec': target_worker.get('Nơi làm việc chính', ''),
        'dia_diem_ngay': f"{meta['dia_diem']}, ngày ... tháng ... năm {meta['nam']}",
        'thu_nhap_mien_thue': f"{thu_nhap_val:,}".replace(",", " "),
        'VND': convert_number_to_vietnamese(thu_nhap_val) if 'convert_number_to_vietnamese' in globals() else ""
    }

    # 5. Render HTML ra PDF
    to_doi_path = os.path.join(target_folder_base, found_to_doi)
    os.makedirs(to_doi_path, exist_ok=True)
    
    safe_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in hoten])
    output_pdf = os.path.join(to_doi_path, f"Cam_ket_Ca_Nhan_{safe_name}.pdf")
    
    try:
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('CK.html')
        html_out = template.render(context)
        
        HTML(string=html_out).write_pdf(output_pdf)
        print(f"Hợp đồng cá nhân đã tạo: {output_pdf}")
        messagebox.showinfo("Thành công", f"Đã tạo file tại: {output_pdf}")
        
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể tạo file: {str(e)}")

def in_phu_luc_to_doi(team_name):
    # 1. Khởi tạo đường dẫn các file JSON
    ttda_path = get_TTDA_JSON()
    dsnc_json_path = get_DSNC_JSON()

    # Đọc thông tin chung dự án từ thong_tin_du_an.JSON
    meta = {
        'ten_cong_ty': "",
        'ten_cong_trinh': "",
        'dia_diem': ""
    }
    if os.path.exists(ttda_path):
        try:
            with open(ttda_path, "r", encoding="utf-8") as f:
                proj_data = json.load(f)
                meta = {
                    'ten_cong_ty': proj_data.get("Tên công ty", ""),
                    'ten_cong_trinh': proj_data.get("Tên công trình", ""),
                    'dia_diem': proj_data.get("Địa điểm", "")
                }
        except Exception as e:
            print(f"Lỗi đọc thông tin dự án: {e}")

    # 2. Đọc dữ liệu nhân công từ danh_sach_nhan_cong.JSON
    data_json = {}
    if os.path.exists(dsnc_json_path):
        try:
            with open(dsnc_json_path, "r", encoding="utf-8") as f:
                data_json = json.load(f)
        except Exception as e:
            print(f"Lỗi đọc dữ liệu nhân công từ JSON: {e}")

    # Chuẩn hóa tên tổ đội cần tìm
    team_name_norm = unicodedata.normalize('NFC', team_name).strip()
    
    # Tìm kiếm tổ đội trong dữ liệu JSON (so sánh không phân biệt chữ hoa/thường hoặc chuẩn hóa)
    target_team_info = None
    actual_team_name = team_name_norm
    
    for key, val in data_json.items():
        if unicodedata.normalize('NFC', key).strip().lower() == team_name_norm.lower():
            target_team_info = val
            actual_team_name = key.strip()
            break

    if not target_team_info or not target_team_info.get("workmen"):
        messagebox.showwarning("Thông báo", "Tổ này không có nhân viên hoặc không tìm thấy dữ liệu!")
        return

    workmen_list = target_team_info.get("workmen", [])

    # 3. Lọc và xác định Tổ trưởng trong tổ
    dai_dien_to = "Chưa xác định"
    ds_nhan_vien = []

    for worker in workmen_list:
        ho_ten = worker.get('Họ và tên', '')
        cccd_raw = str(worker.get('CMT/CCCD/Hộ chiếu', '')).replace('.0', '').split('.')[0]
        cong_viec = worker.get('Công việc', '')

        # Kiểm tra xem có phải tổ trưởng không (dùng hàm normalize_text nếu có, hoặc chuẩn hóa thông thường)
        cv_normalized = unicodedata.normalize('NFC', str(cong_viec)).strip().lower()
        print (cv_normalized)
        if "tổ trưởng" in cv_normalized and dai_dien_to == "Chưa xác định":
            dai_dien_to = ho_ten

        ds_nhan_vien.append({
            'ho_ten': ho_ten,
            'cccd': cccd_raw,
            'cong_viec': cong_viec
        })

    # 4. Chuẩn bị dữ liệu (Context) cho Template
    context = {
        'ten_cong_ty': meta['ten_cong_ty'],
        'ten_cong_trinh': meta['ten_cong_trinh'],
        'dia_diem': meta['dia_diem'],
        'dai_dien_to': dai_dien_to,
        'danh_sach_nhan_vien': ds_nhan_vien
    }

    # 5. Render PDF
    template_dir = get_path('templete')
    try:
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('phu_luc_to_doi.html')
        html_out = template.render(context)
    except Exception as e:
        messagebox.showerror("Lỗi Template", f"Không tìm thấy hoặc lỗi file template 'phu_luc_to_doi.html': {str(e)}")
        return
    
    # Lưu file kết quả
    to_doi_path = get_path(os.path.join(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai), actual_team_name))
    os.makedirs(to_doi_path, exist_ok=True)
    
    safe_team_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in actual_team_name])
    output_pdf = os.path.join(to_doi_path, f"Phu_Luc_Danh_Sach_to_doi_{safe_team_name}.pdf")
    
    try:
        HTML(string=html_out).write_pdf(output_pdf)
        messagebox.showinfo("Hoàn tất", f"Đã xuất phụ lục cho {actual_team_name}")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể tạo file PDF: {str(e)}")
def get_DSNC_path():
    #current_dir = os.path.dirname(os.path.abspath(__file__))
    #parent_dir = os.path.dirname(current_dir)
    #root_dir = os.path.dirname(parent_dir)
    #dsnc_path = os.path.join(root_dir, 'Data', 'dulieuduan.xlsx')
    nctnpath = globalconfig.folder_nhancongthuengoai
    dsnc_path = get_path (os.path.join("Data", f"{nctnpath}/dulieuduan.xlsx"))    
    return dsnc_path
def get_DSNC_JSON():
    nctnpath = getattr(globalconfig, 'ketoan_nhancongthuengoai', '')
        # Lưu ý: Kiểm tra xem file thực tế bạn đang lưu là danh_sach_to_doi.JSON hay danh_sach_nhan_cong.JSON 
        # (Dựa trên hàm add_new_workman, nó ghi vào danh_sach_nhan_cong.JSON)
    json_path = get_path(os.path.join("Data", nctnpath, "danh_sach_to_doi.JSON"))    
    return json_path
def get_TTDA_JSON():
    ttdapath = getattr(globalconfig, 'ketoan_nhancongthuengoai', '')
        # Lưu ý: Kiểm tra xem file thực tế bạn đang lưu là danh_sach_to_doi.JSON hay danh_sach_nhan_cong.JSON 
        # (Dựa trên hàm add_new_workman, nó ghi vào danh_sach_nhan_cong.JSON)
    json_path = get_path(os.path.join("Data", ttdapath, f"{globalconfig.file_thong_tin_du_an}.JSON"))    
    return json_path
    
def in_hop_dong_giao_khoan_to_truong(team_name):
    current_dir = os.path.dirname(os.path.abspath(__file__)) #Bin/module/file
    parent_dir = os.path.dirname(current_dir)
    root_dir = os.path.dirname(parent_dir)
    
    dsnc_path = get_DSNC_path()
    template_dir = get_path('templete')#os.path.join(root_dir, 'templete')
    
    # 1. Đọc dữ liệu
    df1 = pd.read_excel(dsnc_path, sheet_name='NCList', engine='openpyxl', dtype={'CMT/CCCD/Hộ chiếu': str})
    df2 = pd.read_excel(dsnc_path, sheet_name='Công trình', engine='openpyxl')
    df2.columns = df2.columns.str.strip()
    # Lấy thông tin chung
    meta = {
        'ten_giam_doc': df2.loc[0, "Giám đốc"],
        'chi_huy_truong': df2.loc[0, "Chỉ huy trưởng"],
        'ten_cong_ty': df2.loc[0, "Tên công ty"],
        'ten_cong_trinh': df2.loc[0, "Tên công trình"],
        'dia_diem': df2.loc[0, "Địa điểm"],
        'nam': str(df2.loc[0, "Năm thực hiện"]),
        'gia_tri_khoan': df2.loc[0, "Số tiền khoán nhân công"],
        'VND': convert_number_to_vietnamese(df2.loc[0, "Số tiền khoán nhân công"])
    }
        
    # 2. Tìm Tổ trưởng trong tổ
    df1['Công việc'] = df1['Công việc'].apply(lambda x: unicodedata.normalize('NFC', str(x)))
    df1['Tổ đội'] = df1['Tổ đội'].apply(lambda x: unicodedata.normalize('NFC', str(x)))
    df1['Tổ đội'] = df1['Tổ đội'].str.strip()
    df1['Công việc'] = df1['Công việc'].strip()
    team_name = unicodedata.normalize('NFC', team_name)
    to_truong = df1[(df1['Tổ đội'] == team_name.strip()) & 
                    (df1['Công việc'].str.contains('Tổ trưởng', case=False, na=False))]
    
    if to_truong.empty:
        messagebox.showwarning("Thông báo", f"Không tìm thấy Tổ trưởng trong tổ {team_name}!")
        return
    
    row = to_truong.iloc[0]
    thoi_gian_bat_dau_hop_dong = row['Ngày bắt đầu hợp đồng']
    contract_duration = row['Thời gian hợp đồng']
    print(f"thời gian hợp đồng : {contract_duration}")
    contract_duration_unit = row['Đơn vị thời gian hợp đồng']
    thoi_gian_ket_thuc_hop_dong = calculate_deadline(thoi_gian_bat_dau_hop_dong, contract_duration, contract_duration_unit)
    # 3. Chuẩn bị Context
    context = {
        'ho_ten_to_truong': row['Họ và tên'],
        'Mr_Mrs': row['Giới tính'], 
        'ngay_sinh': row['Ngày sinh'].strftime('%d/%m/%Y') if pd.notnull(row['Ngày sinh']) else "",
        'dia_chi': row['Địa chi cư trú'],
        'cccd': str(row['CMT/CCCD/Hộ chiếu']).replace('.0', ''),
        'ngay_cap': row['Ngày cấp'].strftime('%d/%m/%Y') if pd.notnull(row['Ngày cấp']) else "",
        'noi_cap': row['Nơi cấp'],
        'giam_doc': df2.loc[0, "Giám đốc"],
        'ten_cong_ty': meta['ten_cong_ty'],
        'tien_goi_thau_khoan':int(row['Giá trị khoán']),
        'VND':convert_number_to_vietnamese(int(row['Giá trị khoán'])),
        'thoi_gian_hop_dong': str(int(contract_duration)),
        'bat_dau_hop_dong': thoi_gian_bat_dau_hop_dong,
        'ket_thuc_hop_dong': thoi_gian_ket_thuc_hop_dong,
        'don_vị_thoi_gian_hop_dong': contract_duration_unit,
        'nam_thuc_hien': meta['nam'],
        'so_hop_dong': row['Số hợp đồng'],
        'CCCD': row['CMT/CCCD/Hộ chiếu'],
        'ngay_cap_CCCD': row['Ngày cấp'].strftime('%d/%m/%Y') if pd.notnull(row['Ngày cấp']) else "",
        'noi_cap_CCCD': row['Nơi cấp'],
        'so_tai_khoan': row['Số TK'],
        'ten_ngan_hang': row['Tên ngân hàng'],
        'ten_chu_TK': row['Tên chủ TK']
    }
    
    # 4. Render HTML ra PDF
    target_folder = get_path('Output', globalconfig.ketoan_nhancongthuengoai, team_name)
    os.makedirs(target_folder, exist_ok=True)
    output_pdf = os.path.join(target_folder, f"HD_Giao_Khoan_To_đoi_{team_name}.pdf")
    
    try:
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('hop_dong_giao_khoan_to_truong.html')
        html_out = template.render(context)
        
        HTML(string=html_out).write_pdf(output_pdf)
        messagebox.showinfo("Thành công", f"Đã tạo Hợp đồng cho Tổ trưởng: {output_pdf}")
    except Exception as e:
        messagebox.showerror("Lỗi Render", f"Không thể tạo file: {str(e)}")
        
def nhap_luong_dialog(team_name):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(current_dir)
    root_dir = os.path.dirname(parent_dir) 
    dsnc_path = os.path.join(root_dir, 'Data', 'dulieuduan.xlsx')
    
    target_folder = os.path.join(root_dir, 'Data', team_name)
    os.makedirs(target_folder, exist_ok=True)
     
    json_path = os.path.join(root_dir, 'Data', team_name, f"bangchialuong_{team_name}.json")
    #Đọc dữ liệu
    df1 = pd.read_excel(dsnc_path, sheet_name='NCList', engine='openpyxl', dtype={'CMT/CCCD/Hộ chiếu': str})
    df2 = pd.read_excel(dsnc_path, sheet_name='Công trình', engine='openpyxl')
    nam_cong_trinh = str(df2.loc[0, "Năm thực hiện"])
    #tai danh sach nhan vien tu file dulieuduan.xlsx
    team_data = df1[df1['Tổ đội'] == team_name]
    
    if team_data.empty:
        messagebox.showwarning("Thông báo", "Tổ này không có nhân viên!")
        return
    # Lọc người có công việc là 'Tổ trưởng' trong tổ này
    team_data['Công việc_clean'] = team_data['Công việc'].apply(lambda x: unicodedata.normalize('NFC', str(x)))
    totruong_ = "Tổ trưởng"
    totruong_ = unicodedata.normalize('NFC', totruong_)
    to_truong_row = team_data[team_data['Công việc'].str.contains(totruong_, case=False, na=False)]
    if not to_truong_row.empty:
        dai_dien_to = to_truong_row.iloc[0]['Họ và tên']
    else:
        dai_dien_to = "Chưa xác định" # Hoặc tên mặc định nếu không tìm thấy
    # 2. Chuẩn bị dữ liệu (Context)
    ds_nhan_vien = []
    for _, row in team_data.iterrows():
        ds_nhan_vien.append({
            'ho_ten': row['Họ và tên'],
            'cccd': str(row['CMT/CCCD/Hộ chiếu']).replace('.0', ''),
            'cong_viec': row['Công việc'],
            'STK': row['Số TK'],
            'nganhang': row['Tên ngân hàng'],
            'chu_tk': row['Tên chủ TK']
        })
    
    # Đọc dữ liệu cũ nếu đã tồn tại
    saved_data = {}
    if os.path.exists(json_path) and os.path.getsize(json_path) > 0:
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                saved_data = json.load(f)
        except json.JSONDecodeError:
            # Nếu file lỗi định dạng, reset lại về rỗng
            saved_data = {}
    else:
        # Nếu file không tồn tại hoặc kích thước bằng 0, khởi tạo rỗng
        saved_data = {}
    gia_tri_khoan = float(to_truong_row.iloc[0]['Giá trị khoán'])

    dialog = Toplevel()
    dialog.title(f"Nhập lương: {team_name}")
    #dialog.geometry("520x300") # Thiết lập kích thước mặc định

    # --- KHU VỰC ĐẦU (THÔNG TIN) ---
    header_frame = tk.Frame(dialog)
    header_frame.pack(fill="x", padx=10, pady=5)

    tk.Label(header_frame, text=f"Giá trị khoán: {gia_tri_khoan:,.0f}", font=('Arial', 10, 'bold'), fg="blue").pack()
    lbl_tong = tk.Label(header_frame, text="Tổng thu nhập: 0", font=('Arial', 10, 'bold'))
    lbl_tong.pack()

    # --- KHU VỰC DANH SÁCH (CÓ CUỘN) ---
    canvas = tk.Canvas(dialog)
    scrollbar = tk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
             "<Configure>",
             lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
     )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True, padx=10)
    scrollbar.pack(side="right", fill="y")

    # Grid danh sách bên trong scrollable_frame
    tk.Label(scrollable_frame, text="Họ tên", font=('Arial', 10, 'bold')).grid(row=0, column=0)
    tk.Label(scrollable_frame, text="Thu nhập", font=('Arial', 10, 'bold')).grid(row=0, column=1)

    entries = {}
    def cap_nhat_tong(*args):
       tong = 0
       # entries.values() giờ là danh sách các StringVar
       for var in entries.values():
           val = var.get().replace(',', '') # Lấy giá trị trực tiếp từ StringVar
           if val.replace('.', '').isdigit(): # Kiểm tra số
              tong += float(val)
           lbl_tong.config(text=f"Tổng thu nhập: {tong:,.0f}", 
                    fg="red" if tong > gia_tri_khoan else "green")
    entries = {} # Từ điển này sẽ lưu các StringVar
    for i, nv in enumerate(ds_nhan_vien, start=1):
       tk.Label(scrollable_frame, text=nv['ho_ten']).grid(row=i, column=0, sticky='w')
       #var = tk.StringVar(value=saved_data.get(nv['ho_ten'], "4000000"))
       #var.trace_add("write", cap_nhat_tong)
       #entry = tk.Entry(scrollable_frame, textvariable=var)
       #entry.grid(row=i, column=1, pady=2)
       #entries[nv['ho_ten']] = entry
       # Tạo biến riêng biệt cho từng nhân viên
       # Sử dụng saved_data.get để lấy giá trị
       # 1. Tạo biến StringVar độc lập
       nv_var = tk.StringVar(value=saved_data.get(nv['ho_ten'], "4000000"))    
       # 2. Gắn trace trực tiếp vào biến này
       nv_var.trace_add("write", cap_nhat_tong)   
       # 3. Tạo entry với textvariable là biến này
       entry = tk.Entry(scrollable_frame, textvariable=nv_var)
       entry.grid(row=i, column=1, pady=2)   
       # 4. QUAN TRỌNG: Lưu chính biến 'nv_var' vào từ điển thay vì lưu 'entry'
       entries[nv['ho_ten']] = nv_var
   # Sau khi kết thúc vòng lặp thêm nhân viên, thêm đoạn này:
    dialog.update_idletasks() # Tính toán lại kích thước
    
    # Lấy kích thước yêu cầu của nội dung
    width = dialog.winfo_reqwidth()
    height = dialog.winfo_reqheight()
    
    # Thiết lập kích thước cửa sổ dựa trên nội dung (tối đa là 600px cho chiều cao)
    #dialog.geometry(f"{width}x{min(height, 600)}")
    # --- KHU VỰC NÚT BẤM (DƯỚI CÙNG) ---
    footer_frame = tk.Frame(dialog)
    footer_frame.pack(fill="x", side="bottom", pady=10)
    cap_nhat_tong()

    result = {"confirmed": False}

    def save_and_close():
       # 1. Lưu dữ liệu vào JSON
       data_to_save = {name: entry.get() for name, entry in entries.items()}
       os.makedirs(os.path.dirname(json_path), exist_ok=True)
       with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=4)   
       # 2. Chuẩn bị dữ liệu để in PDF
       list_for_pdf = []
       total_sum = 0
       for nv in ds_nhan_vien:
           luong = float(data_to_save.get(nv['ho_ten'], 0))
           total_sum += luong
           list_for_pdf.append({
               'ho_ten': nv['ho_ten'],
               'luong': "{:,.0f}".format(luong), # Định dạng tiền tệ
               'thong_tin': nv['STK'] + "-" + nv['nganhang'] + "-" + nv['chu_tk'],  # Bạn có thể thêm cột STK vào file Excel để lấy tại đây     
           })
       # 3. Render PDF
       template_dir = os.path.join(root_dir, 'templete') # Đảm bảo file HTML nằm ở đây
       env = Environment(loader=FileSystemLoader(template_dir))
       template = env.get_template('bang_chia_luong.html')
    
       html_out = template.render(
           team_name=team_name,
           danh_sach=list_for_pdf,
           tong_cong="{:,.0f}".format(total_sum),
           ten_nguoi_lap= str(to_truong_row.iloc[0]['Họ và tên']),
           VND = convert_number_to_vietnamese(total_sum),
           nam = nam_cong_trinh           
       )    
       # 4. Lưu file PDF
       output_dir = os.path.join(root_dir, 'Output', team_name)
       os.makedirs(output_dir, exist_ok=True)
       output_pdf = os.path.join(output_dir, f"Bảng chia lương_{team_name}.pdf")   
       HTML(string=html_out).write_pdf(output_pdf)
       result["confirmed"] = True
       dialog.destroy()
       messagebox.showinfo("Thành công", f"Đã xuất file PDF tại: {output_pdf}")

    tk.Button(footer_frame, text="Lưu & Xuất PDF", command=save_and_close, bg="#dddddd").pack()
    cap_nhat_tong()
    dialog.wait_window()
    return result["confirmed"], (json.load(open(json_path, 'r', encoding='utf-8')) if result["confirmed"] else {})
    
def quan_ly_danh_muc_cong_viec(team_name):
    nctnpath = globalconfig.folder_nhancongthuengoai
    folder_path = os.path.join(nctnpath, team_name)
    file_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    
    # Đảm bảo file tồn tại
    if not os.path.exists(folder_path): os.makedirs(folder_path)
    if not os.path.exists(file_path):
        with open(file_path, "w", encoding="utf-8") as f: json.dump([], f)

    dialog = tk.Toplevel()
    dialog.title(f"Khai báo danh mục công việc: {team_name}")
    
    # Bảng hiển thị
    columns = ("cv", "dv", "dg")
    tree = ttk.Treeview(dialog, columns=columns, show="headings")
    tree.heading("cv", text="Tên công việc")
    tree.heading("dv", text="Đơn vị tính")
    tree.heading("dg", text="Đơn giá")
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    def load_data():
        for i in tree.get_children(): tree.delete(i)
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            for item in data:
                tree.insert("", "end", values=(item["cong_viec"], item["don_vi"], item["don_gia"]))
    def them_cv():
        # Gọi hàm tạo ở bước trước hoặc popup tại đây
        # 1. Tạo cửa sổ popup nhập liệu
        popup = tk.Toplevel(dialog)
        popup.title("Thêm công việc mới")
        popup.geometry("300x200")

        tk.Label(popup, text="Tên công việc:").pack(pady=5)
        entry_cv = tk.Entry(popup, width=30)
        entry_cv.pack()

        tk.Label(popup, text="Đơn vị tính:").pack(pady=5)
        entry_dv = tk.Entry(popup, width=30)
        entry_dv.pack()

        tk.Label(popup, text="Đơn giá:").pack(pady=5)
        entry_dg = tk.Entry(popup, width=30)
        entry_dg.pack()
        def luu_moi():
            ten = entry_cv.get()
            dv = entry_dv.get()
            dg = entry_dg.get()
            if not (ten and dv and dg):
               messagebox.showwarning("Lỗi", "Vui lòng nhập đầy đủ thông tin!")
               return

            # 2. Đọc dữ liệu hiện có từ file
            with open(file_path, "r", encoding="utf-8") as f:
                 danh_sach = json.load(f)
            # --- KIỂM TRA TRÙNG LẶP ---
            # Chuyển tên nhập vào thành chữ thường để so sánh
            ten_normalized = ten.lower()
            # Kiểm tra nếu tên công việc đã tồn tại trong danh sách
            da_ton_tai = any(item["cong_viec"].strip().lower() == ten_normalized for item in danh_sach)            
            if da_ton_tai:
                messagebox.showerror("Lỗi", f"Công việc '{ten}' đã tồn tại trong danh sách!")
                return
            # ---------------------------
            # 3. Thêm mới
            danh_sach.append({
                "cong_viec": ten,
                "don_vi": dv,
                "don_gia": float(dg),
            })

            # 4. Ghi lại vào file
            with open(file_path, "w", encoding="utf-8") as f:
                 json.dump(danh_sach, f, ensure_ascii=False, indent=4)

            popup.destroy()
            load_data() # Làm mới lại bảng chính
            messagebox.showinfo("Thành công", "Đã thêm công việc!")

        tk.Button(popup, text="Lưu", command=luu_moi).pack(pady=10)        
    def xoa_cv():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn một công việc để xóa!")
            return
            
        # Lấy thông tin dòng đang chọn
        item_values = tree.item(selected[0])['values']
        ten_cv_can_xoa = item_values[0]
        
        # 1. Hộp thoại cảnh báo xác nhận xóa
        if not messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa công việc: '{ten_cv_can_xoa}' không?"):
            return
            
        # 2. Xóa trong file JSON
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                danh_sach = json.load(f)
            
            # Lọc lại danh sách: giữ lại những cái KHÔNG trùng với tên vừa chọn
            # Lưu ý: Nếu tên công việc không duy nhất, cách này sẽ xóa tất cả các dòng có cùng tên đó.
            # Nếu cần chính xác, bạn nên dùng chỉ số index của Treeview để lọc.
            danh_sach_moi = [item for item in danh_sach if item["cong_viec"] != ten_cv_can_xoa]
            
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(danh_sach_moi, f, ensure_ascii=False, indent=4)
                
            # 3. Làm mới danh sách và thông báo
            load_data()
            messagebox.showinfo("Thành công", f"Đã xóa: {ten_cv_can_xoa}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi xóa dữ liệu: {e}")
    def sua_cv():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn một công việc để sửa!")
            return
        
        # Lấy giá trị hiện tại của dòng đang chọn
        item_values = tree.item(selected[0])['values']
        ten_cu, dv_cu, dg_cu = item_values[0], item_values[1], item_values[2]
        
        # Tạo popup sửa
        popup = tk.Toplevel(dialog)
        popup.title("Sửa công việc")
        popup.geometry("300x250")

        tk.Label(popup, text="Tên công việc:").pack(pady=5)
        entry_cv = tk.Entry(popup, width=30)
        entry_cv.insert(0, ten_cu) # Điền sẵn dữ liệu cũ
        entry_cv.pack()

        tk.Label(popup, text="Đơn vị tính:").pack(pady=5)
        entry_dv = tk.Entry(popup, width=30)
        entry_dv.insert(0, dv_cu)
        entry_dv.pack()

        tk.Label(popup, text="Đơn giá:").pack(pady=5)
        entry_dg = tk.Entry(popup, width=30)
        entry_dg.insert(0, dg_cu)
        entry_dg.pack()

        def luu_sua():
            ten_moi = entry_cv.get().strip()
            dv_moi = entry_dv.get().strip()
            dg_moi = entry_dg.get().strip()

            if not (ten_moi and dv_moi and dg_moi):
                messagebox.showwarning("Lỗi", "Vui lòng nhập đầy đủ thông tin!")
                return

            # Đọc file
            with open(file_path, "r", encoding="utf-8") as f:
                danh_sach = json.load(f)
            
            # Kiểm tra trùng tên (nếu tên mới khác tên cũ)
            if ten_moi.lower() != ten_cu.lower():
                if any(item["cong_viec"].strip().lower() == ten_moi.lower() for item in danh_sach):
                    messagebox.showerror("Lỗi", "Tên công việc đã tồn tại!")
                    return

            # Cập nhật dữ liệu
            for item in danh_sach:
                if item["cong_viec"] == ten_cu:
                    item["cong_viec"] = ten_moi
                    item["don_vi"] = dv_moi
                    item["don_gia"] = dg_moi
                    break
            
            # Ghi lại file
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(danh_sach, f, ensure_ascii=False, indent=4)
            
            popup.destroy()
            load_data()
            messagebox.showinfo("Thành công", "Đã cập nhật công việc!")

        tk.Button(popup, text="Lưu thay đổi", command=luu_sua).pack(pady=10)
    load_data()
    # Nút bấm điều khiển
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(fill="x", pady=5)
    tk.Button(btn_frame, text="Thêm", command=them_cv).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Sửa", command=sua_cv).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Xóa", command=xoa_cv).pack(side="left", padx=5)

def phu_luc_cong_viec_hop_dong(team_name):
    # Đường dẫn file
    folder_path = os.path.join("Data", team_name)
    file_cv_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    file_pl_path = os.path.join(folder_path, f"phuluchopdong_{team_name}_{so_hop_dong}.json")
    path_thong_tin_du_an = get_DSNC_path()

    if not os.path.exists(file_cv_path):
        messagebox.showerror("Lỗi", "Chưa có danh mục công việc cho tổ này!")
        return

    dialog = tk.Toplevel()
    dialog.title(f"Phụ lục hợp đồng: {team_name}")
    dialog.geometry("600x400")

    # Bảng hiển thị (có thêm cột Khối lượng và Thành tiền)
    columns = ("cv", "dv", "dg", "kl", "tt")
    tree = ttk.Treeview(dialog, columns=columns, show="headings")
    tree.heading("cv", text="Công việc"); tree.column("cv", width=150)
    tree.heading("dv", text="ĐVT"); tree.column("dv", width=50)
    tree.heading("dg", text="Đơn giá"); tree.column("dg", width=80)
    tree.heading("kl", text="Khối lượng"); tree.column("kl", width=80)
    tree.heading("tt", text="Thành tiền"); tree.column("tt", width=100)
    tree.pack(fill="both", expand=True, padx=10, pady=10)

    # Dictionary lưu các Entry khối lượng
    entries_kl = {}

    # Load dữ liệu
    with open(file_cv_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        for idx, item in enumerate(data):
            # Thêm dòng vào bảng
            tree.insert("", "end", iid=idx, values=(item["cong_viec"], item["don_vi"], item["don_gia"], "", ""))
            
            # Tạo ô nhập khối lượng (đặt trên Canvas hoặc Tổ trưởng qua logic tính toán)
            # Mẹo: Dùng Entry widget đè lên vị trí cột khối lượng
            entry = tk.Entry(dialog, width=10)
            entries_kl[idx] = (entry, item)

    def luu_phu_luc():
        danh_sach_phu_luc = []
        tong_gia_tri = 0
        
        for idx, (entry, item) in entries_kl.items():
            kl_str = entry.get()
            kl = float(kl_str) if kl_str.replace('.', '').isdigit() else 0
            dg = float(item["don_gia"].replace(',', ''))
            thanh_tien = kl * dg
            tong_gia_tri += thanh_tien
            
            danh_sach_phu_luc.append({
                "cong_viec": item["cong_viec"],
                "khoi_luong": kl,
                "thanh_tien": thanh_tien
            })

        # 1. Lưu vào file JSON phụ lục
        with open(file_pl_path, "w", encoding="utf-8") as f:
            json.dump(danh_sach_phu_luc, f, ensure_ascii=False, indent=4)

        # 2. Ghi giá trị vào thongtinduan.xlsx
        try:
            df = pd.read_excel(path_thong_tin_du_an)
            # Giả định tìm dòng có cột 'Tổ đội' == team_name để cập nhật cột 'Giá trị hợp đồng'
            df['Tổ đội'] = df1['Tổ đội'].apply(lambda x: unicodedata.normalize('NFC', str(x)))
            df['Tổ đội'] = df1['Tổ đội'].strip()
            team_name = unicodedata.normalize('NFC', team_name)
            df.loc[df['Tổ đội'] == team_name.strip(), 'Giá trị khoán'] = tong_gia_tri
            #df.to_excel(path_thong_tin_du_an, index=False)
            messagebox.showinfo("Thành công", f"Đã lưu phụ lục và cập nhật dự án!\nTổng: {tong_gia_tri:,.0f}")
        except Exception as e:
            messagebox.showerror("Lỗi Excel", f"Không thể ghi vào file thontinduan.xlsx: {e}")

    tk.Button(dialog, text="Lưu Phụ lục", command=luu_phu_luc).pack(pady=10)
    
def addcontract(team_name):
    
    folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
    os.makedirs(folder_path, exist_ok=True)
    
    # 1. Tự động sinh số hợp đồng lũy tiến
    nam = datetime.now().year
    files = [f for f in os.listdir(folder_path) if f.startswith(f"hop dong_{team_name}")]
    max_id = 0
    import re
    for f in files:
        # Tìm số đứng trước ký tự "-" hoặc tên năm trong file
        # Ví dụ file: "hop dong_TeamA_0005-2026-HĐGK.json"
        match = re.search(r'_(\d{4})-\d{4}-HĐGK', f)
        if match:
            current_id = int(match.group(1))
            if current_id > max_id:
                max_id = current_id
    
    new_id = max_id + 1
    acronym = "".join([word[0].upper() for word in team_name.split()])
    so_hop_dong = f"{new_id:04d}-{nam}-HĐGK-{acronym}" 
    
    # 2. Tải danh mục công việc
    catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    if not os.path.exists(catalog_path):
        messagebox.showerror("Lỗi", "Chưa có file danh mục công việc!")
        globalconfig.root.deiconify()
        globalconfig.root.focus_force() # Đảm bảo cửa sổ chính được chọn ngay lập tức
        return
        
    with open(catalog_path, "r", encoding="utf-8") as f:
        danh_muc = json.load(f)

    # 3. Giao diện
    dialog = tk.Toplevel(globalconfig.root)
    dialog.attributes("-topmost", True) # Đưa cửa sổ lên trên cùng mọi ứng dụng khác
    dialog.title(f"Tạo Hợp đồng: {so_hop_dong}")
    dialog.geometry("600x600") # Tăng chiều cao lên chút để thoáng hơn

    # Tạo một Padding chung cho dialog
    dialog_padding = {'padx': 15, 'pady': 10}

    # --- 1. Tiêu đề ---
    tk.Label(dialog, text=f"Số hợp đồng: {so_hop_dong}", font=('Arial', 12, 'bold')).pack(pady=10)

    # --- 2. Frame thông tin (Grid để căn chỉnh ngay ngắn) ---
    frame_info = tk.Frame(dialog)
    frame_info.pack(fill="x", **dialog_padding)

    tk.Label(frame_info, text="Bắt đầu:").grid(row=0, column=0, sticky="w")
    ent_start = tk.Entry(frame_info); ent_start.grid(row=0, column=1, padx=5, pady=2)

    tk.Label(frame_info, text="Thời gian kéo dài:").grid(row=1, column=0, sticky="w")
    ent_duration = tk.Entry(frame_info); ent_duration.grid(row=1, column=1, padx=5, pady=2)

    tk.Label(frame_info, text="Đơn vị thời gian:").grid(row=2, column=0, sticky="w")
    combo_unit = ttk.Combobox(frame_info, values=["Giờ", "Ngày", "Tháng", "Năm"], state="readonly")
    combo_unit.grid(row=2, column=1, padx=5, pady=2)
    combo_unit.current(1)

    # --- 3. Frame nhập khối lượng ---
    frame_kl = tk.Frame(dialog)
    frame_kl.pack(fill="x", **dialog_padding)
    tk.Label(frame_kl, text="Nhập KL cho dòng chọn:").pack(side="left")
    ent_kl_input = tk.Entry(frame_kl, width=15)
    ent_kl_input.pack(side="left", padx=5)
    #-----4 Frame nut save hợp đồng
    
    
    
    # Hàm cập nhật KL (giữ nguyên logic của bạn)
    def cap_nhat_kl():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn công việc trong bảng!", parent=dialog)
            return
        val = ent_kl_input.get()
        if not val.replace('.', '', 1).isdigit():
            messagebox.showerror("Lỗi", "Khối lượng phải là số!", parent=dialog)
            return
        item = selected[0]
        cur_vals = list(tree.item(item, "values"))
        cur_vals[3] = val
        tree.item(item, values=cur_vals)
        ent_kl_input.delete(0, tk.END)

    #tk.Button(frame_kl, text="Lưu hợp đồng", command=cap_nhat_kl).pack(side="left")
    tk.Button(frame_kl, text="Cập nhật", command=cap_nhat_kl).pack(side="left")
    
    #tk.Button(button_frame, text="Lưu hợp đồng", bg="#ADD8E6", command=save_contract).pack(side="left")
    #button_frame = tk.Frame(dialog)
    #button_frame.pack(fill="x", **dialog_padding)
    
    #button1 = tk.Button(button_frame, text="Lưu hợp đồng", bg="#ADD8E6", command=save_contract)
    #button1.pack(side="left")
    frame_button = tk.Frame(dialog)
    frame_button.pack(fill="x", **dialog_padding, side="bottom")
    tk.Button(frame_button, text="Lưu hợp đồng", command=cap_nhat_kl).pack(side="left")

    # --- 4. Frame chứa Treeview ---
    tree_frame = tk.LabelFrame(dialog, text="Danh mục công việc", padx=5, pady=5)
    #tree_frame.pack(fill="both", expand=True, **dialog_padding)
    tree_frame.pack(fill="both", expand=True, padx=dialog_padding['padx'], pady=(0, 50))
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
    tree = ttk.Treeview(
        tree_frame,
        columns=("cv", "dv", "dg", "kl"),
        show="headings",
        xscrollcommand=h_scroll.set
    )

    h_scroll.config(command=tree.xview)
    tree.pack(side="top", fill="both", expand=True)
    h_scroll.pack(side="bottom", fill="x")

    tree.heading("cv", text="Cong viec"); tree.column("cv", width=200)
    tree.heading("dv", text="DVT");       tree.column("dv", width=50)
    tree.heading("dg", text="Don gia");   tree.column("dg", width=100)
    tree.heading("kl", text="Khoi luong");tree.column("kl", width=100)

    for cv in danh_muc:
        tree.insert("", "end", values=(cv["cong_viec"], cv["don_vi"], cv["don_gia"], "0"))
    auto_resize_columns(tree)
    
    

    # --- 5. Frame chua 2 nut ---
    #button_frame = tk.Frame(dialog)
    #button_frame.pack(side="bottom", fill="x", padx=5, pady=5)
    # --- Frame nút lưu ---
    
    #tk.Button(
    #    button_frame,
    #    text="Luu hop dong",
    #    bg="#ADD8E6",
    #    command=save_contract
    #).pack(side=tk.LEFT, expand=True, fill="x", padx=5)
    #button1 = tk.Button(button_frame, text="Luu hop dong", bg="#ADD8E6", command=save_contract)
    #button1.grid(row=0, column=0, sticky="ew", padx=5)
    #button_frame.columnconfigure(0, weight=1)
    #tk.Button(
    #    button_frame,
    #    text="Mo danh muc",
    #    bg="#ADD8E6",
     #   command=lambda: hien_thi_danh_sach_hop_dong(team_name)
    #).pack(side=tk.LEFT, expand=True, fill="x", padx=5)

    def on_closing():
        hien_thi_danh_sach_hop_dong(team_name)
        dialog.destroy()        
    dialog.protocol("WM_DELETE_WINDOW", on_closing)
    #dialog.state('zoomed')    
    globalconfig.dlg_addhopdong = dialog
        
def lam_moi_tree_hd():
            print(f"team name from global: {globalconfig.curent_team_name}")
            # Xóa sạch dữ liệu cũ
            tree_hd = globalconfig.tree_hd
            file_path = globalconfig.file_path
            team_name = globalconfig.curent_team_name
            folder_path = os.path.join("Data", globalconfig.ketoan_nhancongthuengoai, team_name)
            
            print(f"team name of local: {team_name}")
            
            for i in tree_hd.get_children():
                tree_hd.delete(i)
            
            # Nạp lại dữ liệu từ folder
            for filename in os.listdir(folder_path):
                if filename.startswith(f"hop dong_{team_name}") and filename.endswith(".json"):
                    file_path = os.path.join(folder_path, filename)
                    with open(file_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    
                    # Node gốc: Hợp đồng
                    parent_id = tree_hd.insert("", "end", text=data["so_hop_dong"], values=(
                        data["bat_dau"], data["time_unit"], data["duration"], f"{data['tong_gia_tri']:,.0f}"
                    ), tags=(file_path,))

                    # Node Phụ lục
                    node_pl_root = tree_hd.insert(parent_id, "end", text="Phụ lục")
                    if "phu_luc" in data:
                        for pl in data["phu_luc"]:
                            tong_pl = sum(item["thanh_tien"] for item in pl["danh_sach_bo_sung"])
                            tree_hd.insert(node_pl_root, "end", text=pl["ten_phu_luc"], values=(
                                pl["ngay_ky"], "-", "-", f"{tong_pl:,.0f}"
                            ))
                    
                    # Node Nghiệm thu (CẬP NHẬT Ở ĐÂY)
                    node_nt_root = tree_hd.insert(parent_id, "end", text="Nghiệm thu")
                    if "nghiem_thu" in data:
                        for i, nt in enumerate(data["nghiem_thu"], start=1):
                            # Tính tổng giá trị của đợt nghiệm thu này
                            tong_nt = sum(ct["thanh_tien_ky_nay"] for ct in nt["chi_tiet"])
                            tree_hd.insert(node_nt_root, "end", text=f"Đợt {i} - {nt['ngay']}", values=(
                                nt["ngay"], "-", "-", f"{tong_nt:,.0f}"
                            ))
                  
                    # Node Thanh toán
                    node_tt_root = tree_hd.insert(parent_id, "end", text="Thanh toán")
                    if "thanh_toan" in data:
                        for i, tt in enumerate(data["thanh_toan"], start=1):
                            # Lấy ngày lập và tổng giá trị thực tế của đợt thanh toán
                            ngay_tt = tt.get("ngay_lap", "N/A")
                            tong_thuc_te = tt.get("so_tien_thuc_te_tong", 0)
                            
                            # Chèn vào treeview
                            tree_hd.insert(node_tt_root, "end", text=f"{tt.get('ten_thanh_toan', f'Đợt {i}')}", values=(
                                ngay_tt, "-", "-", f"{tong_thuc_te:,.0f}"
                            ))
                    
                    # Tự động mở rộng các node
                    tree_hd.item(parent_id, open=True)
                    tree_hd.item(node_pl_root, open=True)
                    tree_hd.item(node_nt_root, open=True)
                    tree_hd.item(node_tt_root, open=True) # Mở rộng luôn node thanh toán  
def clear_container():
    for widget in globalconfig.global_container.winfo_children():
        widget.destroy()
def clean_cell(val):
    """Làm sạch giá trị ô để hiển thị text thuần túy"""
    if pd.isna(val): 
        return ""
    # Chuyển về string, xóa khoảng trắng thừa
    s = str(val).strip()
    
    # Nếu là dạng số float kết thúc bằng .0, loại bỏ .0
    if s.endswith('.0'):
        s = s[:-2]
        
    return s
def print_selected_team():
    #print (f"do the print_selected_team")
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy text của item được chọn (tên Tổ đội)
    team_name = globalconfig.curent_team_name#tree.item(selected_item, 'text')
    
    if messagebox.askyesno("Xác nhận", f"Bạn muốn in toàn bộ cam kết cho tổ: {team_name}?"):
        inhopdongnhancong(selected_team=team_name)
        messagebox.showinfo("Hoàn tất", "Đã in xong Cam kết cho tổ đội!")
def print_team_appendix():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    
    if messagebox.askyesno("Xác nhận", f"In phụ lục danh sách cho: {team_name}?"):
        in_phu_luc_to_doi(team_name)
        #messagebox.showinfo("Thành công", "Đã xuất phụ lục danh sách tổ!") 
def add_hopdong():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
    os.makedirs(folder_path, exist_ok=True)
    catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    if not os.path.exists(catalog_path):
        messagebox.showerror("Lỗi", "Chưa có file danh mục công việc!")
        return
    addcontract(team_name)
    #globalconfig.root = root
    #root.withdraw() 
def print_selected_person():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    
    # Kiểm tra xem người dùng có chọn đúng nhân viên không
    if not tree.parent(selected_item):
        messagebox.showwarning("Thông báo", "Vui lòng chọn một nhân viên cụ thể!")
        return
        
    values = tree.item(selected_item, 'values') # (HoTen, MST, CCCD)
    hoten, mst, cccd = values[0], values[1], values[2]
    
    if messagebox.askyesno("Xác nhận", f"In cam kết cho {hoten} (CCCD: {cccd})?"):
        # Truyền 3 tham số định danh, không cần team_name
        inhopdong_ca_nhan(hoten, mst, cccd)    
def show_custom_menu(event, root_window):
    # 3. Logic xác định item trong Treeview trước tiên
    tree = globalconfig.tree_todoi
    item = tree.identify_row(event.y)

    if not item:
        return  # Không có item thì không hiện menu

    tree.selection_set(item)
    
    # Xác định đây là node cha (Tổ đội) hay node con (Nhân viên)
    is_team_node = (tree.parent(item) == "")
    
    if is_team_node:
        # Lấy trực tiếp tên tổ từ text của node cha
        parent_text = tree.item(item, "text")
        team_name = parent_text.split(" - Tổ trưởng:")[0].strip()
    else:
        # Nếu click vào nhân viên, tìm ngược lên node cha của nó để lấy tên tổ
        parent_id = tree.parent(item)
        parent_text = tree.item(parent_id, "text")
        team_name = parent_text.split(" - Tổ trưởng:")[0].strip()
    
    # Cập nhật luôn vào global để các hàm khác dùng chung nếu cần
    globalconfig.curent_team_name = team_name

    # 1. Tạo menu Toplevel
    menu = ctk.CTkToplevel(root_window)
    menu.overrideredirect(True)
    menu.attributes("-topmost", True)
    menu.geometry(f"+{event.x_root}+{event.y_root}")
    menu.lift()

    # Hàm đóng menu an toàn
    def close_menu(e=None):
        if menu.winfo_exists():
            menu.grab_release()
            menu.destroy()

    # Khung chứa menu
    frame = ctk.CTkFrame(menu, corner_radius=8, fg_color="#FFFFFF", border_width=1, border_color="#D1D1D1")
    frame.pack(padx=2, pady=2)

    # Hàm tạo mục menu chuẩn
    def add_menu_item(text, icon_path, command_func):
        img = None
        if icon_path and os.path.exists(icon_path):
            try:
                img = ctk.CTkImage(light_image=Image.open(icon_path), size=(20, 20))
            except Exception:
                pass
        
        def on_click():
            close_menu()
            # Thực thi tác vụ sau một nhịp nhỏ để UI giải phóng sự kiện cửa sổ
            root_window.after(20, command_func)

        btn = ctk.CTkButton(
            frame, text=text, image=img, compound="left", anchor="w",
            fg_color="transparent", text_color="#333333", hover_color="#E8E8E8",
            height=35, width=220, command=on_click
        )
        btn.pack(fill="x", padx=4, pady=2)

    # Đắp nội dung menu dựa theo loại node
    if is_team_node:
        add_menu_item("Thêm thành viên tổ", get_path("icons/edit.png"), lambda: add_new_workman(root_window, team_name, workman_data=None))
        add_menu_item("In cam kết cho tổ", get_path("icons/print.png"), lambda: print_selected_team())
        add_menu_item("In phụ lục danh sách tổ", get_path("icons/list.png"), lambda: print_team_appendix())
        add_menu_item("Danh mục công việc", get_path("icons/edit.png"), lambda: danh_muc_cong_viec())
        add_menu_item("Tạo hợp đồng khoán", get_path("icons/edit.png"), lambda: mo_dialog_hop_dong(team_name))
        add_menu_item("Danh mục hợp đồng", get_path("icons/edit.png"), lambda: hien_thi_danh_sach_hop_dong(team_name))
    else:
        add_menu_item("In Cam kết cho người này", get_path("icons/edit.png"), lambda: print_selected_person())
        add_menu_item("Sửa thông tin người này", get_path("icons/edit.png"), lambda: sua_thanh_vien_chon())

    # 2. Xử lý click ra ngoài vùng menu để đóng thay vì dùng FocusOut dễ bị lỗi
    def check_click_outside(event):
        try:
            # Kiểm tra xem tọa độ click chuột có nằm trong khung cửa sổ menu không
            x_root, y_root = event.x_root, event.y_root
            m_x, m_y = menu.winfo_rootx(), menu.winfo_rooty()
            m_w, m_h = menu.winfo_width(), menu.winfo_height()
            
            if not (m_x <= x_root <= m_x + m_w and m_y <= y_root <= m_y + m_h):
                close_menu()
        except Exception:
            close_menu()

    # Lắng nghe sự kiện click toàn cục trên cửa sổ gốc để bắt sự kiện bấm ra ngoài
    root_window.bind("<Button-1>", check_click_outside, add="+")
    
    # Gắn sự kiện hủy bind khi menu bị hủy để tránh rò rỉ bộ nhớ
    def on_destroy(e):
        try:
            root_window.unbind("<Button-1>")
        except Exception:
            pass
            
    menu.bind("<Destroy>", on_destroy)
        
def danh_muc_cong_viec():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    team_name = unicodedata.normalize('NFC', team_name)    
    quan_ly_danh_muc_cong_viec(team_name)
def add_hopdong():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    # Lấy tên tổ từ node cha (cột #0)
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
    os.makedirs(folder_path, exist_ok=True)
    catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    if not os.path.exists(catalog_path):
        messagebox.showerror("Lỗi", "Chưa có file danh mục công việc!")
        return
    addcontract(team_name)
    #globalconfig.root = root
    root.withdraw()
def hien_thi_ds_hopdong():
    tree = globalconfig.tree_todoi
    selected_item = tree.selection()[0]
    team_name = tree.item(selected_item, 'text').split(" - ")[0].split(" (TT:")[0].strip()
    hien_thi_danh_sach_hop_dong(team_name)
def cap_nhat_team_name_selected(event=None):
    tree = globalconfig.tree_todoi
    # 1. Kiểm tra xem có mục nào được chọn không (tránh lỗi Index)
    selection = tree.selection()    
    selected_item = selection[0]
    
    # 2. Tìm nút gốc (Tổ đội)
    # Nếu tree.parent() trả về chuỗi rỗng (""), tức là người dùng đã chọn chính Tổ đội.
    # Nếu trả về ID khác, tức là người dùng chọn Nhân viên -> ID cha chính là Tổ đội.
    parent_id = tree.parent(selected_item)
    root_item = parent_id if parent_id != "" else selected_item
    
    # 3. Lấy tên tổ đội từ nút gốc
    item_text = tree.item(root_item, 'text')
    
    # 4. Tách chuỗi lấy tên
    team_name = item_text.split(" - ")[0].split(" (TT:")[0].strip()
    globalconfig.curent_team_name = team_name
    print(f"Đã click chọn team: {team_name}")
    
def load_names():
        clear_container()
        # 1. Cổng kiểm soát license
        is_valid = check_license_and_launch(globalconfig.root, globalconfig.status_label)
        if not is_valid:
            print("Bản quyền không hợp lệ. Đang dừng chương trình...")
            return

        # 2. Tạo UI Treeview
        table_frame = tk.Frame(globalconfig.global_container)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        columns = ('HoTen', 'MST', 'CCCD')
        tree = ttk.Treeview(table_frame, columns=columns, show='tree headings')
        tree.bind("<Button-3>", lambda event: show_custom_menu(event, globalconfig.root))
        tree.heading("#0", text="Tổ đội / Tổ trưởng", anchor="w")
        tree.column("#0", width=250, anchor="w")
        tree.heading('HoTen', text='Họ và tên')
        tree.heading('MST', text='MST')
        tree.heading('CCCD', text='CMT/CCCD/Hộ chiếu')
        tree.column('HoTen', width=150)
        tree.column('MST', width=100)
        tree.column('CCCD', width=150)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.bind("<<TreeviewSelect>>", cap_nhat_team_name_selected)
        
        # Lưu vào globalconfig
        globalconfig.tree_todoi = tree

        # Biến để lưu trữ node cần highlight
        target_node_to_highlight = None
        target_team_name = getattr(globalconfig, 'curent_team_name', '')
        if target_team_name:
            target_team_name = unicodedata.normalize('NFC', target_team_name).strip()

        # 3. Nạp dữ liệu từ file JSON theo cấu trúc Dictionary
        try:
            json_path = get_DSNC_JSON()
            print(f"Đang đọc dữ liệu từ: {json_path}")
            
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                
                # Duyệt qua từng tổ đội (key là tên tổ đội, value chứa "workmen")
                for to_doi, team_info in data.items():
                    workmen_list = team_info.get("workmen", [])
                    
                    # Tìm xem trong danh sách nhân công ai giữ chức "Tổ trưởng"
                    to_truong_name = "Chưa xác định"
                    for worker in workmen_list:
                        cong_viec = normalize_text(str(worker.get("Công việc", "")))
                        if "tổ trưởng" in cong_viec.lower():
                            to_truong_name = worker.get("Họ và tên", "Chưa xác định")
                            break
                    
                    # Tạo node cha cho Tổ đội trên Treeview
                    node_cha = tree.insert("", "end", text=f"{to_doi} - Tổ trưởng: {to_truong_name}", open=True)
                    
                    # Kiểm tra nếu tên tổ đội này khớp với curent_team_name thì lưu lại để highlight sau
                    if target_team_name and unicodedata.normalize('NFC', str(to_doi)).strip().lower() == target_team_name.lower():
                        target_node_to_highlight = node_cha

                    # Chèn các thành viên (workmen) vào node con
                    for worker in workmen_list:
                        tree.insert(node_cha, "end", values=(
                            clean_cell(worker.get('Họ và tên', '')), 
                            clean_cell(worker.get('MST', '')), 
                            clean_cell(worker.get('CMT/CCCD/Hộ chiếu', ''))
                        ))
                
                # Thực hiện highlight và focus vào node tổ đội phù hợp nếu tìm thấy
                if target_node_to_highlight:
                    tree.selection_set(target_node_to_highlight)
                    tree.focus(target_node_to_highlight)
                    tree.see(target_node_to_highlight)
            else:
                print(f"Không tìm thấy file JSON tại: {json_path}")
                
        except Exception as e:
            print(f"Lỗi khi load dữ liệu JSON: {e}")
            messagebox.showerror("Lỗi", f"Không thể tải danh sách tổ đội: {e}")
def show_custom_menu_danh_sach_hop_dong(event, root_window):
    tree = globalconfig.tree_hd
    item = tree.identify_row(event.y)

    if not item:
        return  # Không có item thì không hiện menu

    tree.selection_set(item)
    
    # Hàm phụ trợ: Truy vết ngược lên node gốc để lấy file_path chính xác từ tags
    def get_file_path_from_node(node):
        node_goc = node
        while tree.parent(node_goc) != "":
            node_goc = tree.parent(node_goc)
        tags = tree.item(node_goc, "tags")
        return tags[0] if tags and len(tags) > 0 else getattr(globalconfig, 'file_path', '')

    file_path = get_file_path_from_node(item)

    # Xác định cấp độ node
    parent_id = tree.parent(item)
    parent_text = tree.item(parent_id, "text") if parent_id != "" else ""
    item_text = tree.item(item, "text")
    
    # Phân loại nhóm node
    is_root_contract = (parent_id == "") # Node gốc (Hợp đồng)
    
    # Click trực tiếp vào các node tiêu đề nhóm cha
    is_phu_luc_header = (item_text == "Phụ lục" and parent_id != "")
    is_nghiem_thu_header = (item_text == "Nghiệm thu" and parent_id != "")
    is_thanh_toan_header = (item_text == "Thanh toán" and parent_id != "")

    # Click vào các mục con nằm bên trong nhóm
    is_phu_luc_item = (parent_text == "Phụ lục")
    is_nghiem_thu_item = (parent_text == "Nghiệm thu")
    is_thanh_toan_item = (parent_text == "Thanh toán")

    # 3. Tạo menu Toplevel
    menu = ctk.CTkToplevel(root_window)
    menu.overrideredirect(True)
    menu.attributes("-topmost", True)
    menu.geometry(f"+{event.x_root}+{event.y_root}")
    menu.lift()

    def close_menu(e=None):
        if menu.winfo_exists():
            menu.grab_release()
            menu.destroy()

    frame = ctk.CTkFrame(menu, corner_radius=8, fg_color="#FFFFFF", border_width=1, border_color="#D1D1D1")
    frame.pack(padx=2, pady=2)

    def add_menu_item(text, icon_path, command_func):
        img = None
        if icon_path and os.path.exists(icon_path):
            try:
                img = ctk.CTkImage(light_image=Image.open(icon_path), size=(20, 20))
            except Exception:
                pass
        
        def on_click():
            close_menu()
            root_window.after(20, command_func)

        btn = ctk.CTkButton(
            frame, text=text, image=img, compound="left", anchor="w",
            fg_color="transparent", text_color="#333333", hover_color="#E8E8E8",
            height=35, width=240, command=on_click
        )
        btn.pack(fill="x", padx=4, pady=2)

    # 4. Đắp nội dung menu linh hoạt theo cấp bậc node được click
    if is_root_contract:
        add_menu_item("Sửa hợp đồng", get_path("icons/edit.png"), lambda: sua_hop_dong(item, file_path))
        add_menu_item("In hợp đồng", get_path("icons/print.png"), lambda: in_hop_dong(file_path))
        add_menu_item("Xóa hợp đồng", get_path("icons/delete.png"), lambda: xoa_hop_dong(item))
        add_menu_item("Tạo đề nghị nghiệm thu", get_path("icons/list.png"), lambda: mo_cua_so_nghiem_thu(file_path))
        add_menu_item("Tạo đề nghị thanh toán", get_path("icons/list.png"), lambda: mo_cua_so_thanh_toan(file_path))
        add_menu_item("Thêm phụ lục khối lượng", get_path("icons/edit.png"), lambda: mo_cua_so_phu_luc(file_path))

    elif is_phu_luc_header:
        add_menu_item("Thêm phụ lục khối lượng", get_path("icons/edit.png"), lambda: mo_cua_so_phu_luc(file_path))

    elif is_nghiem_thu_header:
        add_menu_item("Tạo đề nghị nghiệm thu", get_path("icons/list.png"), lambda: mo_cua_so_nghiem_thu(file_path))

    elif is_thanh_toan_header:
        add_menu_item("Tạo đề nghị thanh toán", get_path("icons/list.png"), lambda: mo_cua_so_thanh_toan(file_path))

    elif is_phu_luc_item:
        add_menu_item(f"Sửa phụ lục {item_text}", get_path("icons/edit.png"), lambda: sua_phu_luc(item))
        add_menu_item(f"Xóa phụ lục {item_text}", get_path("icons/delete.png"), lambda: xoa_phu_luc(item))
        add_menu_item(f"In phụ lục {item_text}", get_path("icons/print.png"), lambda: in_phu_luc_bo_sung(item, file_path))

    elif is_nghiem_thu_item:
        add_menu_item(f"In nghiệm thu {item_text}", get_path("icons/print.png"), lambda: in_nghiem_thu_cong_viec(item, file_path))
        if check_last_item(item):
            add_menu_item(f"Sửa nghiệm thu {item_text}", get_path("icons/edit.png"), lambda: sua_nghiem_thu(item))
            add_menu_item(f"Xóa nghiệm thu {item_text}", get_path("icons/delete.png"), lambda: xoa_nghiem_thu(item))

    elif is_thanh_toan_item:
        if check_last_item(item):         
            add_menu_item(f"Sửa thanh toán {item_text}", get_path("icons/edit.png"), lambda: sua_thanh_toan(item))
            add_menu_item(f"Xóa thanh toán {item_text}", get_path("icons/delete.png"), lambda: xoa_thanh_toan(item))
        add_menu_item(f"In thanh toán {item_text}", get_path("icons/print.png"), lambda: in_thanh_toan_chia_thu_nhap(item, file_path))

    # 5. Xử lý click ra ngoài vùng menu để đóng
    def check_click_outside(event):
        try:
            x_root, y_root = event.x_root, event.y_root
            m_x, m_y = menu.winfo_rootx(), menu.winfo_rooty()
            m_w, m_h = menu.winfo_width(), menu.winfo_height()
            
            if not (m_x <= x_root <= m_x + m_w and m_y <= y_root <= m_y + m_h):
                close_menu()
        except Exception:
            close_menu()

    root_window.bind("<Button-1>", check_click_outside, add="+")
    
    def on_destroy(e):
        try:
            root_window.unbind("<Button-1>")
        except Exception:
            pass
            
    menu.bind("<Destroy>", on_destroy)   
def lam_moi_treeview():
        lam_moi_tree_hd()
        return  
def xoa_phu_luc(item):
        tree_hd = globalconfig.tree_hd
        if not messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa phụ lục này?"):
            return
            
        # Lấy thông tin cần thiết để xóa
        ten_pl = tree_hd.item(item, "text")
        node_goc = item
        while tree_hd.parent(node_goc) != "":
            node_goc = tree_hd.parent(node_goc)
        
        file_path = tree_hd.item(node_goc, "tags")[0]
        
        # Đọc, sửa dữ liệu và lưu lại
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            
        data["phu_luc"] = [pl for pl in data["phu_luc"] if pl["ten_phu_luc"] != ten_pl]
        
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
            
        lam_moi_treeview() # Cập nhật lại giao diện
        messagebox.showinfo("Thành công", "Đã xóa phụ lục!")
def xoa_hop_dong(item):
        tree_hd = globalconfig.tree_hd
        if not messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa hợp đồng này?"):
            return
            
        # Lấy thông tin cần thiết để xóa
        ten_hop_dong = tree_hd.item(item, "text")
        node_goc = item
        while tree_hd.parent(node_goc) != "":
            node_goc = tree_hd.parent(node_goc)
        
        file_path = tree_hd.item(node_goc, "tags")[0]
        
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"Đã xóa file: {file_path}")
                # Sau khi xóa file, bạn nên xóa luôn tag trong Treeview (xem tiếp phần 2)
            except PermissionError:
                print("Không thể xóa file vì file đang được mở hoặc không có quyền.")
            except Exception as e:
                print(f"Lỗi khi xóa file: {e}")
        else:
            print("File không tồn tại.")
            
        lam_moi_treeview() # Cập nhật lại giao diện
        messagebox.showinfo("Thành công", f"Đã xóa hợp đồng {ten_hop_dong}!")  
def xoa_nghiem_thu(item):
            tree_hd = globalconfig.tree_hd
            if not messagebox.askyesno("Xác nhận", "Xóa đợt nghiệm thu này?"): return
            
            # index là số thứ tự của đợt nghiệm thu trong node "Nghiệm thu"
            idx = tree_hd.index(item)
            
            with open(globalconfig.file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            if "nghiem_thu" in data and idx < len(data["nghiem_thu"]):
                data["nghiem_thu"].pop(idx)
                with open(globalconfig.file_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                lam_moi_treeview()      
def in_phu_luc_bo_sung(item, file_path):
    tree_hd = globalconfig.tree_hd
    template_dir = get_path('templete')
    team_name = globalconfig.curent_team_name
    
    # 1. Đọc dữ liệu hợp đồng/phụ lục từ file JSON truyền vào
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    # 2. Lấy thông tin chung dự án từ get_TTDA_JSON()
    ttda_path = get_TTDA_JSON()
    meta = {
        'ten_cong_trinh': "",
        'dia_diem': "",
        'giam_doc': ""
    }
    if os.path.exists(ttda_path):
        try:
            with open(ttda_path, "r", encoding="utf-8") as f_meta:
                proj_data = json.load(f_meta)
                meta = {
                    'ten_cong_trinh': proj_data.get("Tên công trình", ""),
                    'dia_diem': proj_data.get("Địa điểm", ""),
                    'giam_doc': proj_data.get("Giám đốc", "")
                }
        except Exception as e:
            print(f"Lỗi đọc thông tin dự án: {e}")

    # 3. Đọc dữ liệu nhân công từ get_DSNC_JSON() để tìm Tổ trưởng của tổ
    dsnc_json_path = get_DSNC_JSON()
    to_truong_data = None
    
    team_name_norm = unicodedata.normalize('NFC', team_name).strip()
    
    if os.path.exists(dsnc_json_path):
        try:
            with open(dsnc_json_path, "r", encoding="utf-8") as f_dsnc:
                dsnc_data = json.load(f_dsnc)
                
                # Tìm đúng tổ đội
                target_team_info = None
                for k, v in dsnc_data.items():
                    if unicodedata.normalize('NFC', k).strip().lower() == team_name_norm.lower():
                        target_team_info = v
                        break
                
                if target_team_info and "workmen" in target_team_info:
                    for worker in target_team_info["workmen"]:
                        cong_viec = unicodedata.normalize('NFC', str(worker.get('Công việc', ''))).strip().lower()
                        if "tổ trưởng" in cong_viec:
                            to_truong_data = worker
                            break
        except Exception as e:
            print(f"Lỗi đọc dữ liệu nhân công: {e}")

    if not to_truong_data:
        messagebox.showwarning("Thông báo", f"Không tìm thấy Tổ trưởng trong tổ {team_name}!")
        return 

    ten_pl = tree_hd.item(item, "text")
    print(f"Tên phụ lục: {ten_pl}")
    print(f"Tên file hợp đồng: {file_path}")

    # Tìm phụ lục cần in
    pl_data = next((pl for pl in data.get("phu_luc", []) if pl.get("ten_phu_luc") == ten_pl), None)
    if not pl_data:
        messagebox.showerror("Lỗi", "Không tìm thấy dữ liệu phụ lục!")
        return

    Ten_phu_luc = ten_pl
    Du_an = meta['ten_cong_trinh']
    Dia_diem = meta['dia_diem']

    def lay_so_hop_dong_tu_phu_luc(item_phu_luc):
        current_node = item_phu_luc
        while tree_hd.parent(current_node) != "":
            current_node = tree_hd.parent(current_node)
        so_hop_dong = tree_hd.item(current_node, "text")
        return so_hop_dong         

    So_hop_dong = lay_so_hop_dong_tu_phu_luc(item)
    
    # Lấy thông tin chi tiết bổ sung đầu tiên (hoặc xử lý theo danh sách nếu cần)
    danh_sach_bs = pl_data.get("danh_sach_bo_sung", [])
    if not danh_sach_bs:
        messagebox.showerror("Lỗi", "Danh sách bổ sung khối lượng trống!")
        return
        
    cong_viec = danh_sach_bs[0].get("cong_viec", "")
    don_vi_tinh = lay_don_vi(team_name_norm, cong_viec)
    khoi_luong = danh_sach_bs[0].get("khoi_luong", 0) 
    don_gia = lay_don_gia(team_name_norm, cong_viec)
    thanh_tien = danh_sach_bs[0].get("thanh_tien", 0) 
    
    Ten_Giao_Thau = meta['giam_doc']
    Ten_Nhan_Thau = to_truong_data.get("Họ và tên", "")

    # 4. Chuẩn bị Context
    context = {
        'Ten_phu_luc': Ten_phu_luc,
        'Du_an': Du_an, 
        'Dia_diem': Dia_diem,
        'So_hop_dong': So_hop_dong,
        'cong_viec': cong_viec,
        'don_vi_tinh': don_vi_tinh,
        'khoi_luong': khoi_luong,
        'don_gia': f"{don_gia:,.0f}".replace(",", " "),
        'thanh_tien': f"{thanh_tien:,.0f}".replace(",", " "),
        'Ten_Giao_Thau': Ten_Giao_Thau,
        'Ten_Nhan_Thau': Ten_Nhan_Thau,
        'VND': convert_number_to_vietnamese(int(thanh_tien)) if thanh_tien else ""            
    }

    # 5. Render file PDF
    target_folder = get_path(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai, team_name_norm))
    os.makedirs(target_folder, exist_ok=True)
    
    # Làm sạch tên file để tránh lỗi hệ điều hành
    safe_pl_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in Ten_phu_luc])
    output_pdf = os.path.join(target_folder, f"Phu_luc_bo_sung_khoi_luong_{So_hop_dong}_{safe_pl_name}.pdf")
    
    try:
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('phu_luc_bo_sung_khoi_luong.html')
        html_out = template.render(context)
        
        HTML(string=html_out).write_pdf(output_pdf)
        messagebox.showinfo("Thành công", f"Đã tạo phụ lục bổ sung khối lượng: {output_pdf}")
    except Exception as e:
        messagebox.showerror("Lỗi Render", f"Không thể tạo file: {str(e)}") 
def sua_phu_luc(item):
        tree_hd = globalconfig.tree_hd
        # 1. Tìm thông tin file cha
        node_goc = item
        while tree_hd.parent(node_goc) != "":
            node_goc = tree_hd.parent(node_goc)
        file_path = tree_hd.item(node_goc, "tags")[0]
        
        # 2. Lấy tên phụ lục và dữ liệu
        ten_pl = tree_hd.item(item, "text")
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Tìm phụ lục cần sửa
        pl_data = next((pl for pl in data["phu_luc"] if pl["ten_phu_luc"] == ten_pl), None)
        if not pl_data:
            messagebox.showerror("Lỗi", "Không tìm thấy dữ liệu phụ lục!")
            return

        # 3. Mở cửa sổ sửa (Truyền thêm pl_data và file_path để lưu lại)
        dialog_window = mo_cua_so_phu_luc(globalconfig.file_path, pl_data)
        globalconfig.root.wait_window(dialog_window)        
        lam_moi_tree_hd()            
def in_nghiem_thu_cong_viec(item, file_path):
    tree_hd = globalconfig.tree_hd
    team_name = globalconfig.curent_team_name
    template_dir = get_path('templete')
    
    # 1. Lấy thông tin từ Treeview
    ten_nghiem_thu = tree_hd.item(item, "text")  # Lấy text của item được chọn
    parent_item = tree_hd.parent(item)            # Lấy ID cha
    grandparent_item = tree_hd.parent(parent_item) # Lấy ID cha của cha
    ten_hop_dong = tree_hd.item(grandparent_item, "text") # Lấy tên HĐ từ cha của cha        
    
    # Lấy dữ liệu nghiệm thu từ file JSON hợp đồng
    idx = tree_hd.index(item)
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    nghiem_thu_data = data["nghiem_thu"][idx]
    
    # 2. Lấy thông tin chung dự án từ get_TTDA_JSON()
    ttda_path = get_TTDA_JSON()
    meta = {
        'ten_cong_ty': "",
        'ten_cong_trinh': "",
        'dia_diem': "",
        'giam_doc': "",
        'chi_huy_truong': "",
        'nam': ""
    }
    if os.path.exists(ttda_path):
        try:
            with open(ttda_path, "r", encoding="utf-8") as f_meta:
                proj_data = json.load(f_meta)
                meta = {
                    'ten_cong_ty': proj_data.get("Tên công ty", ""),
                    'ten_cong_trinh': proj_data.get("Tên công trình", ""),
                    'dia_diem': proj_data.get("Địa điểm", ""),
                    'giam_doc': proj_data.get("Giám đốc", ""),
                    'chi_huy_truong': proj_data.get("Chỉ huy trưởng", ""),
                    'nam': str(proj_data.get("Năm thực hiện", ""))
                }
        except Exception as e:
            print(f"Lỗi đọc thông tin dự án: {e}")

    # 3. Đọc dữ liệu nhân công từ get_DSNC_JSON() để tìm Tổ trưởng
    dsnc_json_path = get_DSNC_JSON()
    to_truong_data = None
    team_name_norm = unicodedata.normalize('NFC', team_name).strip()
    
    if os.path.exists(dsnc_json_path):
        try:
            with open(dsnc_json_path, "r", encoding="utf-8") as f_dsnc:
                dsnc_data = json.load(f_dsnc)
                
                target_team_info = None
                for k, v in dsnc_data.items():
                    if unicodedata.normalize('NFC', k).strip().lower() == team_name_norm.lower():
                        target_team_info = v
                        break
                
                if target_team_info and "workmen" in target_team_info:
                    for worker in target_team_info["workmen"]:
                        cong_viec_w = unicodedata.normalize('NFC', str(worker.get('Công việc', ''))).strip().lower()
                        if "tổ trưởng" in cong_viec_w:
                            to_truong_data = worker
                            break
        except Exception as e:
            print(f"Lỗi đọc dữ liệu nhân công: {e}")

    if not to_truong_data:
        messagebox.showwarning("Thông báo", f"Không tìm thấy Tổ trưởng trong tổ {team_name}!")
        return 

    # 4. Chuẩn bị danh sách chi tiết nghiệm thu
    chi_tiet_full = []
    tong_hd, tong_truoc, tong_nay, tong_luy_ke = 0, 0, 0, 0
    ngay = nghiem_thu_data.get("ngay", "")
    parts = ngay.split('/')
    thang = int(parts[1]) if len(parts) >= 2 and parts[1].isdigit() else ""
    
    for sub_item in nghiem_thu_data.get("chi_tiet", []):
        cong_viec = sub_item["cong_viec"]
        kl_truoc = sub_item.get("luy_ke_truoc", 0)
        kl_nay = sub_item.get("kl_nghiem_thu_ky_nay", 0)
        don_gia = sub_item.get("don_gia", 0)
        
        # Tính toán các giá trị
        kl_luy_ke = kl_truoc + kl_nay
        gia_tri_hd = sub_item.get("kl_hop_dong", 0) * don_gia
        gia_tri_truoc = kl_truoc * don_gia
        gia_tri_nay = sub_item.get("thanh_tien_ky_nay", 0)
        gia_tri_luy_ke = gia_tri_truoc + gia_tri_nay
        don_vi_tinh = lay_don_vi(team_name_norm, cong_viec) 
        
        chi_tiet_full.append({
            "cong_viec": cong_viec,
            "don_vi": don_vi_tinh,
            "kl_hd": sub_item.get("kl_hop_dong", 0),
            "kl_truoc": kl_truoc,
            "kl_nay": kl_nay,
            "kl_luy_ke": kl_luy_ke,
            "don_gia": don_gia,
            "gia_tri_hd": gia_tri_hd,
            "gia_tri_truoc": gia_tri_truoc,
            "gia_tri_nay": gia_tri_nay,
            "gia_tri_luy_ke": gia_tri_luy_ke
        })
        
        tong_hd += gia_tri_hd
        tong_truoc += gia_tri_truoc
        tong_nay += gia_tri_nay
        tong_luy_ke += gia_tri_luy_ke       

    context = {
        'So_hop_dong': ten_hop_dong,
        'Du_an': meta['ten_cong_trinh'],
        'cong_ty': meta['ten_cong_ty'],
        'Nam': meta['nam'],
        'thang': thang,
        'chi_tiet': chi_tiet_full,
        'tong_hd': tong_hd,
        'tong_truoc': tong_truoc,
        'tong_nay': tong_nay,
        'tong_luy_ke': tong_luy_ke,                        
        'Ten_Giam_Doc': meta['giam_doc'],
        'Ten_Chi_Huy_Truong': meta['chi_huy_truong'], 
        'Ten_Dai_Dien_Nhan_Thau': to_truong_data.get("Họ và tên", "")
    }

    # 5. Render PDF
    target_folder = get_path(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai, team_name_norm))
    os.makedirs(target_folder, exist_ok=True)
    
    safe_nt_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in nghiem_thu_data.get('ten_nghiem_thu', '')])
    safe_hd_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in ten_hop_dong])
    output_pdf = os.path.join(target_folder, f"Nghiem_thu_khoi_luong_{safe_nt_name}_{safe_hd_name}.pdf")
       
    try:
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('nghiem_thu_khoi_luong_cong_viec.html')
        html_out = template.render(context)
        
        HTML(string=html_out).write_pdf(output_pdf)
        messagebox.showinfo("Thành công", f"Đã tạo file nghiệm thu: {output_pdf}")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể tạo PDF: {str(e)}")
def in_thanh_toan_chia_thu_nhap(item, file_path):
        tree_hd = globalconfig.tree_hd
        # Lấy thông tin cơ bản
        team_name = globalconfig.curent_team_name
        template_dir = get_path('templete')
        
        ten_thanh_toan = tree_hd.item(item, "text")
        parent_item = tree_hd.parent(item)
        grandparent_item = tree_hd.parent(parent_item)
        ten_hop_dong = tree_hd.item(grandparent_item, "text")
        
        idx = tree_hd.index(item)
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        thanh_toan_record = data["thanh_toan"][idx]
        chi_tiet = thanh_toan_record["chi_tiet_phuc_vu"]
        for ct in chi_tiet:
            ct['tai_khoan'] = get_so_tai_khoan(ct['ten'])        

        # 1. Lấy thông tin chung dự án từ get_TTDA_JSON()
        ttda_path = get_TTDA_JSON()
        meta = {
            'ten_cong_ty': "",
            'nam': "",
            'giam_doc': "",
            'ke_toan_trung': "",
            'nguoi_lap_bieu': ""
        }
        if os.path.exists(ttda_path):
            try:
                with open(ttda_path, "r", encoding="utf-8") as f_meta:
                    proj_data = json.load(f_meta)
                    meta = {
                        'ten_cong_ty': proj_data.get("Tên công ty", ""),
                        'nam': str(proj_data.get("Năm thực hiện", "")),
                        'giam_doc': proj_data.get("Giám đốc", ""),
                        'ke_toan_trung': proj_data.get("Kế toán trưởng", ""),
                        'nguoi_lap_bieu': proj_data.get("Người lập biểu", "")
                    }
            except Exception as e:
                print(f"Lỗi đọc thông tin dự án: {e}")

        # Tính toán các biến cần thiết
        tong_tt_ky_nay = sum(item.get("gia_tri_tt_ky_nay", 0) for item in chi_tiet)
        tong_thuc_te = thanh_toan_record["so_tien_thuc_te_tong"]
        team_name_norm = unicodedata.normalize('NFC', team_name).strip()

        # Context dùng chung cho cả 2 template
        context = {
            "team_name": team_name_norm,
            "ten_thanh_toan": ten_thanh_toan,
            "Nam": meta['nam'],
            "cong_ty": meta['ten_cong_ty'],
            "ngay_lap": thanh_toan_record.get("ngay_lap", ""),
            "chi_tiet": chi_tiet,
            "tong_tt_ky_nay": tong_tt_ky_nay,
            "tong_thuc_te": tong_thuc_te,
            "tong_thuc_te_tong": tong_thuc_te,
            "tong_tien_chu": convert_number_to_vietnamese(tong_thuc_te),
            "giam_doc": meta['giam_doc'],
            "ke_toan": meta['ke_toan_trung'],
            "nguoi_lap": meta['nguoi_lap_bieu'],
            # Truyền hàm để gọi trực tiếp trong template
            "get_so_tai_khoan": get_so_tai_khoan,
            "doc_so_thanh_chu": convert_number_to_vietnamese,
            "ten_doi_truong": get_ten_doi_truong(team_name_norm) 
        }
        
        target_folder = get_path(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai, team_name_norm))
        os.makedirs(target_folder, exist_ok=True)
        
        safe_tt_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in ten_thanh_toan])
        safe_hd_name = "".join([c if c.isalnum() or c in (' ', '_', '-') else '_' for c in ten_hop_dong])

        try:
            env = Environment(loader=FileSystemLoader(template_dir))
            
            # 1. In Bảng Thanh Toán
            template1 = env.get_template('bang_thanh_toan.html')
            html1 = template1.render(context)
            path1 = os.path.join(target_folder, f"Thanh_toan_{safe_hd_name}_{safe_tt_name}.pdf")
            HTML(string=html1).write_pdf(path1)
            
            # 2. In Bảng Chia Thu Nhập
            template2 = env.get_template('bang_chia_thu_nhap.html')
            html2 = template2.render(context)
            path2 = os.path.join(target_folder, f"Chia_thu_nhap_{safe_hd_name}_{safe_tt_name}.pdf")
            HTML(string=html2).write_pdf(path2)
            
            messagebox.showinfo("Thông báo", f"Đã xuất 2 file PDF tại thư mục:\n{target_folder}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo PDF: {str(e)}")

def sua_nghiem_thu(item):
            tree_hd = globalconfig.tree_hd        
            idx = tree_hd.index(item)
            # Bạn cần cập nhật hàm mo_cua_so_nghiem_thu để nhận thêm tham số này
            dialog_window = mo_cua_so_nghiem_thu(globalconfig.file_path, edit_index=idx) 
            globalconfig.root.wait_window(dialog_window)
            lam_moi_tree_hd()                
def sua_thanh_toan(item):
    tree_hd = globalconfig.tree_hd
    # 1. Tìm chỉ số của node thanh toán trong danh sách con của node "Thanh toán"
    parent_id = tree_hd.parent(item)
    all_children = tree_hd.get_children(parent_id)
    index_in_tree = all_children.index(item)
    
    # 2. Mở cửa sổ với index đã xác định
    # Giả sử bạn lưu path trong tag hoặc biến toàn cục
    file_path = globalconfig.file_path 
    dialog_window = mo_cua_so_thanh_toan(globalconfig.file_path, edit_index=index_in_tree)
    globalconfig.root.wait_window(dialog_window)
    lam_moi_tree_hd()
def xoa_thanh_toan(item):
    tree_hd = globalconfig.tree_hd
    if not messagebox.askyesno("Xác nhận", "Bạn có chắc chắn muốn xóa bản ghi thanh toán này?"):
        return

    # 1. Xác định index
    parent_id = tree_hd.parent(item)
    all_children = tree_hd.get_children(parent_id)
    index_to_delete = all_children.index(item)
    
    # 2. Xóa trong dữ liệu (data)
    # Lưu ý: Bạn cần đảm bảo biến 'data' đang chứa dữ liệu mới nhất
    with open(globalconfig.file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    if "thanh_toan" in data and index_to_delete < len(data["thanh_toan"]):
        del data["thanh_toan"][index_to_delete]
        
        # 3. Lưu lại file
        with open(globalconfig.file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
            
        # 4. Làm mới lại cây hiển thị
        lam_moi_tree_hd()
        messagebox.showinfo("Thành công", "Đã xóa thanh toán!")        
def mo_cua_so_phu_luc(file_path, pl_data=None):
        #print(f" mo_cua_so_phu_luc : file hợp đồng được chọn: {file_path}")
        win_pl = tk.Toplevel()
        win_pl.title("Sửa phụ lục" if pl_data else "Thêm phụ lục")
        win_pl.grab_set()
        # Đọc danh mục công việc
        team_name = globalconfig.curent_team_name
        folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
        catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")#file danh mục công việc
        
        try:
            with open(catalog_path, "r", encoding="utf-8") as f:
                danh_muc_db = json.load(f)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không tìm thấy file danh mục: {e}")
            return            
        # Các Widget... (Label, Entry, Combobox như cũ)
        # ... ent_ngay_ky, combo_cv, ent_khoi_luong ...

        # Nếu là chế độ sửa, điền dữ liệu cũ vào
        # 1. PHẢI KHỞI TẠO ENTRY TRƯỚC
        # Chỉ giữ lại các trường cần nhập thủ công
        tk.Label(win_pl, text="Ngày ký (dd/mm/yyyy):").grid(row=0, column=0, padx=5, pady=5)
        ent_ngay_ky = tk.Entry(win_pl)
        ngay_hien_tai = datetime.now().strftime("%d/%m/%Y")           
        ent_ngay_ky.delete(0, tk.END)               
        ent_ngay_ky.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(win_pl, text="Chọn công việc:").grid(row=1, column=0, padx=5, pady=5)
        combo_cv = ttk.Combobox(win_pl, values=[item['cong_viec'] for item in danh_muc_db], state="readonly")
        combo_cv.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(win_pl, text="Khối lượng:").grid(row=2, column=0, padx=5, pady=5)
        ent_khoi_luong = tk.Entry(win_pl)
        ent_khoi_luong.grid(row=2, column=1, padx=5, pady=5)
        if pl_data:
            ent_ngay_ky.insert(0, pl_data["ngay_ky"])
            # Giả định pl_data["danh_sach_bo_sung"][0] là công việc chính
            cv_cu = pl_data["danh_sach_bo_sung"][0]["cong_viec"]
            combo_cv.set(cv_cu)
            ent_khoi_luong.insert(0, str(pl_data["danh_sach_bo_sung"][0]["khoi_luong"]))
        else:
            ent_ngay_ky.insert(0, ngay_hien_tai)
        def luu_phu_luc():
            try:
                # 1. Validate dữ liệu nhập vào
                if not ent_ngay_ky.get().strip():
                    messagebox.showerror("Lỗi", "Vui lòng nhập ngày ký!")
                    return
                ngay_nhap = ent_ngay_ky.get().strip()
                try:
                   # Kiểm tra định dạng và tính hợp lệ của ngày (ví dụ: không có ngày 32 hoặc tháng 13)
                   datetime.strptime(ngay_nhap, '%d/%m/%Y')
                except ValueError:
                   messagebox.showerror("Lỗi", "Ngày ký không hợp lệ! Vui lòng nhập đúng định dạng dd/mm/yyyy.")
                   return
                
                try:
                    khoi_luong_moi = float(ent_khoi_luong.get())
                except ValueError:
                    messagebox.showerror("Lỗi", "Khối lượng phải là số!")
                    return

                # 2. Đọc dữ liệu
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                
                if "phu_luc" not in data:
                    data["phu_luc"] = []

                # 3. Cập nhật dữ liệu
                if pl_data:
                    # Tìm index
                    pl_idx = next((i for i, pl in enumerate(data["phu_luc"]) 
                                 if pl["ten_phu_luc"] == pl_data["ten_phu_luc"]), None)
                    
                    if pl_idx is not None:
                        # Cập nhật thông tin cơ bản
                        data["phu_luc"][pl_idx]["ngay_ky"] = ent_ngay_ky.get()
                        
                        # Cập nhật khối lượng và tính lại thành tiền
                        item_bo_sung = data["phu_luc"][pl_idx]["danh_sach_bo_sung"][0]
                        don_gia = float(item_bo_sung["price"]) # Dùng đơn giá cũ trong file
                        
                        item_bo_sung["khoi_luong"] = khoi_luong_moi
                        item_bo_sung["thanh_tien"] = khoi_luong_moi * don_gia
                    else:
                        messagebox.showerror("Lỗi", "Không tìm thấy phụ lục trong hệ thống!")
                        return
                else:
                    # Logic thêm mới
                    ten_cv = combo_cv.get()
                    if not ten_cv:
                        messagebox.showwarning("Cảnh báo", "Vui lòng chọn công việc!")
                        return
                    
                    # 1. Tìm đơn giá từ danh mục
                    item_danh_muc = next((x for x in danh_muc_db if x["cong_viec"] == ten_cv), None)
                    if item_danh_muc is None:
                        messagebox.showerror("Lỗi", f"Không tìm thấy thông tin đơn giá cho công việc: {ten_cv}")
                        return  # Dừng hàm nếu không tìm thấy dữ liệu
                    don_gia = float(item_danh_muc["don_gia"])
                    #print(f"Đã cập đơn giá: {don_gia}")
                    # 2. Tạo tên phụ lục tự động (tìm số nhỏ nhất còn thiếu)
                    existing_ids = []
                    for pl in data["phu_luc"]:
                        # Trích xuất số từ tên "PLxx"
                        try:
                            # Lấy phần số sau chữ "PL"
                            num = int(pl["ten_phu_luc"].replace("PL", ""))
                            existing_ids.append(num)
                        except:
                            continue
                    
                    # Tìm số nhỏ nhất chưa tồn tại bắt đầu từ 1
                    new_id = 1
                    while new_id in existing_ids:
                        new_id += 1
                    
                    new_ten_pl = f"PL{new_id:03d}"
                    
                    
                    # 3. Tạo cấu trúc phụ lục mới
                    new_pl = {
                        "ten_phu_luc": new_ten_pl,
                        "ngay_ky": ent_ngay_ky.get(),
                        "danh_sach_bo_sung": [{
                            "cong_viec": ten_cv,
                            "khoi_luong": khoi_luong_moi,
                            "price": don_gia,
                            "thanh_tien": khoi_luong_moi * don_gia
                        }]
                    }
                    data["phu_luc"].append(new_pl)

                # 4. Ghi file an toàn
                with open(file_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                
                lam_moi_treeview()
                messagebox.showinfo("Thành công", "Đã lưu thay đổi!")
                win_pl.destroy()

            except Exception as e:
                messagebox.showerror("Lỗi hệ thống", f"Có lỗi xảy ra khi lưu: {e}")

        btn_luu = tk.Button(win_pl, text="Lưu Phụ lục", command=luu_phu_luc)
        btn_luu.grid(row=4, columnspan=2, pady=10) # Sử dụng grid thay cho pack
        return win_pl
def mo_cua_so_nghiem_thu(file_path, edit_index=None):
    import re
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    nghiem_thu_list = data.get("nghiem_thu", [])
    # --- LOGIC TÍNH TÊN ĐỢT ---
    def get_next_ten():
        ds_ten = [nt.get("ten_nghiem_thu", "") for nt in nghiem_thu_list]
        # Tìm các số trong chuỗi "Đợt X"
        max_dot = 0
        for t in ds_ten:
            match = re.search(r'Đợt (\d+)', t)
            if match:
               max_dot = max(max_dot, int(match.group(1)))
        return f"Đợt {max_dot + 1}"                        
            
    # 1. Chuẩn bị dữ liệu
    tong_hop = {}
    # Lấy danh sách gốc (Hợp đồng + Phụ lục)
    ds_nguon = [data.get("danh_sach_chi_tiet", [])] + [pl.get("danh_sach_bo_sung", []) for pl in data.get("phu_luc", [])]
    
    for items in ds_nguon:
        for item in items:
            cv = item["cong_viec"]
            if cv not in tong_hop:
                tong_hop[cv] = {"hd": 0, "luy_ke_truoc": 0, "don_gia": item.get("price", 0)}
            tong_hop[cv]["hd"] += item["khoi_luong"]

    # Tính lũy kế trước (trừ đợt đang sửa nếu có)
    for i, nt in enumerate(data.get("nghiem_thu", [])):
        if edit_index is not None and i == edit_index:
            continue
        for ct in nt.get("chi_tiet", []):
            if ct["cong_viec"] in tong_hop:
                tong_hop[ct["cong_viec"]]["luy_ke_truoc"] += ct.get("kl_nghiem_thu_ky_nay", 0)

    # 2. Giao diện
    win_nt = tk.Toplevel()
    ten_hien_tai = nghiem_thu_list[edit_index]["ten_nghiem_thu"] if edit_index is not None else get_next_ten()
    win_nt.title(f"{ten_hien_tai} - {'Sửa' if edit_index is not None else 'Tạo mới nghiệm thu'}")
    win_nt.grab_set()
    win_nt.geometry("850x450")
    print(f"Tên nghiệm thu hiện tại: {ten_hien_tai}")
    # --- Thêm dòng hướng dẫn tại đây ---
    lbl_instruction = tk.Label(
        win_nt, 
        text="* Click đúp vào ô 'Kỳ này' để thay đổi số liệu", 
        font=("Arial", 10, "italic"), 
        fg="blue"
    )
    lbl_instruction.pack(anchor="w", padx=10, pady=5)
    # 3. Hàm lưu
    def luu():
        chi_tiet_nt = []
        for iid, info in rows.items():
            vals = tree.item(iid, "values")
            kl = float(vals[3])
            if kl > 0:
                chi_tiet_nt.append({
                    "cong_viec": info["cv"],
                    "don_gia": info["don_gia"],
                    "kl_hop_dong": info["hd"],
                    "luy_ke_truoc": info["luy_ke_truoc"],
                    "kl_nghiem_thu_ky_nay": kl,
                    "thanh_tien_ky_nay": kl * info["don_gia"]
                })
        
        ngay_hien_tai = datetime.now().strftime("%d/%m/%Y")

        if edit_index is not None:
            ngay_cu = nghiem_thu_list[edit_index].get("ngay", ngay_hien_tai)
            # Giữ nguyên tên cũ khi sửa
            data["nghiem_thu"][edit_index] = {
                "ten_nghiem_thu": ten_hien_tai, 
                "ngay": ngay_cu, 
                "chi_tiet": chi_tiet_nt
            }
        else:
            # Tạo mới với tên đợt tự động tăng
            data.setdefault("nghiem_thu", []).append({
                "ten_nghiem_thu": get_next_ten(), 
                "ngay": ngay_hien_tai, 
                "chi_tiet": chi_tiet_nt
            })
            
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        messagebox.showinfo("Thành công", "Đã lưu nghiệm thu!")
        lam_moi_treeview()
        win_nt.destroy()        
    #nút lưu nghiệm thu
    frame_button = tk.Frame(win_nt)
    frame_button.pack(fill="x", pady=10, side="bottom")
    frame_button.grid_columnconfigure(0, weight=1)
    # Cấu hình nút bấm với phong cách hiện đại
    btn = tk.Button(
        frame_button, 
        text="Lưu nghiệm thu", 
        command=luu,
        bg="#007AFF",        # Màu xanh dương (hoặc màu thương hiệu của bạn)
        fg="white",          # Chữ trắng
        relief="flat",       # Bỏ viền nổi mặc định
        padx=20,             # Tạo khoảng cách ngang để nút to hơn
        pady=8,              # Tạo khoảng cách dọc
        activebackground="#005bb5", # Màu khi nhấn chuột vào
        activeforeground="white",
        cursor="hand2"       # Hiển thị bàn tay khi di chuột vào
    )
    btn.grid(row=0, column=0)
    #tk.Button(frame_button, text="Lưu nghiệm thu", command=luu).pack(side="left")
    #tk.Button(win_nt, text="Lưu Nghiệm Thu", command=luu).pack(pady=10)
    
    
    cols = ("công việc", "Hợp đồng", "Kỳ trước", "Kỳ này", "Lũy kế")
    tree = ttk.Treeview(win_nt, columns=cols, show="headings")
    for col in cols: tree.heading(col, text=col.upper())
    tree.pack(fill="both", expand=True)

    rows = {}
    # Nạp dữ liệu vào bảng
    for cv, info in tong_hop.items():
        # Nếu đang ở chế độ sửa, lấy giá trị của đợt đó
        kl_ky_nay = 0
        if edit_index is not None:
            for ct in data["nghiem_thu"][edit_index]["chi_tiet"]:
                if ct["cong_viec"] == cv:
                    kl_ky_nay = ct.get("kl_nghiem_thu_ky_nay", 0)
        
        iid = tree.insert("", "end", values=(cv, info["hd"], info["luy_ke_truoc"], kl_ky_nay, info["luy_ke_truoc"] + kl_ky_nay))
        rows[iid] = {"cv": cv, **info}
        
    auto_resize_columns(tree)    

    # Logic nhập liệu (giữ như trước)
    def on_double_click(event):
        item = tree.identify_row(event.y)
        if tree.identify_column(event.x) == "#4":
            # Tạo popup nhập liệu đơn giản
            # Tạo cửa sổ con
            input_win = tk.Toplevel(win_nt)
            input_win.title("Nhập khối lượng kỳ này")

            # 1. Cấu hình để nó thành Dialog OnTop và Modal
            input_win.transient(win_nt) 
            input_win.grab_set()        

            # 2. Tạo nội dung
            entry = tk.Entry(input_win, width=40)
            entry.insert(0, tree.item(item, "values")[3])
            entry.pack(padx=20, pady=20)
            entry.focus_set()

            # 3. Căn giữa cửa sổ con so với cửa sổ cha
            input_win.update_idletasks() # Quan trọng: Phải gọi lệnh này để hệ thống tính toán kích thước trước
            width = input_win.winfo_width()
            height = input_win.winfo_height()

            x = win_nt.winfo_x() + (win_nt.winfo_width() // 2) - (width // 2)
            y = win_nt.winfo_y() + (win_nt.winfo_height() // 2) - (height // 2)

            input_win.geometry(f"+{x}+{y}")
            def save():
                kl = float(entry.get())
                tree.set(item, "Kỳ này", kl)
                tree.set(item, "Lũy kế", rows[item]["luy_ke_truoc"] + kl)
                input_win.destroy()
            tk.Button(input_win, text="OK", command=save).pack()
    tree.bind("<Double-1>", on_double_click)


    return win_nt
def mo_cua_so_thanh_toan(file_path, edit_index=None):
        import re
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        team_name = globalconfig.curent_team_name
        team_name_norm = unicodedata.normalize('NFC', team_name).strip()

        # 1. Lấy danh sách thành viên của tổ đội từ get_DSNC_JSON()
        dsnc_json_path = get_DSNC_JSON()
        team_members = []
        
        if os.path.exists(dsnc_json_path):
            try:
                with open(dsnc_json_path, "r", encoding="utf-8") as f_dsnc:
                    dsnc_data = json.load(f_dsnc)
                    
                    target_team_info = None
                    for k, v in dsnc_data.items():
                        if unicodedata.normalize('NFC', k).strip().lower() == team_name_norm.lower():
                            target_team_info = v
                            break
                    
                    if target_team_info and "workmen" in target_team_info:
                        for worker in target_team_info["workmen"]:
                            w_name = worker.get('Họ và tên', '').strip()
                            if w_name:
                                team_members.append(w_name)
            except Exception as e:
                print(f"Lỗi đọc dữ liệu nhân công từ JSON: {e}")

        # 2. Tính toán hạn mức
        thanh_toan_list = data.get("thanh_toan", [])
        tong_nghiem_thu = sum(c.get("thanh_tien_ky_nay", 0) for nt in data.get("nghiem_thu", []) for c in nt.get("chi_tiet", []))
        print(f"Tổng nghiệm thu: {tong_nghiem_thu}")
        
        tong_da_thanh_toan = sum(
            float(item.get('so_tien_thuc_te', 0)) 
            for i, tt in enumerate(thanh_toan_list) 
            if i != edit_index 
            for item in tt.get('chi_tiet_phuc_vu', [])
        )
        tong_tam_ung = sum(
            float(item.get('khau_tru_tam_ung', 0)) 
            for i, tt in enumerate(thanh_toan_list) 
            if i != edit_index 
            for item in tt.get('chi_tiet_phuc_vu', [])
        )
        tong_khau_tru_thue = sum(
            float(item.get('khau_tru_thue', 0)) 
            for i, tt in enumerate(thanh_toan_list) 
            if i != edit_index 
            for item in tt.get('chi_tiet_phuc_vu', [])
        )
        print(f"Tổng đã thanh toán: {tong_da_thanh_toan}")
        
        tong_con_no_cu = sum(
            float(item.get('con_no_lai', 0)) 
            for i, tt in enumerate(thanh_toan_list) 
            if i != edit_index 
            for item in tt.get('chi_tiet_phuc_vu', [])
        )
        han_muc_toi_da = tong_nghiem_thu - tong_da_thanh_toan - tong_tam_ung - tong_khau_tru_thue

        # --- UI Setup ---
        win = tk.Toplevel()
        win.title(f"Phân bổ thanh toán {team_name_norm}")
        win.grab_set() 
        win.geometry("950x500")

        # Frame thông tin
        frame_info = tk.Frame(win)
        frame_info.pack(fill="x", padx=10, pady=5)
        lbl_han_muc = tk.Label(frame_info, text=f"Hạn mức tối đa: {han_muc_toi_da:,.0f}", font=('Arial', 10, 'bold'))
        lbl_han_muc.pack(side="left")
        lbl_tong_da_nhap = tk.Label(frame_info, text="Tổng TT kỳ này: 0", font=('Arial', 10, 'bold'))
        lbl_tong_da_nhap.pack(side="left", padx=20)
        
        def save_payment():
                try:
                    phan_bo = []
                    tong_thuc_te = 0
                    
                    for name, row in entry_rows.items():
                        v = [float(e.get().replace(',', '') or 0) for e in row["entries"]]
                        thuc_te_str = row["ent_tien_thuc_te"].get().replace(',', '')
                        thuc_te = float(thuc_te_str or 0)
                        
                        if any(x < 0 for x in v) or thuc_te < 0:
                            raise ValueError("Số âm không hợp lệ")
                        
                        con_no = v[1]
                        
                        vals = {
                            "ten": name, 
                            "gia_tri_tt_ky_nay": v[0], 
                            "gia_tri_giu_lai": v[1], 
                            "cong_no_ky_truoc": v[2], 
                            "khau_tru_tam_ung": v[3], 
                            "khau_tru_thue": v[4], 
                            "so_tien_thuc_te": thuc_te, 
                            "con_no_lai": con_no
                        }
                        
                        phan_bo.append(vals)
                        tong_thuc_te += thuc_te
                        
                    if round(tong_thuc_te, 2) > round(han_muc_toi_da, 2):
                       messagebox.showwarning("Cảnh báo", 
                       f"Tổng thực tế ({tong_thuc_te:,.0f}) đã vượt quá hạn mức ({han_muc_toi_da:,.0f})!")
                       return 
                       
                    record = {
                        "ten_thanh_toan": ent_ten_thanh_toan.get(), 
                        "ngay_lap": ent_ngay_lap.get(), 
                        "chi_tiet_phuc_vu": phan_bo, 
                        "so_tien_thuc_te_tong": tong_thuc_te
                    }
                    
                    if edit_index is not None: 
                        data["thanh_toan"][edit_index] = record
                    else: 
                        data.setdefault("thanh_toan", []).append(record)
                    
                    with open(file_path, "w", encoding="utf-8") as f: 
                        json.dump(data, f, ensure_ascii=False, indent=4)
                    
                    messagebox.showinfo("Thành công", "Đã lưu dữ liệu!")
                    win.destroy()
                    lam_moi_tree_hd()
                    
                except Exception as e:
                    print(f"Lỗi lưu dữ liệu: {e}")
                    messagebox.showerror("Lỗi", "Kiểm tra lại dữ liệu nhập (Phải là số)!")

        # Khung tên và ngày
        frame_info_chung = tk.Frame(win)
        frame_info_chung.pack(fill="x", padx=10, pady=5)
        tk.Label(frame_info_chung, text="Tên thanh toán:").pack(side="left")
        ent_ten_thanh_toan = tk.Entry(frame_info_chung)
        ent_ten_thanh_toan.pack(side="left", padx=5)
        tk.Label(frame_info_chung, text="Ngày lập:").pack(side="left", padx=(10, 0))
        ent_ngay_lap = tk.Entry(frame_info_chung)
        ent_ngay_lap.insert(0, datetime.now().strftime("%d/%m/%Y"))
        ent_ngay_lap.pack(side="left", padx=5)

        # Logic gợi ý tên
        if edit_index is None:
            ds_ten = [tt.get("ten_thanh_toan", "") for tt in thanh_toan_list]
            max_lan = max([int(re.search(r'(\d+)', t).group(1)) for t in ds_ten if re.search(r'Thanh toán lần (\d+)', t, re.I)], default=0)
            ent_ten_thanh_toan.insert(0, f"Thanh toán lần {max_lan + 1}")
        else:
            record = thanh_toan_list[edit_index]
            ent_ten_thanh_toan.insert(0, record.get("ten_thanh_toan", ""))
            ent_ngay_lap.delete(0, "end")
            ent_ngay_lap.insert(0, record.get("ngay_lap", ""))

        frame_button = tk.Frame(win, pady=10, bg="lightblue")
        frame_button.pack(fill="x", side="bottom")
        frame_button.columnconfigure(0, weight=1)
        
        btn = tk.Button(
            frame_button, 
            text="Lưu thanh toán", 
            command=save_payment,
            bg="#007AFF", 
            fg="white", 
            relief="flat", 
            padx=20, 
            pady=8, 
            activebackground="#005bb5", 
            activeforeground="white",
            cursor="hand2" 
        )
        btn.pack()
        
        canvas = tk.Canvas(win)
        scrollbar = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
        
        scrollable_frame = tk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Headers
        headers = ["Tên thành viên", "TT kỳ này", "Giữ lại", "Nợ kỳ trước", "Tạm ứng", "Thuế", "Thực tế", "Còn nợ lại"]
        for i, h in enumerate(headers):
            tk.Label(scrollable_frame, text=h, font=("Arial", 10, "bold"), borderwidth=1, relief="solid", bg="lightgray", width=12).grid(row=0, column=i, sticky="nsew")
            
        def get_no_cu_truoc(name):
                for tt in reversed(data.get("thanh_toan", [])):
                    for item in tt.get("chi_tiet_phuc_vu", []):
                        if item.get("ten") == name: 
                            return item.get("con_no_lai", 0)
                return 0

        # Tạo bảng nhập liệu
        entry_rows = {}
        for idx, name in enumerate(team_members, start=1):
            tk.Label(scrollable_frame, text=name, borderwidth=1, relief="solid", width=15).grid(row=idx, column=0)
            
            row_entries = []
            for col in range(1, 7):
                ent = tk.Entry(scrollable_frame, width=12)
                ent.grid(row=idx, column=col, padx=1, pady=1)
                if col < 6:
                    ent.bind('<KeyRelease>', lambda e: update_tien_thuc_te())
                row_entries.append(ent)
                if col == 3 and edit_index is None: 
                    ent.insert(0, str(get_no_cu_truoc(name)))
            
            ent_tien_thuc_te = tk.Entry(scrollable_frame, width=12)
            ent_tien_thuc_te.grid(row=idx, column=6, padx=1, pady=1)
            
            lbl_con_no = tk.Label(scrollable_frame, text="0", borderwidth=1, relief="solid", width=12)
            lbl_con_no.grid(row=idx, column=7, padx=1, pady=1)
            
            entry_rows[name] = {
                "entries": row_entries, 
                "lbl_con_no": lbl_con_no, 
                "ent_tien_thuc_te": ent_tien_thuc_te
            }

        if edit_index is not None:
            for item in thanh_toan_list[edit_index].get("chi_tiet_phuc_vu", []):
                if item["ten"] in entry_rows:
                    en = entry_rows[item["ten"]]["entries"]
                    et = entry_rows[item["ten"]]["ent_tien_thuc_te"] 
                    vals = [item["gia_tri_tt_ky_nay"], item["gia_tri_giu_lai"], item["cong_no_ky_truoc"], item["khau_tru_tam_ung"], item["khau_tru_thue"]]
                    for i in range(5): 
                        en[i].insert(0, str(vals[i]))
                    et.insert(0, str(item["so_tien_thuc_te"]))

        def update_calculations():
            tong_ky_nay = 0
            for name, row in entry_rows.items():
                en = row["entries"]
                v = [float(e.get() or 0) for e in en]
                con_no = v[1]
                row["lbl_con_no"].config(text=f"{con_no:,.0f}")
                tong_ky_nay += v[0]
            lbl_tong_da_nhap.config(text=f"Tổng TT kỳ này: {tong_ky_nay:,.0f}", fg="red" if tong_ky_nay > han_muc_toi_da else "black")

        def update_tien_thuc_te():
            tong_ky_nay = 0
            for name, row in entry_rows.items():
                en = row["entries"]
                try:
                    v = [float(e.get() or 0) for e in en]                    
                    tien_thuc_te = v[0] - v[1] + v[2] - v[3] - v[4]                    
                    row["ent_tien_thuc_te"].delete(0, tk.END)
                    row["ent_tien_thuc_te"].insert(0, f"{tien_thuc_te:,.0f}")
                    con_no = v[1]
                    row["lbl_con_no"].config(text=f"{con_no:,.0f}")
                    tong_ky_nay += v[0]
                except Exception as e:
                    print(f"Lỗi cập nhật tại {name}: {e}")
            lbl_tong_da_nhap.config(text=f"Tổng TT kỳ này: {tong_ky_nay:,.0f}", fg="red" if tong_ky_nay > han_muc_toi_da else "black")
            
        update_calculations() 
        update_tien_thuc_te()

        return win
def cap_nhat_so_hd_selected(event=None):
    team_name = globalconfig.curent_team_name
    tree_hd = globalconfig.tree_hd
    selected = tree_hd.selection()
    if not selected: return
    
    item = selected[0]
    
    # Hàm đệ quy hoặc vòng lặp để tìm lên node gốc nhất
    curr = item
    while tree_hd.parent(curr) != "":
        curr = tree_hd.parent(curr)
    
    # Sau khi lên tới gốc, lấy text (Số HĐ)
    so_hd = tree_hd.item(curr)["text"]
    globalconfig.so_hop_dong_selected = so_hd
    file_name = f"hop dong_{team_name}_{so_hd}.json"
    file_path = os.path.join("Data", globalconfig.ketoan_nhancongthuengoai, team_name, file_name)
    globalconfig.file_path = file_path
    #print(f"số hợp đồng chọn: {so_hd} filepath: {file_path}")  
def in_hop_dong(file_path):
    template_dir = get_path('templete')
    team_name = globalconfig.curent_team_name
    
    # 1. Đọc dữ liệu hợp đồng từ file json truyền vào
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    # 2. Lấy thông tin chung dự án từ get_TTDA_JSON()
    ttda_path = get_TTDA_JSON()
    meta = {
        'ten_cong_ty': "",
        'ten_cong_trinh': "",
        'dia_diem': "",
        'giam_doc': "",
        'nam': ""
    }
    if os.path.exists(ttda_path):
        try:
            with open(ttda_path, "r", encoding="utf-8") as f_meta:
                proj_data = json.load(f_meta)
                meta = {
                    'ten_cong_ty': proj_data.get("Tên công ty", ""),
                    'ten_cong_trinh': proj_data.get("Tên công trình", ""),
                    'dia_diem': proj_data.get("Địa điểm", ""),
                    'giam_doc': proj_data.get("Giám đốc", ""),
                    'nam': str(proj_data.get("Năm thực hiện", ""))
                }
        except Exception as e:
            print(f"Lỗi đọc thông tin dự án: {e}")

    # 3. Đọc dữ liệu nhân công từ get_DSNC_JSON() để tìm Tổ trưởng
    dsnc_json_path = get_DSNC_JSON()
    to_truong_data = None
    
    team_name_norm = unicodedata.normalize('NFC', team_name).strip()
    
    if os.path.exists(dsnc_json_path):
        try:
            with open(dsnc_json_path, "r", encoding="utf-8") as f_dsnc:
                dsnc_data = json.load(f_dsnc)
                
                # Tìm đúng tổ đội
                target_team_info = None
                for k, v in dsnc_data.items():
                    if unicodedata.normalize('NFC', k).strip().lower() == team_name_norm.lower():
                        target_team_info = v
                        break
                
                if target_team_info and "workmen" in target_team_info:
                    for worker in target_team_info["workmen"]:
                        cong_viec = unicodedata.normalize('NFC', str(worker.get('Công việc', ''))).strip().lower()
                        if "tổ trưởng" in cong_viec:
                            to_truong_data = worker
                            break
        except Exception as e:
            print(f"Lỗi đọc dữ liệu nhân công: {e}")

    if not to_truong_data:
        messagebox.showwarning("Thông báo", f"Không tìm thấy Tổ trưởng trong tổ {team_name}!")
        return 

    # 4. Tính toán thông số thời gian và chi tiết
    thoi_gian_bat_dau_hop_dong = data.get("bat_dau")    
    contract_duration = data.get("duration")
    contract_duration_unit = data.get("time_unit")
    so_hop_dong = data.get("so_hop_dong")
    thoi_gian_ket_thuc_hop_dong = calculate_deadline(thoi_gian_bat_dau_hop_dong, contract_duration, contract_duration_unit)
    
    chi_tiet = data.get("danh_sach_chi_tiet", [])
    for ct in chi_tiet:
        ct['don_vi'] = lay_don_vi(team_name, ct['cong_viec'])

    tong_gia_tri_val = float(data.get("tong_gia_tri", 0))

    # 5. Chuẩn bị Context chung cho Template
    context = {
        'ho_ten_to_truong': to_truong_data.get('Họ và tên', ''),
        'Mr_Mrs': to_truong_data.get('Giới tính', ''), 
        'ngay_sinh': str(to_truong_data.get('Ngày sinh', '')),
        'dia_chi': to_truong_data.get('Địa chỉ cư trú', ''),
        'CCCD': str(to_truong_data.get('CMT/CCCD/Hộ chiếu', '')).replace('.0', ''),
        'ngay_cap': str(to_truong_data.get('Ngày cấp', '')),
        'noi_cap': to_truong_data.get('Nơi cấp', ''),
        'giam_doc': meta['giam_doc'],
        'ten_cong_ty': meta['ten_cong_ty'], 
        'ten_cong_trinh': meta['ten_cong_trinh'],
        'dia_diem': meta['dia_diem'],
        'tien_goi_thau_khoan': f"{tong_gia_tri_val:,.0f}".replace(",", " "),
        'VND': convert_number_to_vietnamese(int(tong_gia_tri_val)),
        'tong_gia_tri': data.get("tong_gia_tri"),
        'thoi_gian_hop_dong': str(int(contract_duration)) if contract_duration else "",
        'bat_dau_hop_dong': thoi_gian_bat_dau_hop_dong,
        'ket_thuc_hop_dong': thoi_gian_ket_thuc_hop_dong,
        'don_vị_thoi_gian_hop_dong': contract_duration_unit,
        'nam_thuc_hien': meta['nam'],
        'so_hop_dong': so_hop_dong,
        'so_tai_khoan': to_truong_data.get('Số TK', ''),
        'ten_ngan_hang': to_truong_data.get('Tên ngân hàng', ''),
        'ten_chu_TK': to_truong_data.get('Tên chủ TK', ''),
        'chi_tiet': chi_tiet, 
        'ten_doi_truong': to_truong_data.get('Họ và tên', '')
    }

    # 6. Render HTML ra PDF
    target_folder = get_path(os.path.join('Output', globalconfig.ketoan_nhancongthuengoai, team_name_norm))
    os.makedirs(target_folder, exist_ok=True)
   
    try:
        env = Environment(loader=FileSystemLoader(template_dir))
        
        # In Hợp đồng chính
        template_hd = env.get_template('hop_dong_giao_khoan_to_truong.html')
        html_hd = template_hd.render(context)
        path_hd = os.path.join(target_folder, f"HD_Giao_Khoan_To_doi_{team_name_norm}.pdf")
        HTML(string=html_hd).write_pdf(path_hd)
        
        # In Phụ lục giá
        template_pl = env.get_template('phu_luc_gia.html')
        html_pl = template_pl.render(context)
        path_pl = os.path.join(target_folder, f"Phu_Luc_Gia_{so_hop_dong}_{team_name_norm}.pdf")
        HTML(string=html_pl).write_pdf(path_pl)
        
        messagebox.showinfo("Thành công", "Đã tạo thành công Hợp đồng và Phụ lục giá!")
    except Exception as e:
        messagebox.showerror("Lỗi Render", f"Không thể tạo file: {str(e)}")
def check_last_item(item):
    tree_hd = globalconfig.tree_hd
    parent_id = tree_hd.parent(item)
    children_list = tree_hd.get_children(parent_id)
    if item == children_list[-1]:
        return True
    else:
        return False
                
def hien_thi_danh_sach_hop_dong(team_name):
    print("goi hàm hien_thi_danh_sach_hop_dong")
    #try:
    #   globalconfig.dlg_addhopdong.destroy()
    #except :
    #   pass
       
    #try:
    #    globalconfig.root.withdraw()   
    #except :
    #   passload_names
    clear_container()
    
    #capture the curent team_name and update global variable
    globalconfig.curent_team_name = team_name
    #globalconfig.so_hop_dong_selected
    folder_path = get_path(os.path.join(os.path.join('Data', globalconfig.ketoan_nhancongthuengoai), team_name))
    #list_win = tk.Toplevel()
    #list_win.title(f"Danh sách hợp đồng - {team_name}")
    #list_win.geometry("600x400")

    # Bảng danh sách
    columns = ("so_hd", "bat_dau", "unit", "duration", "gia_tri")
    # Sửa đoạn khởi tạo Treeview
    #tree_hd = ttk.Treeview(list_win, columns=("col1", "col2", "col3", "col4"), show="tree headings")
    style = ttk.Style()
    style.configure("MyCustom.TLabelframe.Label", 
                font=("Helvetica", 14, "bold"), 
                foreground="blue")
    container = ttk.LabelFrame(globalconfig.global_container, text=f"{globalconfig.curent_team_name}", padding=0, style="MyCustom.TLabelframe")
    container.pack(fill="both", expand=True, padx=0, pady=0)
    
    tree_hd = ttk.Treeview(container, columns=("col1", "col2", "col3", "col4"), show="tree headings")
    tree_hd.bind("<Button-3>", lambda event: show_custom_menu_danh_sach_hop_dong(event, globalconfig.root))
    tree_hd.heading("#0", text="Số HĐ / Phụ lục") # Cột chính để phân cấp
    tree_hd.heading("col1", text="Bắt đầu"); tree_hd.heading("col2", text="Đơn vị TG")
    tree_hd.heading("col3", text="Duration"); tree_hd.heading("col4", text="Giá trị")
    tree_hd.column("#0", width=150)
    # THÊM DÒNG NÀY ĐỂ HIỂN THỊ
    tree_hd.pack(fill="both", expand=True, padx=10, pady=10)
    globalconfig.tree_hd = tree_hd

            
    menu_ngu_canh = tk.Menu(globalconfig.root, tearoff=0)  
            
    def handle_right_click(event):
                item = tree_hd.identify_row(event.y)
                if not item: return
                tree_hd.selection_set(item)
                
                menu_ngu_canh.delete(0, "end")
                
                parent_id = tree_hd.parent(item)
                parent_text = tree_hd.item(parent_id, "text") if parent_id != "" else ""
                
                # Logic kiểm tra vị trí
                is_phu_luc = (parent_text == "Phụ lục")
                is_nghiem_thu = (parent_text == "Nghiệm thu")
                is_thanh_toan = (parent_text == "Thanh toán")
                
                if parent_id == "": # Node gốc (Hợp đồng)
                    menu_ngu_canh.add_command(label="Sửa hợp đồng", command=lambda: sua_hop_dong(item, globalconfig.file_path)) # Thêm dòng này
                    menu_ngu_canh.add_command(label="In hợp đồng", command=lambda: in_hop_dong(globalconfig.file_path)) #globalconfig.file_path đã đại diện cho cả số HD
                    menu_ngu_canh.add_command(label="Xóa hợp đồng", command=lambda: xoa_hop_dong(item)) #globalconfig.file_path đã đại diện cho cả số HD
                    menu_ngu_canh.add_command(label="Tạo đề nghị nghiệm thu", command=lambda: mo_cua_so_nghiem_thu(globalconfig.file_path))
                    menu_ngu_canh.add_command(label="Tạo đề nghị thanh toán", command=lambda: mo_cua_so_thanh_toan(globalconfig.file_path))
                    menu_ngu_canh.add_command(label="Thêm phụ lục khối lượng", command=lambda: mo_cua_so_phu_luc(globalconfig.file_path))
                
                elif is_phu_luc:
                    menu_ngu_canh.add_command(label=f"Sửa phụ lục {tree_hd.item(item, "text")}", command=lambda: sua_phu_luc(item))
                    menu_ngu_canh.add_command(label=f"Xóa phụ lục {tree_hd.item(item, "text")}", command=lambda: xoa_phu_luc(item))
                    menu_ngu_canh.add_command(label=f"In phụ lục {tree_hd.item(item, "text")}", command=lambda: in_phu_luc_bo_sung(item, globalconfig.file_path))
                    
                elif is_nghiem_thu:
                    menu_ngu_canh.add_command(label=f"In nghiệm thu {tree_hd.item(item, "text")}", command=lambda: in_nghiem_thu_cong_viec(item, globalconfig.file_path))
                    if check_last_item(item):
                        menu_ngu_canh.add_command(label=f"Sửa nghiệm thu {tree_hd.item(item, "text")}", command=lambda: sua_nghiem_thu(item))
                        menu_ngu_canh.add_command(label=f"Xóa nghiệm thu {tree_hd.item(item, "text")}", command=lambda: xoa_nghiem_thu(item))
                elif is_thanh_toan:
                    if check_last_item(item):                        
                        menu_ngu_canh.add_command(label=f"Sửa thanh toán {tree_hd.item(item, "text")}", command=lambda: sua_thanh_toan(item))
                        menu_ngu_canh.add_command(label=f"Xóa thanh toán {tree_hd.item(item, "text")}", command=lambda: xoa_thanh_toan(item))
                    menu_ngu_canh.add_command(label=f"In thanh toán {tree_hd.item(item, "text")}", command=lambda: in_thanh_toan_chia_thu_nhap(item, globalconfig.file_path))
                menu_ngu_canh.post(event.x_root, event.y_root)
        
    def btn_them_phu_luc_click():
        selected = tree_hd.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn hợp đồng!")
            return
        #team_name = globalconfig.curent_team_name -> team_name có sẵn trong hàm mẹ   
        #so_hop_dong = .globalconfig.so_hop_dong_selected    
        #folder_path = os.path.join("Data", globalconfig.curent_team_name) #Data/team_name
        # Lấy node hiện tại
        current_node = selected[0]
        
        # Tìm lên node gốc (nơi chứa tags là file_path)
        # Nếu node hiện tại là con, parent sẽ là node cha, cứ thế cho đến khi parent là ""
        node_goc = current_node
        while tree_hd.parent(node_goc) != "":
            node_goc = tree_hd.parent(node_goc)
        
        # Lấy tags từ node gốc
        tags = tree_hd.item(node_goc, "tags")
        if not tags:
            messagebox.showerror("Lỗi", "Không tìm thấy thông tin tệp của hợp đồng này.")
            return            
        file_path = tags[0] # Đây chính là đường dẫn file chính xác
        
        #file_path = os.path.join(folder_path, f"hop dong_{team_name}_{so_hop_dong}.json")#file hop dong
        # Đọc danh mục công việc
        catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")#file danh mục công việc
        
        try:
            with open(catalog_path, "r", encoding="utf-8") as f:
                danh_muc_db = json.load(f)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không tìm thấy file danh mục: {e}")
            return

        win_pl = tk.Toplevel()
        win_pl.title("Thêm phụ lục hợp đồng")
        
        # Chỉ giữ lại các trường cần nhập thủ công
        tk.Label(win_pl, text="Ngày ký (dd/mm/yyyy):").grid(row=0, column=0, padx=5, pady=5)
        ent_ngay_ky = tk.Entry(win_pl)
        ngay_hien_tai = datetime.now().strftime("%d/%m/%Y")
        ent_ngay_ky.insert(0, ngay_hien_tai)
        ent_ngay_ky.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(win_pl, text="Chọn công việc:").grid(row=1, column=0, padx=5, pady=5)
        combo_cv = ttk.Combobox(win_pl, values=[item['cong_viec'] for item in danh_muc_db], state="readonly")
        combo_cv.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(win_pl, text="Khối lượng:").grid(row=2, column=0, padx=5, pady=5)
        ent_khoi_luong = tk.Entry(win_pl)
        ent_khoi_luong.grid(row=2, column=1, padx=5, pady=5)

        def luu_phu_luc():
            try:
                # 1. Đọc lại file hợp đồng để lấy số lượng phụ lục mới nhất (phòng trường hợp người khác vừa lưu)
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                
                # --- CẢI TIẾN TẠO TÊN TỰ ĐỘNG ---
                existing_names = [pl.get("ten_phu_luc", "") for pl in data.get("phu_luc", [])]
                
                # Trích xuất các số từ tên: ví dụ "phụ lục_ 001" -> 1
                import re
                existing_numbers = []
                for name in existing_names:
                    match = re.search(r'(\d+)$', name) # Tìm số ở cuối chuỗi
                    if match:
                        existing_numbers.append(int(match.group(1)))
                
                # Tìm số nhỏ nhất chưa xuất hiện (bắt đầu từ 1)
                new_count = 1
                while new_count in existing_numbers:
                    new_count += 1
                
                ten_pl_auto = f"phụ lục_ {new_count:03d}"
                # --------------------------------
                
                # 3. Lấy thông tin công việc
                cv_chon = next((item for item in danh_muc_db if item['cong_viec'] == combo_cv.get()), None)
                if not cv_chon:
                    messagebox.showwarning("Thông báo", "Vui lòng chọn công việc!")
                    return
                
                khoi_luong = float(ent_khoi_luong.get())
                don_gia = float(cv_chon['don_gia'])
                
                # 4. Tạo cấu trúc dữ liệu
                new_item = {
                    "cong_viec": cv_chon['cong_viec'],
                    "khoi_luong": khoi_luong,
                    "price": don_gia,
                    "thanh_tien": khoi_luong * don_gia
                }
                
                if "phu_luc" not in data:
                    data["phu_luc"] = []
                
                data["phu_luc"].append({
                    "ten_phu_luc": ten_pl_auto,
                    "ngay_ky": ent_ngay_ky.get(),
                    "danh_sach_bo_sung": [new_item]
                })
                
                # 5. Ghi lại file
                with open(file_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                # --- CẬP NHẬT MỚI ---
                lam_moi_treeview() # Gọi hàm làm mới để cập nhật giao diện
                # --------------------
                messagebox.showinfo("Thành công", f"Đã thêm {ten_pl_auto} vào hợp đồng!")
                win_pl.destroy()
                
            except Exception as e:
                print(f"file_path: {file_path} , {so_hop_dong}")
                messagebox.showerror("Lỗi", f"Dữ liệu không hợp lệ: {e}")

        tk.Button(win_pl, text="Lưu Phụ lục", command=luu_phu_luc).grid(row=3, columnspan=2, pady=10)  
    # 3. Gán sự kiện cho Treeview
    # Button-3 là chuột phải trên Windows/Linux
    #tree_hd.bind("<Button-3>", handle_right_click)
    # 1. Click chuột trái (hoặc dùng phím điều hướng)
    tree_hd.bind("<<TreeviewSelect>>", cap_nhat_so_hd_selected)
    print(f"hàm hien_thi_danh_sach_hop_dong đọc thư mục folder_path : {folder_path}")
    # Đọc tất cả file hợp đồng trong folder
    for filename in os.listdir(folder_path):
        if filename.startswith(f"hop dong_{team_name}") and filename.endswith(".json"):
            file_path = os.path.join(folder_path, filename)
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            print(f"đọc file hợp đồng: {file_path}")
            # 1. Node Hợp đồng (Gốc)
            parent_id = tree_hd.insert("", "end", text=data["so_hop_dong"], values=(
                data["bat_dau"], data["time_unit"], data["duration"], f"{data['tong_gia_tri']:,.0f}"
            ), tags=(file_path,))

            # 2. Node trung gian "Phụ lục"
            node_pl_root = tree_hd.insert(parent_id, "end", text="Phụ lục")
            
            if "phu_luc" in data:
                for pl in data["phu_luc"]:
                    # Tính tổng tiền của phụ lục đó
                    tong_pl = sum(item["thanh_tien"] for item in pl["danh_sach_bo_sung"])
                    # Thêm vào dưới node "Phụ lục"
                    tree_hd.insert(node_pl_root, "end", text=pl["ten_phu_luc"], values=(
                        pl["ngay_ky"], "-", "-", f"{tong_pl:,.0f}"
                    ))
            # 3. Node trung gian "Nghiệm thu" (Chuẩn bị cho tương lai)
            node_tt_root = tree_hd.insert(parent_id, "end", text="Nghiệm thu")            
            # 4. Node trung gian "Thanh toán" (Chuẩn bị cho tương lai)
            node_tt_root = tree_hd.insert(parent_id, "end", text="Thanh toán")
        # Sau này khi có dữ liệu thanh toán, bạn chỉ cần lặp và chèn vào node_tt_root tương tự như phụ lục
        
    def in_hop_dong_pdf():
        selected = tree_hd.selection()
        if not selected:
            messagebox.showwarning("Thông báo", "Vui lòng chọn một hợp đồng!")
            return
        
        file_path = tree_hd.item(selected[0])["tags"][0]
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Xuất PDF cơ bản
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(200, 10, txt=f"HOP DONG: {data['so_hop_dong']}", ln=True, align='C')
        pdf.set_font("Arial", size=12)
        pdf.ln(10)
        pdf.cell(200, 10, txt=f"Bat dau: {data['bat_dau']} | Duration: {data['duration']}", ln=True)
        pdf.cell(200, 10, txt=f"Tong gia tri: {data['tong_gia_tri']:,.0f} VND", ln=True)
        
        output_name = file_path.replace(".json", ".pdf")
        pdf.output(output_name)
        messagebox.showinfo("Thành công", f"Đã xuất file PDF tại: {output_name}")
      
    def open_root():
        load_names()
        #list_win.destroy()
        #globalconfig.root.deiconify()
        #globalconfig.root.state('zoomed')
        #globalconfig.root.focus_force() # Đảm bảo cửa sổ chính được chọn ngay lập tức
        
        
    ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
    ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"    
    #button_frame = tk.Frame(globalconfig.root)
    button_frame = tk.Frame(globalconfig.global_container)
    button_frame.pack(pady=10)
    
    common_width = 200
    button_1 = ctk.CTkButton(
    master=button_frame, 
    text="Trở lại danh sách tổ đội",
    width=common_width,             # Chiều rộng lớn hơn
    height=45,             # Chiều cao lớn hơn
    corner_radius=10,      # Bo tròn cạnh (giá trị càng lớn càng tròn)
    font=("Arial", 14, "bold"),
    fg_color="#3B8ED0",    # Màu nền chính
    hover_color="#275E8D", # Màu khi di chuột vào (tạo hiệu ứng phản hồi)
    command=open_root
    )   
    button_1.pack(side="left", padx=10) # Lưu ý: Đừng để .pack() sau constructor nếu bạn cần dùng biến button_1 sau này
    
    button_2 = ctk.CTkButton(
    master=button_frame, 
    text="Mở thư mục in",
    width=common_width,             # Chiều rộng lớn hơn
    height=45,             # Chiều cao lớn hơn
    corner_radius=10,      # Bo tròn cạnh (giá trị càng lớn càng tròn)
    font=("Arial", 14, "bold"),
    fg_color="#3B8ED0",    # Màu nền chính
    hover_color="#275E8D", # Màu khi di chuột vào (tạo hiệu ứng phản hồi)
    command=open_output
    )
    button_2.pack(side="left", padx=10)
    
    lam_moi_treeview()
    def on_closing():
        globalconfig.root.deiconify()
        globalconfig.root.focus_force() # Đảm bảo cửa sổ chính được chọn ngay lập tức 
        globalconfig.root.state('zoomed')
        #list_win.destroy()        
    #list_win.protocol("WM_DELETE_WINDOW", on_closing)
    #list_win.state('zoomed')
def sinh_so_hd_tu_dong(team_name):
    import re
    # Định nghĩa thư mục chứa dữ liệu của team
    folder_path = os.path.join("Data", globalconfig.ketoan_nhancongthuengoai, team_name)
    
    # Lấy năm hiện tại
    nam = datetime.now().year
    
    # Kiểm tra xem folder có tồn tại không
    if not os.path.exists(folder_path):
        return f"0001-{nam}-HĐGK"
    
    # Lấy danh sách file hợp đồng của team
    files = [f for f in os.listdir(folder_path) if f.startswith(f"hop dong_{team_name}")]
    
    max_id = 0
    # Sử dụng Regex để trích xuất số thứ tự từ tên file
    # Giả định cấu trúc file là: "hop dong_TeamA_0005-2026-HĐGK.json"
    pattern = re.compile(rf"hop dong_{team_name}_(\d{{4}})-\d{{4}}-HĐGK")
    
    for f in files:
        match = pattern.search(f)
        if match:
            current_id = int(match.group(1))
            if current_id > max_id:
                max_id = current_id
    
    # Tính số hợp đồng tiếp theo và trả về chuỗi định dạng
    new_id = max_id + 1
    return f"{new_id:04d}-{nam}-HĐGK"
def mo_dialog_hop_dong(team_name, data=None, file_path=None):
    folder_path = os.path.join("Data", globalconfig.ketoan_nhancongthuengoai, team_name)
    os.makedirs(folder_path, exist_ok=True)
    # (Giữ nguyên logic sinh số HĐ nếu data is None)
    # Nếu data có sẵn (chế độ sửa), dùng data["so_hop_dong"]
    
    so_hop_dong = data["so_hop_dong"] if data else sinh_so_hd_tu_dong(team_name)
    # 2. Tải danh mục công việc
    catalog_path = os.path.join(folder_path, f"khai báo danh mục công việc_{team_name}.json")
    if not os.path.exists(catalog_path):
        messagebox.showerror("Lỗi", "Chưa có file danh mục công việc!")
        return
    with open(catalog_path, "r", encoding="utf-8") as f:
        danh_muc = json.load(f)
    # 3. Giao diện
    
    dialog = tk.Toplevel()
    dialog.title(f"{'Sửa' if data else 'Tạo'} Hợp đồng: {so_hop_dong}")
    dialog.grab_set()
    # ... (Giữ nguyên code tạo các ô Input và Treeview của bạn) ...
    dialog.geometry("600x650")
    tk.Label(dialog, text=f"Số hợp đồng: {so_hop_dong}", font=('Arial', 10, 'bold')).pack(pady=5)
    # Frame thông tin
    frame_info = tk.Frame(dialog)
    frame_info.pack(padx=10, pady=5)
    tk.Label(frame_info, text="Bắt đầu:").grid(row=0, column=0); ent_start = tk.Entry(frame_info); ent_start.grid(row=0, column=1)
    tk.Label(frame_info, text="Thời gian kéo dài:").grid(row=1, column=0); ent_duration = tk.Entry(frame_info); ent_duration.grid(row=1, column=1)
    #tk.Label(frame_info, text="Đơn vị thời gian:").grid(row=2, column=0); ent_unit = tk.Entry(frame_info); ent_unit.grid(row=2, column=1)
    tk.Label(frame_info, text="Đơn vị thời gian:").grid(row=2, column=0)
    # Định nghĩa danh sách các đơn vị
    danh_sach_don_vi = ["Giờ", "Ngày", "Tháng", "Năm"]
    # Tạo Combobox
    combo_unit = ttk.Combobox(frame_info, values=danh_sach_don_vi, state="readonly")
    combo_unit.grid(row=2, column=1)
    # Mặc định chọn giá trị đầu tiên (ví dụ: "Ngày")
    combo_unit.current(1)
    
    # --- BỔ SUNG: Frame nhập khối lượng ---
    frame_kl = tk.Frame(dialog)
    frame_kl.pack(pady=5)
    tk.Label(frame_kl, text="Nhập KL cho dòng chọn:").pack(side="left")
    ent_kl_input = tk.Entry(frame_kl, width=15)
    ent_kl_input.pack(side="left", padx=5)
    def cap_nhat_kl():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn công việc trong bảng!")
            return
        val = ent_kl_input.get()
        if not val.replace('.', '', 1).isdigit():
            messagebox.showerror("Lỗi", "Khối lượng phải là số!")
            return
        
        # Cập nhật vào cột thứ 4 (index 3)
        item = selected[0]
        cur_vals = list(tree.item(item, "values"))
        cur_vals[3] = val
        tree.item(item, values=cur_vals)
        ent_kl_input.delete(0, tk.END)
    

    #tk.Button(frame_kl, text="Cập nhật", command=cap_nhat_kl).pack(side="left")
    # --- 1. CẤU HÌNH TOÀN CỤC (Đặt ở đây là tốt nhất) ---
    ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
    ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"
    # SỬA Ở ĐÂY:
    button_kl = ctk.CTkButton(
        master=frame_kl,      # 1. Đổi master thành frame_kl
        text="Nhập",
        width=70,
        height=25,
        corner_radius=10,
        font=("Arial", 10, "bold"),
        fg_color="#FF9800",
        hover_color="#CC7A00",
        command=cap_nhat_kl
    )
    # 2. Pack nó cùng phía với các widget khác
    button_kl.pack(side="left", padx=10)
    
    #tạo frame chứ nút save
    def save_contract():
        # ... (Giữ nguyên logic kiểm tra dữ liệu của bạn) ...
            # 1. Kiểm tra để trống hoặc sai định dạng ngày
            start_date = ent_start.get().strip()
            if not start_date:
                messagebox.showerror("Lỗi", "Vui lòng nhập ngày bắt đầu!")
                return
            if not is_valid_date(start_date):
                messagebox.showerror("Lỗi", "Ngày bắt đầu phải có định dạng dd/mm/yyyy và là ngày hợp lệ!")
                return

            # 2. Kiểm tra Thời gian kéo dài (phải là số nguyên/thực dương)
            duration_val = ent_duration.get().strip()
            try:
                duration = float(duration_val)
                if duration <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Lỗi", "Thời gian kéo dài phải là số dương!")
                return

            # 1. Kiểm tra dữ liệu và tính toán danh sách chi tiết mới
            tong_gia_tri = 0
            chi_tiet = []
            for child in tree.get_children():
                values = tree.item(child)["values"]
                kl = float(values[3])
                if kl < 0:
                    messagebox.showerror("Lỗi", "khối lượng phải là số >= 0!")
                    return
                dg = float(values[2].replace(",", ""))
                thanh_tien = kl * dg
                tong_gia_tri += thanh_tien
                chi_tiet.append({"cong_viec": values[0], "khoi_luong": kl, "price": dg, "thanh_tien": thanh_tien})
            if tong_gia_tri <= 0:
                    messagebox.showerror("Lỗi", "Tổng giá trị phải là số Dương!")
                    return
            # 2. Xử lý dữ liệu để lưu
            if data:
                # Chế độ Sửa: Cập nhật các trường hợp đồng vào data cũ
                contract_data = data  # Lấy dữ liệu hiện tại (bao gồm cả phu_luc, nghiem_thu)
                contract_data.update({
                    "bat_dau": ent_start.get(),
                    "time_unit": combo_unit.get(),
                    "duration": float(ent_duration.get()),
                    "tong_gia_tri": tong_gia_tri,
                    "danh_sach_chi_tiet": chi_tiet
                })
            else:
                # Chế độ Tạo mới: Tạo cấu trúc mới (không có phụ lục/nghiệm thu ban đầu)
                contract_data = {
                    "so_hop_dong": so_hop_dong,
                    "bat_dau": ent_start.get(),
                    "time_unit": combo_unit.get(),
                    "duration": float(ent_duration.get()),
                    "tong_gia_tri": tong_gia_tri,
                    "danh_sach_chi_tiet": chi_tiet,
                    "phu_luc": [],      # Khởi tạo rỗng
                    "nghiem_thu": []    # Khởi tạo rỗng
                }
            
            # 3. Ghi file
            path = file_path if file_path else os.path.join(folder_path, f"hop dong_{team_name}_{so_hop_dong}.json")
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(contract_data, f, ensure_ascii=False, indent=4)
                messagebox.showinfo("OK", "Đã lưu hợp đồng và giữ nguyên phụ lục/nghiệm thu!")
                dialog.destroy()
                #lam_moi_tree_hd()
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu: {e}")    
    frame_button = tk.Frame(dialog)
    frame_button.pack(fill="x", side="bottom")
    # --- 1. CẤU HÌNH TOÀN CỤC (Đặt ở đây là tốt nhất) ---
    ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
    ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"
    #tk.Button(dialog, text="Lưu", command=save_contract).pack()
    button = ctk.CTkButton(
    master=frame_button, 
    text="Lưu",
    width=150,             # Chiều rộng lớn hơn
    height=45,             # Chiều cao lớn hơn
    corner_radius=20,      # Bo tròn cạnh (giá trị càng lớn càng tròn)
    font=("Arial", 14, "bold"),
    fg_color="#3B8ED0",    # Màu nền chính
    hover_color="#275E8D", # Màu khi di chuột vào (tạo hiệu ứng phản hồi)
    command=save_contract
    )
    button.pack(pady=20)
    
# 1. Tạo Frame chứa cả Treeview và Scrollbar để dễ Tổ trưởng bố cục
    tree_frame = tk.Frame(dialog)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

    # 2. Tạo Scrollbar ngang
    h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
    
    # 3. Tạo Treeview, gắn xscrollcommand vào scrollbar
    tree = ttk.Treeview(tree_frame, columns=("cv", "dv", "dg", "kl"), 
                        show="headings", xscrollcommand=h_scroll.set)
    
    # 4. Gắn scrollbar vào treeview
    h_scroll.config(command=tree.xview)
    
    # 5. Đặt các thành phần vào Frame
    tree.pack(side="top", fill="both", expand=True)
    h_scroll.pack(side="bottom", fill="x")
    tree.heading("cv", text="Công việc"); tree.heading("dv", text="ĐVT")
    tree.heading("dg", text="Đơn giá"); tree.heading("kl", text="Khối lượng")
    tree.pack(fill="both", expand=True, padx=10)
    def cap_nhat_dialog(event = None):
        selection = tree.selection()    
        selected_item = selection[0]
        cur_vals = list(tree.item(selected_item, "values"))
        gia_tri_can_dien = cur_vals[3]
        ent_kl_input.delete(0, "end")
        ent_kl_input.insert(0, gia_tri_can_dien)
        
    tree.bind("<<TreeviewSelect>>", cap_nhat_dialog)
    for cv in danh_muc:
        tree.insert("", "end", values=(cv["cong_viec"], cv["don_vi"], f"{int(float(cv['don_gia'])):,}", "0"))    
    # Nếu đang ở chế độ sửa, đổ dữ liệu vào
    # Nếu đang ở chế độ sửa, cập nhật khối lượng vào các dòng đã có
    if data:
        ent_start.insert(0, data["bat_dau"])
        ent_duration.insert(0, str(data["duration"]))
        combo_unit.set(data["time_unit"])
        
        # Tạo một dictionary để tra cứu nhanh công việc cũ theo tên
        dict_data_cu = {item["cong_viec"]: item["khoi_luong"] for item in data["danh_sach_chi_tiet"]}
        
        # Duyệt qua các dòng đã insert từ danh mục để điền khối lượng cũ vào
        for item_id in tree.get_children():
            values = list(tree.item(item_id, "values"))
            ten_cv = values[0]
            if ten_cv in dict_data_cu:
                values[3] = str(dict_data_cu[ten_cv]) # Cập nhật lại cột khối lượng (index 3)
                tree.item(item_id, values=values)
    auto_resize_columns(tree)

    
    
    #code để dialog bật ra giữa màn hình
    dialog.update_idletasks()

    # 2. Lấy kích thước màn hình và kích thước cửa sổ
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    width = dialog.winfo_width()
    height = dialog.winfo_height()

    # 3. Tính toán tọa độ x, y để căn giữa
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)

    # 4. Áp dụng vị trí mới
    dialog.geometry(f'+{x}+{y}')

    # Mẹo: Thêm dòng này nếu muốn người dùng không được tương tác với cửa sổ chính
    # cho đến khi đóng dialog (tạo hiệu ứng modal)
    #dialog.grab_set()
    return dialog
def sua_hop_dong(item, file_path):
    # 1. Đọc dữ liệu từ file JSON được chọn
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể đọc file: {e}")
        return

    # 2. Gọi hàm chung ở chế độ sửa
    # Lấy team_name từ đường dẫn (ví dụ: Data/TeamA/...)
    team_name = os.path.basename(os.path.dirname(file_path))
    dialog_window = mo_dialog_hop_dong(team_name, data=data, file_path=file_path)
    globalconfig.root.wait_window(dialog_window)
    lam_moi_tree_hd()
def lam_moi_tree_todoi():
    # Kiểm tra xem tree_todoi đã được khởi tạo hay chưa
    if not hasattr(globalconfig, 'tree_todoi') or globalconfig.tree_todoi is None:
        print("Treeview tổ đội chưa được khởi tạo.")
        return

    tree = globalconfig.tree_todoi

    # 1. Xóa sạch dữ liệu cũ trên Treeview
    for item in tree.get_children():
        tree.delete(item)

    # 2. Lấy đường dẫn file JSON từ hàm get_DSNC_JSON()
    json_path = get_DSNC_JSON()

    # 3. Đọc dữ liệu từ file JSON và nạp lại vào Treeview
    try:
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Duyệt qua từng tổ đội trong cấu trúc Dictionary
            for to_doi, team_info in data.items():
                workmen_list = team_info.get("workmen", [])

                # Tìm tên tổ trưởng dựa vào trường "Công việc"
                to_truong_name = "Chưa xác định"
                for worker in workmen_list:
                    cong_viec = str(worker.get("Công việc", ""))
                    if "tổ trưởng" in cong_viec.strip().lower():
                        to_truong_name = worker.get("Họ và tên", "Chưa xác định")
                        break

                # Tạo node cha cho Tổ đội (vẫn hiển thị bình thường dù workmen_list rỗng)
                node_cha = tree.insert("", "end", text=f"{to_doi} - Tổ trưởng: {to_truong_name}", open=True)

                # Chèn danh sách nhân công vào node con (nếu có)
                for worker in workmen_list:
                    tree.insert(node_cha, "end", values=(
                        clean_cell(worker.get('Họ và tên', '')), 
                        clean_cell(worker.get('MST', '')), 
                        clean_cell(worker.get('CMT/CCCD/Hộ chiếu', ''))
                    ))
        else:
            print(f"Không tìm thấy file dữ liệu tại: {json_path}")

    except Exception as e:
        print(f"Lỗi khi cập nhật tree_todoi: {e}")
    #show dialog with text để nhập tên tổ đội
    #save tên tổ vào file Data/{ketoan_nhancongthuengoai}/danh_sach_to_doi.JSON
    #kiểm tra trùng lên trong danh_sach_to_doi.JSON , nếu không trùng thì lưu , trùng thì hiển thị thông báo và thoát
def add_new_team(parent):
    # 1. Cấu hình đường dẫn
    folder_path = f"Data/{globalconfig.ketoan_nhancongthuengoai}"
    file_path = get_path(os.path.join(folder_path, "danh_sach_to_doi.JSON"))
    
    if not os.path.exists(get_path(folder_path)):
        os.makedirs(get_path(folder_path))

    # 2. Xử lý logic lưu dữ liệu
    def save_team():
        new_name = entry_team.get().strip()
        if not new_name: return

        # Đọc dữ liệu cũ (đọc thành dict thay vì list)
        data = {} 
        if os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            except:
                data = {}

        # Kiểm tra trùng key
        if new_name in data:
            messagebox.showerror("Lỗi", f"Tổ đội '{new_name}' đã tồn tại!")
            return

        # Khởi tạo node con (workmen) cho team này
        data[new_name] = {"workmen": []}
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        
        dialog.destroy()
        #cập nhật tree_todoi        
        lam_moi_tree_todoi()
    # 3. Tạo Dialog Modal
    dialog = tk.Toplevel(parent) # Gán parent để làm modal
    dialog.title("Khai báo tên tổ đội")
    dialog.geometry("350x200")
    
    # Biến cửa sổ thành Modal
    dialog.transient(parent) 
    dialog.grab_set() 

    # Khung chứa nội dung (Padding cho thoáng)
    frame = tk.Frame(dialog, padx=20, pady=20)
    frame.pack(fill="both", expand=True)

    # Title
    tk.Label(frame, text="Khai báo tên tổ đội", font=("Arial", 14, "bold")).pack(pady=(0, 20))

    # Row Tên tổ đội
    row_frame = tk.Frame(frame)
    row_frame.pack(fill="x", pady=5)
    
    tk.Label(row_frame, text="Tên tổ đội", font=("Arial", 10)).pack(side="left", padx=(0, 10))
    entry_team = tk.Entry(row_frame, width=25, font=("Arial", 10))
    entry_team.pack(side="left", expand=True, fill="x")
    entry_team.focus()

    # Nút Tạo mới
    btn_create = tk.Button(frame, text="Tạo mới tổ đội", command=save_team, 
                           bg="#5b9bd5", fg="white", font=("Arial", 10, "bold"), 
                           height=2, padx=10)
    btn_create.pack(pady=20)

    # Chờ dialog đóng
    parent.wait_window(dialog)
from datetime import datetime

def add_new_workman(parent, teamname, workman_data=None):
    # Xác định chế độ: Thêm mới hay Chỉnh sửa
    is_edit = workman_data is not None

    # 1. Cấu hình đường dẫn
    folder_path = get_path(f"Data/{globalconfig.ketoan_nhancongthuengoai}")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    file_path = get_DSNC_JSON()

    # 2. Tạo cửa sổ Modal
    dialog = tk.Toplevel(parent)
    title_text = f"Sửa thông tin nhân công: {workman_data.get('Họ và tên', '')}" if is_edit else f"Thêm nhân công vào tổ: {teamname}"
    dialog.title(title_text)
    dialog.geometry("700x450")
    dialog.grab_set() 
    dialog.transient(parent)

    # Danh sách các trường cần nhập
    field_names = [
        "Họ và tên", "Ngày sinh", "Giới tính", "MST", 
        "CMT/CCCD/Hộ chiếu", "Ngày cấp", "Nơi cấp", 
        "Số TK", "Tên ngân hàng", "Tên chủ TK", 
        "Địa chỉ cư trú", "Nơi làm việc chính", "Công việc"
    ]
    
    entries = {}

    # 3. Tạo khung nhập liệu
    main_frame = ttk.Frame(dialog, padding="20")
    main_frame.pack(fill="both", expand=True)

    header_text = f"SỬA THÔNG TIN NHÂN CÔNG - {teamname}" if is_edit else f"THÊM NHÂN CÔNG - {teamname}"
    ttk.Label(main_frame, text=header_text, font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=(0, 20))

    # Vòng lặp tạo label và widget nhập liệu (Bố cục 2 cột)
    for idx, field in enumerate(field_names):
        row = (idx // 2) + 1
        col = (idx % 2) * 2
        
        ttk.Label(main_frame, text=field).grid(row=row, column=col, sticky="w", pady=5, padx=5)
        # Nếu là trường "Giới tính", sử dụng Combobox với lựa chọn Nam/Nữ
        if field == "Giới tính":
            entry = ttk.Combobox(main_frame, width=28, values=["Nam", "Nữ"], state="readonly")
            if is_edit and field in workman_data:
                entry.set(str(workman_data[field]))
            else:
                entry.set("Nam") # Mặc định là Nam khi thêm mới
        # Nếu là trường "Công việc", sử dụng Combobox với 2 lựa chọn
        elif field == "Công việc":
            entry = ttk.Combobox(main_frame, width=28, values=["Tổ trưởng", "Nhân công"], state="readonly")
            if is_edit and field in workman_data:
                entry.set(str(workman_data[field]))
            else:
                entry.set("Nhân công") # Giá trị mặc định khi thêm mới
        else:
            entry = ttk.Entry(main_frame, width=30)
            if is_edit and field in workman_data:
                entry.insert(0, str(workman_data[field]))
      
            
        entry.grid(row=row, column=col+1, pady=5, padx=5)
        entries[field] = entry

    # Đọc dữ liệu từ file JSON
    data = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except:
            data = {}

    # Đảm bảo tổ đội tồn tại trong data
    if teamname not in data:
        data[teamname] = {"workmen": []}

    # Xử lý lưu dữ liệu (Thêm mới hoặc Cập nhật)
    def save_workman():
        new_data = {field: entries[field].get().strip() for field in entries}
        
        # 1. Kiểm tra không được để trống bất kỳ trường nào
        for field in field_names:
            if not new_data[field]:
                messagebox.showerror("Lỗi", f"Trường '{field}' không được để trống!")
                entries[field].focus()
                return

        # 2. Kiểm tra định dạng ngày tháng hợp lệ (dd/mm/yyyy)
        date_fields = ["Ngày sinh", "Ngày cấp"]
        for field in date_fields:
            val = new_data[field]
            try:
                datetime.strptime(val, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Lỗi định dạng", f"Trường '{field}' không đúng định dạng hoặc ngày không tồn tại!\nVui lòng nhập theo định dạng: dd/mm/yyyy (Ví dụ: 25/12/1990)")
                entries[field].focus()
                return

        current_workmen = data[teamname]["workmen"]

        if is_edit:
            # Chế độ Sửa: Tìm nhân công cũ dựa vào STT hoặc CCCD/MST và cập nhật
            updated = False
            for i, worker in enumerate(current_workmen):
                if (worker.get("STT") == workman_data.get("STT")) or \
                   (worker.get("CMT/CCCD/Hộ chiếu") == workman_data.get("CMT/CCCD/Hộ chiếu") and worker.get("Họ và tên") == workman_data.get("Họ và tên")):
                    
                    new_data["STT"] = workman_data.get("STT", i + 1)
                    current_workmen[i] = new_data
                    updated = True
                    break
            
            if not updated:
                new_data["STT"] = len(current_workmen) + 1
                current_workmen.append(new_data)
                
            success_msg = f"Đã cập nhật thông tin nhân công trong {teamname}"
        else:
            # Chế độ Thêm mới: Tự động gán STT và append vào danh sách
            new_data["STT"] = len(current_workmen) + 1
            current_workmen.append(new_data)
            success_msg = f"Đã thêm nhân công vào {teamname}"

        # Ghi đè lại file JSON
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể ghi file dữ liệu: {ex}")
            return

        messagebox.showinfo("Thành công", success_msg)
        dialog.destroy()
        
        # Tự động làm mới lại Treeview sau khi lưu thành công
        if 'lam_moi_tree_todoi' in globals():
            lam_moi_tree_todoi()

    # Nút Lưu
    btn_save = ttk.Button(main_frame, text="Lưu thông tin", command=save_workman)
    btn_save.grid(row=(len(field_names) // 2) + 2, column=0, columnspan=4, pady=20)

    parent.wait_window(dialog)
def sua_thanh_vien_chon():
    # 1. Kiểm tra Treeview đã khởi tạo chưa
    tree = getattr(globalconfig, 'tree_todoi', None)
    if not tree:
        messagebox.showwarning("Cảnh báo", "Bảng dữ liệu tổ đội chưa được khởi tạo!")
        return
        
    # 2. Lấy dòng đang được chọn trên giao diện
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn một thành viên cần sửa thông tin!")
        return
        
    item_id = selected_items[0]
    parent_id = tree.parent(item_id)
    
    # 3. Kiểm tra xem người dùng có chọn nhầm node cha (tên tổ đội) không
    if not parent_id:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn một thành viên cụ thể bên trong tổ đội, không chọn tên tổ đội!")
        return
        
    # 4. Tách lấy tên tổ đội từ node cha
    parent_text = tree.item(parent_id, "text")
    teamname = parent_text.split(" - Tổ trưởng:")[0].strip()
    
    # 5. Lấy các giá trị đang hiển thị trên dòng (HoTen, MST, CCCD)
    values = tree.item(item_id, "values")
    if not values or len(values) < 3:
        messagebox.showerror("Lỗi", "Không đọc được dữ liệu của dòng được chọn!")
        return
        
    ho_ten_chon, mst_chon, cccd_chon = values[0], values[1], values[2]
    
    # 6. Đọc tệp JSON chứa danh sách nhân công
    json_path = get_DSNC_JSON()
    if not os.path.exists(json_path):
        messagebox.showerror("Lỗi", "Không tìm thấy tệp dữ liệu nhân công!")
        return
        
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không đọc được tệp JSON: {e}")
        return
        
    # 7. Tìm đúng bản ghi thông tin chi tiết của nhân công đó trong tổ đội tương ứng
    workman_data = None
    if teamname in data and "workmen" in data[teamname]:
        for worker in data[teamname]["workmen"]:
            # Khớp điều kiện dựa vào CCCD và Họ tên
            if worker.get("CMT/CCCD/Hộ chiếu") == cccd_chon and worker.get("Họ và tên") == ho_ten_chon:
                workman_data = worker
                break
                
    # 8. Gọi hàm add_new_workman ở chế độ Sửa (is_edit = True)
    if workman_data:
        add_new_workman(globalconfig.root, teamname, workman_data=workman_data)
    else:
        messagebox.showerror("Lỗi", "Không tìm thấy dữ liệu chi tiết của thành viên này trong hệ thống!")    

def nhap_thong_tin_du_an(parent):
    # 1. Cấu hình đường dẫn lưu file JSON
    folder_path = get_path(f"Data/{globalconfig.ketoan_nhancongthuengoai}")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    file_path = get_path(os.path.join(folder_path, f"{globalconfig.file_thong_tin_du_an}.JSON"))

    # 2. Tạo cửa sổ Modal
    dialog = tk.Toplevel(parent)
    dialog.title("Khai báo thông tin dự án")
    dialog.geometry("650x550")
    dialog.grab_set()
    dialog.transient(parent)

    # Danh sách các trường thông tin theo bảng
    field_names = [
        "Tên công ty", 
        "Tên công trình", 
        "Giám đốc", 
        "Chỉ huy trưởng", 
        "Kế toán trưởng", 
        "Người lập biểu", 
        "Địa điểm", 
        "Năm thực hiện",
        "Thu nhập năm miễn thuế"
    ]
    
    entries = {}

    # 3. Tạo khung giao diện nhập liệu
    main_frame = ttk.Frame(dialog, padding="20")
    main_frame.pack(fill="both", expand=True)

    ttk.Label(main_frame, text="THÔNG TIN CHUNG DỰ ÁN", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 20))

    # Đọc dữ liệu cũ từ file JSON nếu có (dùng .get() để chống lỗi thiếu trường dữ liệu)
    old_data = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                old_data = json.load(f)
        except Exception:
            old_data = {}

    # Vòng lặp tạo Label và Entry theo dạng 1 cột dọc trực quan, dễ điền
    for idx, field in enumerate(field_names):
        row = idx + 1
        
        ttk.Label(main_frame, text=field, font=("Arial", 10)).grid(row=row, column=0, sticky="w", pady=6, padx=5)
        entry = ttk.Entry(main_frame, width=45, font=("Arial", 10))
        
        # Điền dữ liệu cũ an toàn, nếu thiếu trường sẽ không báo lỗi mà để trống
        val = old_data.get(field, "")
        if val:
            entry.insert(0, str(val))
            
        entry.grid(row=row, column=1, pady=6, padx=5, sticky="ew")
        entries[field] = entry

    # Cấu hình co giãn cột trong khung
    main_frame.columnconfigure(1, weight=1)

    # 4. Hàm xử lý lưu thông tin vào JSON
    def save_project_info():
        # Lấy dữ liệu và loại bỏ khoảng trắng thừa
        new_data = {field: entries[field].get().strip() for field in field_names}
        
        # Kiểm tra không được để trống các trường bắt buộc
        if not new_data["Tên công ty"] or not new_data["Tên công trình"]:
            messagebox.showerror("Lỗi", "Tên công ty và Tên công trình không được để trống!", parent=dialog)
            return

        # Kiểm tra năm thực hiện phải là số (nếu có nhập)
        nam_thuc_hien = new_data["Năm thực hiện"]
        if nam_thuc_hien and not nam_thuc_hien.isdigit():
            messagebox.showerror("Lỗi định dạng", "Trường 'Năm thực hiện' phải là định dạng số (Ví dụ: 2026)!", parent=dialog)
            entries["Năm thực hiện"].focus()
            return

        # Ghi dữ liệu vào file JSON
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(new_data, f, ensure_ascii=False, indent=4)
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể ghi file dữ liệu: {ex}", parent=dialog)
            return

        messagebox.showinfo("Thành công", "Đã lưu thông tin dự án thành công!", parent=dialog)
        dialog.destroy()

    # Nút Lưu thông tin
    btn_save = ttk.Button(main_frame, text="Lưu thông tin", command=save_project_info)
    btn_save.grid(row=len(field_names) + 1, column=0, columnspan=2, pady=25)

    parent.wait_window(dialog)    